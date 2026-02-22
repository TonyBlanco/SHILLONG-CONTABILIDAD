# -*- coding: utf-8 -*-
"""
models/ExportadorExcelMensual.py — MOTOR DE EXPORTACIÓN v3.7.2
-----------------------------------------------------------------
Genera reportes Excel profesionales con:
- Totales calculados (Debe, Haber y SALDO NETO).
- Formato de moneda.
- Estilos (Negritas, colores).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
import datetime


class ExportadorExcelMensual:
    
    @staticmethod
    def _aplicar_estilo_sisters(ws, titulo_reporte, subtitulo="", banco=""):
        """
        Aplica el diseño profesional de la imagen:
        - Logo de Sisters Hospitallers
        - Cabecera morada (#7030A0)
        - CDAD SHILLONG
        """
        # 1. Logo (Esquinas superiores)
        logo_path = Path("assets/logo/logo hospitaller.jpg")
        if logo_path.exists():
            try:
                img = Image(str(logo_path))
                img.width = 110
                img.height = 45
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"[Excel] No se pudo cargar logo: {e}")

        # Colores
        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        morado_claro = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        blanco = Font(color="FFFFFF", bold=True)
        centro = Alignment(horizontal="center", vertical="center")
        borde_fino = Side(style="thin", color="000000")
        border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

        # 2. Encabezado Derecho (Merged B1:F2)
        # Título principal morado
        ws.merge_cells("D1:F2")
        cell_titulo = ws["D1"]
        cell_titulo.value = "CDAD SHILLONG"
        cell_titulo.fill = morado
        cell_titulo.font = Font(color="FFFFFF", bold=True, size=14)
        cell_titulo.alignment = centro

        # Fecha y reporte
        ws["D3"].value = f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y')}"
        ws["D3"].font = Font(bold=True, size=10)
        ws["D3"].alignment = Alignment(horizontal="center")
        ws.merge_cells("D3:F3")

        # Tipo de Reporte (Libro Diario / Banco)
        if titulo_reporte:
            ws.merge_cells("A5:C6")
            cell_rep = ws["A5"]
            cell_rep.value = titulo_reporte.upper()
            cell_rep.fill = morado
            cell_rep.font = blanco
            cell_rep.alignment = centro
            
            ws.merge_cells("D5:F6")
            cell_sub = ws["D5"]
            cell_sub.value = subtitulo.upper() if subtitulo else "CDAD SHILLONG"
            cell_sub.fill = morado
            cell_sub.font = blanco
            cell_sub.alignment = centro

        return 8 # Fila donde empiezan los headers

    @staticmethod
    def _añadir_pie_sisters(ws, start_row, total_ingresos, total_gastos, saldo_inicial, banco=""):
        """Tabla de resumen final y cuadros de firma."""
        row = start_row + 2
        
        # Colores
        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cyan = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        borde_fino = Side(style="thin", color="000000")
        border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        
        # 1. RESUMEN (Derecha)
        # Headers del resumen
        ws.cell(row=row, column=4, value="RESUMEN").font = Font(bold=True)
        ws.cell(row=row, column=5, value="CONCEPTOS").fill = morado
        ws.cell(row=row, column=5).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        row += 1
        
        # Filas
        conceptos = [
            ("Total Ingresos", total_ingresos),
            ("Total Gastos", total_gastos),
            ("SALDO ACTUAL", total_ingresos - total_gastos)
        ]
        
        for text, val in conceptos:
            ws.cell(row=row, column=5, value=text).border = border
            ws.cell(row=row, column=6, value=val).border = border
            ws.cell(row=row, column=6).number_format = "#,##0.00"
            if "SALDO" in text:
                ws.cell(row=row, column=5).font = Font(bold=True)
                ws.cell(row=row, column=6).font = Font(bold=True)
                ws.cell(row=row, column=5).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row += 1
            
        # Check CORRECTO
        ws.cell(row=row, column=5, value="Check").alignment = Alignment(horizontal="right")
        cell_check = ws.cell(row=row, column=6, value="CORRECTO")
        cell_check.font = Font(color="008000", bold=True)
        cell_check.border = border
        cell_check.alignment = Alignment(horizontal="center")
        row += 2

        # 2. FIRMAS (Izquierda y Derecha)
        firma_row = row
        # Preparado
        ws.merge_cells(start_row=firma_row, start_column=1, end_row=firma_row, end_column=2)
        cell_p = ws.cell(row=firma_row, column=1, value="Preparado")
        cell_p.fill = morado; cell_p.font = Font(color="FFFFFF", bold=True); cell_p.alignment = Alignment(horizontal="center")
        
        ws.merge_cells(start_row=firma_row+1, start_column=1, end_row=firma_row+3, end_column=2)
        for r in range(firma_row+1, firma_row+4):
            for c in range(1, 3): 
                ws.cell(row=r, column=c).border = border
        
        ws.cell(row=firma_row+4, column=1, value="Nombre:").font = Font(size=8)
        ws.cell(row=firma_row+4, column=2).fill = cyan
        ws.cell(row=firma_row+5, column=1, value="Cargo:").font = Font(size=8)
        ws.cell(row=firma_row+5, column=2).fill = cyan

        # Revisado/Autorizado
        ws.merge_cells(start_row=firma_row, start_column=4, end_row=firma_row, end_column=5)
        cell_r = ws.cell(row=firma_row, column=4, value="Revisado/Autorizado")
        cell_r.fill = morado; cell_r.font = Font(color="FFFFFF", bold=True); cell_r.alignment = Alignment(horizontal="center")
        
        ws.merge_cells(start_row=firma_row+1, start_column=4, end_row=firma_row+3, end_column=5)
        for r in range(firma_row+1, firma_row+4):
            for c in range(4, 6): 
                ws.cell(row=r, column=c).border = border
        
        ws.cell(row=firma_row+4, column=4, value="Nombre:").font = Font(size=8)
        ws.cell(row=firma_row+4, column=5).fill = cyan
        ws.cell(row=firma_row+5, column=4, value="Cargo:").font = Font(size=8)
        ws.cell(row=firma_row+5, column=5).fill = cyan

    @staticmethod
    def _formato_moneda(ws, row, cols_indices):
        """Aplica formato #,##0.00 a las celdas indicadas."""
        for col in cols_indices:
            cell = ws.cell(row=row, column=col)
            cell.number_format = "#,##0.00"

    # ============================================================
    #  EXPORTACIÓN GENERAL — FORMATO LIBRO TEST
    # ============================================================
    @staticmethod
    def exportar_general(ruta_archivo, datos, periodo_str):
        """Exportación modelo SISTERS HOSPITALERS."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Diario"

        # 1. Aplicar Estilo Sisters
        start_row = ExportadorExcelMensual._aplicar_estilo_sisters(ws, "Libro Diario", periodo_str)
        
        # 2. Saldos Iniciales (Fila destacada)
        saldo_inicial = 0.0
        for m in datos:
            if "inicial" in str(m.get("concepto", "")).lower():
                saldo_inicial = float(m.get("saldo", 0) or 0)
                break
        
        ws.cell(row=start_row, column=4, value="SALDO INICIAL").font = Font(bold=True)
        cell_si = ws.cell(row=start_row, column=6, value=saldo_inicial)
        cell_si.font = Font(bold=True)
        cell_si.number_format = "#,##0.00"
        cell_si.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarillo
        ws.cell(row=start_row, column=5).border = ws.cell(row=start_row, column=6).border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
        row_idx = start_row + 1

        # 3. Headers Tabla
        headers = ["FECHA", "CONCEPTO", "CUENTA", "INGRESOS", "GASTOS", "SALDO"]
        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        borde_fino = Side(style="thin", color="000000")
        border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.fill = morado; cell.font = Font(color="FFFFFF", bold=True); cell.border = border; cell.alignment = Alignment(horizontal="center")
            # Anchos
            letra = get_column_letter(c)
            if h == "CONCEPTO": ws.column_dimensions[letra].width = 45
            elif h == "FECHA": ws.column_dimensions[letra].width = 12
            else: ws.column_dimensions[letra].width = 15
        row_idx += 1

        total_ing = 0.0
        total_gas = 0.0
        
        # 4. Datos
        for m in datos:
            if "inicial" in str(m.get("concepto", "")).lower(): continue
            
            # Priorizar keys explícitas, fallback a debe/haber estándar
            ing = float(m.get("ingresos") if "ingresos" in m else m.get("haber", 0) or 0)
            gas = float(m.get("gastos") if "gastos" in m else m.get("debe", 0) or 0)
            total_ing += ing
            total_gas += gas
            
            row = [
                m.get("fecha"),
                m.get("concepto"),
                str(m.get("cuenta", "")),
                ing,
                gas,
                m.get("saldo")
            ]
            
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                cell.border = border
                if c >= 4:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        # Resumen Final y Firmas
        ExportadorExcelMensual._añadir_pie_sisters(ws, row_idx, total_ing, total_gas, saldo_inicial)

        try:
            wb.save(ruta_archivo)
            return True
        except Exception as e:
            raise e

    # ============================================================
    #  EXPORTACIÓN AGRUPADA (NO SE TOCA)
    # ============================================================
    @staticmethod
    def exportar_agrupado(ruta_archivo, grupos_data, periodo_str, titulo_agrupacion):
        """Exportación agrupada (Banco / Cuenta) con diseño SISTERS."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Agrupado"

        # 1. Cabecera Sisters
        report_title = "Libro Diario Banco" if "Banco" in titulo_agrupacion else f"Reporte por {titulo_agrupacion}"
        row_idx = ExportadorExcelMensual._aplicar_estilo_sisters(ws, report_title, periodo_str)
        
        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        morado_claro = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        borde_fino = Side(style="thin", color="000000")
        border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        
        total_gral_ing = 0.0
        total_gral_gas = 0.0

        for nombre_grupo, movimientos in grupos_data.items():
            # --------------------------------------------------------
            # ESTILO BLOQUE POR BANCO (Según Imagen)
            # --------------------------------------------------------
            
            # HEADER BANCO: [SALDO INICIAL {N}] ... [BANCO (Cyan)]
            # Fila: MORADO COMPLETO
            # A: "SALDO INICIAL {N}" (Amarillo/Blanco)
            # B-C: "CONCEPTOS - {BANCO}" (Morado/Cyan)
            
            # 1. Título "SALDO INICIAL" (Col A)
            cell_si_lbl = ws.cell(row=row_idx, column=1, value=f"SALDO INICIAL")
            cell_si_lbl.font = Font(color="FFFFFF", bold=True, size=9)
            cell_si_lbl.fill = morado
            cell_si_lbl.alignment = Alignment(horizontal="center", vertical="center")
            cell_si_lbl.border = border

            # 2. Header "FECHA" (Col B)
            cell_fecha = ws.cell(row=row_idx, column=2, value="FECHA")
            cell_fecha.font = Font(color="FFFFFF", bold=True)
            cell_fecha.fill = morado
            cell_fecha.alignment = Alignment(horizontal="center", vertical="center")
            cell_fecha.border = border

            # 3. Header "CONCEPTO - {BANCO}" (Col C)
            # Solo columna C, sin merge con B.
            cell_banco = ws.cell(row=row_idx, column=3, value=f"CONCEPTO - {str(nombre_grupo).upper()}")
            cell_banco.font = Font(color="00FFFF", bold=True) # Cyan
            cell_banco.fill = morado
            cell_banco.alignment = Alignment(horizontal="center", vertical="center")
            cell_banco.border = border
            
            # 4. Headers Columnas (D, E, F) -> INGRESOS, GASTOS, SALDO
            headers_right = ["INGRESOS", "GASTOS", "SALDO"]
            for i, h in enumerate(headers_right):
                cell = ws.cell(row=row_idx, column=4+i, value=h)
                cell.fill = morado
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            row_idx += 1

            # --------------------------------------------------------
            # FILA DE SALDO INICIAL (Amarillo)
            # --------------------------------------------------------
            saldo_ini_grupo = 0.0
            for m in movimientos:
                if "inicial" in str(m.get("concepto", "")).lower():
                    saldo_ini_grupo = float(m.get("saldo", 0) or 0)
                    break
            
            # Col A: Fondo Amarillo vacío o con valor
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws.cell(row=row_idx, column=1).border = border
            
            # Col F (Saldo): Valor Saldo Inicial
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5) # Espacio vacio
            for c in range(2, 6): ws.cell(row=row_idx, column=c).border = border

            cell_saldo_ini = ws.cell(row=row_idx, column=6, value=saldo_ini_grupo)
            cell_saldo_ini.number_format = "#,##0.00"
            cell_saldo_ini.font = Font(bold=True)
            cell_saldo_ini.border = border
            
            row_idx += 1

            # --------------------------------------------------------
            # TABLA DE MOVIMIENTOS
            # --------------------------------------------------------
            # Headers Tabla (Fecha, Concepto...)
            # Según imagen: NO HAY HEADERS REPETIDOS AQUÍ, SOLO DATOS
            # Headers globales ya están arriba? NO, la imagen tiene headers en la fila del banco.
            # REVISION IMAGEN:
            # Fila 1 (Morada): [SALDO INICIAL 1] [FECHA] [CONCEPTO - BANCO] [INGRESOS] [GASTOS] [SALDO]
            # Fila 2 (Datos): [Amarillo] [Fecha] [Concepto] ...
            
            # CORRECCION DISEÑO SEGUN IMAGEN (Re-interpretación):
            # La imagen muestra bloques repetitivos.
            # Row 1 (Morada): Col1="SALDO INICIAL 1", Col2="FECHA", Col3="CONCEPTO - BANCO", Col4="INGRESOS", Col5="GASTOS", Col6="SALDO"
            # Row 2 (Amarillo inicia): Col1 (Amarillo), Col2..Col6 (Datos Saldo Inicial)
            

            sub_ing = 0.0
            sub_gas = 0.0
            
            for m in movimientos:
                if "inicial" in str(m.get("concepto", "")).lower(): continue
                
                ing = float(m.get("ingresos") if "ingresos" in m else m.get("haber", 0) or 0)
                gas = float(m.get("gastos") if "gastos" in m else m.get("debe", 0) or 0)
                sub_ing += ing; sub_gas += gas
                
                # Columnas: A(Vacío), B(Fecha), C(Concepto), D(Ing), E(Gas), F(Saldo)
                ws.cell(row=row_idx, column=1).border = border # Col A vacia borde
                ws.cell(row=row_idx, column=2, value=m.get("fecha")).border = border
                ws.cell(row=row_idx, column=3, value=m.get("concepto")).border = border
                ws.cell(row=row_idx, column=4, value=ing).border = border
                ws.cell(row=row_idx, column=5, value=gas).border = border
                ws.cell(row=row_idx, column=6, value=m.get("saldo")).border = border
                
                ExportadorExcelMensual._formato_moneda(ws, row_idx, [4, 5, 6])
                row_idx += 1

            # Separador visual entre bancos
            row_idx += 2
            
            total_gral_ing += sub_ing
            total_gral_gas += sub_gas

        # Pie final de reporte
        ExportadorExcelMensual._añadir_pie_sisters(ws, row_idx, total_gral_ing, total_gral_gas, 0)

        try:
            wb.save(ruta_archivo)
            return True
        except Exception as e:
            raise e
