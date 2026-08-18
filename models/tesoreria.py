# -*- coding: utf-8 -*-
"""
tesoreria.py — SHILLONG CONTABILIDAD v3.8.0 PRO
Lógica del RESUMEN DE TESORERÍA (saldo acumulado por mes) + exportación Excel.
Sin dependencias Qt: lo comparten InformesView (Informes BI) y TesoreriaView.
"""

import json
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MESES_EXCEL = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def obtener_bancos_config():
    """Bancos/tesorería desde data/bancos.json (Caja + bancos + Cambio Euros + Contrapartida)."""
    try:
        with open("data/bancos.json", "r", encoding="utf-8") as f:
            return [b["nombre"] for b in json.load(f).get("banks", [])]
    except (IOError, json.JSONDecodeError, KeyError):
        return ["Caja", "Cambio Euros", "Contrapartida"]


def cargar_saldos_iniciales_año(año):
    """
    Saldo inicial real de cada tesorería para el año: usa el PRIMER mes del año
    que tenga saldo registrado en saldos_mensuales.json (no asume enero).
    Cada banco toma su inicial del primer mes en el que aparece.
    """
    try:
        with open("data/saldos_mensuales.json", "r", encoding="utf-8") as f:
            data = json.load(f)
    except (IOError, json.JSONDecodeError, AttributeError):
        return {}
    saldos = data.get("saldos", {})
    inicial = {}
    for mes in range(1, 13):
        mes_data = saldos.get(f"{año}-{mes:02d}")
        if not isinstance(mes_data, dict):
            continue
        for banco, valores in mes_data.items():
            if banco in inicial or not isinstance(valores, dict):
                continue
            inicial[banco] = float(valores.get("inicial", 0.0) or 0.0)
    return inicial


def _resolver_cuentas_tesoreria(movimientos):
    """
    Devuelve (etiqueta_a_cb, cb_etiquetas):
    - etiqueta_a_cb: mapa etiqueta-banco -> cuenta_banco real (570, 5721, …)
      cuando el movimiento trae el campo cuenta_banco. Así los movimientos que
      NO lo traen (p.ej. meses posteriores) se agrupan en la misma cuenta real.
    - cb_etiquetas: cuenta_banco -> {etiqueta: n} para elegir el nombre visible
      (la etiqueta más frecuente de cada cuenta).
    """
    etiqueta_a_cb = {}
    cb_etiquetas = defaultdict(lambda: defaultdict(int))
    for m in movimientos:
        cb = str(m.get("cuenta_banco", "") or "").strip()
        b = str(m.get("banco", "") or "Caja").strip()
        if cb:
            etiqueta_a_cb.setdefault(b, cb)
            cb_etiquetas[cb][b] += 1
    return etiqueta_a_cb, cb_etiquetas


def _clave_tesoreria(m, etiqueta_a_cb):
    """
    Clave real de tesorería de un movimiento.
    Prioriza la etiqueta del banco (mapeada a su cuenta real): en los datos reales
    el campo cuenta_banco de meses posteriores a veces quedó con 570 aunque el
    banco sea otro (p.ej. 'Post- office sr Sindhu' con cuenta_banco=570).
    Solo usa el cuenta_banco propio si la etiqueta es desconocida.
    """
    cb = str(m.get("cuenta_banco", "") or "").strip()
    b = str(m.get("banco", "") or "Caja").strip()
    cb_resuelto = etiqueta_a_cb.get(b)
    if cb_resuelto:
        return "CB:" + cb_resuelto
    if cb:
        return "CB:" + cb
    return "N:" + b


def _nombre_tesoreria(clave, cb_etiquetas):
    """Nombre visible de una clave: Caja para 570, si no la etiqueta banco más común."""
    if clave.startswith("CB:"):
        codigo = clave[3:]
        if codigo == "570":
            return "Caja"
        cont = cb_etiquetas.get(codigo, {})
        if cont:
            return max(cont, key=cont.get)
        return "Cuenta " + codigo
    return clave[2:]


def datos_tesoreria(movimientos, año):
    """
    Devuelve (bancos, acum):
    - bancos: lista ordenada de tesorerías REALES sin duplicados
      (Caja primero, luego cuentas con movimiento, luego Cambio Euros/Contrapartida
      y bancos de configuración con saldo inicial).
    - acum[banco] = lista de 12 saldos acumulados a fin de cada mes:
      saldo(m) = saldo_inicial + Σ (haber − debe) de movimientos PAGADOS
      con fecha ≤ fin del mes m del año indicado.

    Agrupa por la cuenta de tesorería REAL (cuenta_banco: 570, 5721…5742) en lugar
    del nombre libre del banco, para que un mismo banco/cuenta no aparezca repetido
    con nombres distintos (p.ej. "Union Bank" y "Union Bank, sr Elisa" = 5725).
    """
    bancos_config = obtener_bancos_config()
    saldos_iniciales = cargar_saldos_iniciales_año(año)
    etiqueta_a_cb, cb_etiquetas = _resolver_cuentas_tesoreria(movimientos)

    # Acumular por clave real (CB:<cuenta_banco> o N:<banco>)
    acum = defaultdict(lambda: [0.0] * 12)
    ultimo_mes = 0  # último mes del año con movimientos reales (corte del informe)
    for m in movimientos:
        f_raw = str(m.get("fecha", ""))
        try:
            if "/" in f_raw:
                d, mm, a = map(int, f_raw.split("/"))
            elif "-" in f_raw:
                p = f_raw.split("-")
                if int(p[0]) > 1000:
                    a, mm, d = int(p[0]), int(p[1]), int(p[2])
                else:
                    d, mm, a = int(p[0]), int(p[1]), int(p[2])
            else:
                continue
        except (ValueError, IndexError):
            continue

        if a != año or not (1 <= mm <= 12):
            continue

        try:
            h = float(str(m.get("haber", 0)).replace(",", "."))
            dv = float(str(m.get("debe", 0)).replace(",", "."))
        except (ValueError, TypeError):
            continue

        if str(m.get("estado", "")).lower() != "pagado":
            continue

        neto = h - dv
        if not neto:
            continue

        if mm > ultimo_mes:
            ultimo_mes = mm

        clave = _clave_tesoreria(m, etiqueta_a_cb)
        # Acumulado: suma a todos los meses desde el mes del movimiento
        for idx in range(mm - 1, 12):
            acum[clave][idx] += neto

    def cortar(serie):
        """Deja None en los meses posteriores al último con movimientos reales.
        Así el informe no muestra meses 'planos' inventados (p.ej. hasta diciembre
        cuando la data real termina en julio)."""
        if ultimo_mes <= 0:
            return [None] * 12
        return serie[:ultimo_mes] + [None] * (12 - ultimo_mes)

    # Nombre visible y saldo inicial por clave
    nombres = {clave: _nombre_tesoreria(clave, cb_etiquetas) for clave in acum}

    def inicial_de(clave):
        nombre = nombres[clave]
        if nombre in saldos_iniciales:
            return float(saldos_iniciales[nombre])
        if clave.startswith("CB:"):
            for b in cb_etiquetas.get(clave[3:], {}):
                if b in saldos_iniciales:
                    return float(saldos_iniciales[b])
        return 0.0

    # Fusionar claves con el mismo nombre visible (p.ej. CB:570 y N:Caja)
    por_nombre = {}
    inicial_por_nombre = {}
    for clave, serie in acum.items():
        nombre = nombres[clave]
        if nombre not in por_nombre:
            por_nombre[nombre] = [0.0] * 12
            inicial_por_nombre[nombre] = inicial_de(clave)
        else:
            if not inicial_por_nombre[nombre]:
                inicial_por_nombre[nombre] = inicial_de(clave)
        for i in range(12):
            por_nombre[nombre][i] += serie[i]

    # Construir filas
    filas = []

    # 1) Caja siempre (cuenta real 570 + etiqueta "Caja" + saldo inicial)
    caja_ini = float(saldos_iniciales.get("Caja", 0.0) or 0.0)
    caja_serie = por_nombre.get("Caja", [0.0] * 12)
    filas.append(("Caja", cortar([caja_ini + caja_serie[i] for i in range(12)])))

    # 2) Cuentas reales (CB) con movimientos, ordenadas por código
    for clave in sorted((c for c in acum if c.startswith("CB:") and c != "CB:570"), key=lambda c: c[3:]):
        nombre = nombres[clave]
        if nombre in {f[0] for f in filas}:
            continue
        filas.append((nombre, cortar([inicial_por_nombre.get(nombre, 0.0) + por_nombre[nombre][i] for i in range(12)])))

    # 3) Bancos sin cuenta_banco que no se fusionaron
    usados = {f[0] for f in filas}
    for nombre in sorted(por_nombre):
        if nombre in usados:
            continue
        filas.append((nombre, cortar([inicial_por_nombre.get(nombre, 0.0) + por_nombre[nombre][i] for i in range(12)])))

    # 4) Bancos de configuración sin movimientos: solo si tienen saldo inicial
    #    o son Cambio Euros/Contrapartida (evita filas vacías que confunden).
    usados = {f[0] for f in filas}
    especiales = {"Cambio Euros", "Contrapartida"}
    for b in bancos_config:
        if b == "Caja" or b in usados:
            continue
        ini = float(saldos_iniciales.get(b, 0.0) or 0.0)
        if b in especiales or ini:
            filas.append((b, cortar([ini] * 12)))

    orden = [f[0] for f in filas]
    return orden, {f[0]: f[1] for f in filas}


def exportar_excel_tesoreria(ruta, año, bancos, acum):
    """Exporta el Resumen de Tesorería a un archivo xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tesorería"

    titulo = ws.cell(1, 1)
    titulo.value = f"RESUMEN DE TESORERÍA {año} — SALDO ACUMULADO POR MES"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor="7030A0")

    header_row = 3
    ws.cell(header_row, 1).value = "BANCO / TESORERÍA"
    for c, mes in enumerate(MESES_EXCEL, start=2):
        ws.cell(header_row, c).value = mes
    for c in range(1, 14):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    totales = [None] * 12  # None = mes sin data real (celda vacía)
    r = header_row
    for banco in bancos:
        r += 1
        ws.cell(r, 1).value = banco
        vals = acum[banco]
        for c in range(12):
            v = vals[c]
            if v is None:
                continue  # mes sin movimientos → celda vacía
            if totales[c] is None:
                totales[c] = 0.0
            totales[c] += v
            cell = ws.cell(r, c + 2)
            cell.value = round(v, 2)
            cell.number_format = "#,##0.00"

    r += 1
    ws.cell(r, 1).value = "TOTAL"
    ws.cell(r, 1).font = Font(bold=True)
    for c in range(12):
        v = totales[c]
        if v is None:
            continue  # ningún banco tiene data en ese mes → TOTAL vacío
        cell = ws.cell(r, c + 2)
        cell.value = round(v, 2)
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(ruta)
