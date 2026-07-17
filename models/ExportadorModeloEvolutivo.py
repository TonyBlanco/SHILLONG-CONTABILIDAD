# -*- coding: utf-8 -*-
"""
models/ExportadorModeloEvolutivo.py

Fills the official blank template ("Modelo evolutivo presupuestario 26 - blank.xlsx")
with real data from totales_exactos, preserving the original formatting exactly.

Public API:
  exportar_modelo_evolutivo(ruta_archivo, totales_exactos, cuentas_by_section,
                             anio, periodo_str, nombre_por_cuenta, ruta_plantilla)

Column layout in template (1-based):
  C(3) = account code  D(4) = name  E(5) = MES CORRIENTE  F(6) = BANCO
  G(7) = CAJA  H(8) = MESES ANTERIORES  I(9) = SUMA ACUMULADO
  J(10) = PRESUPUESTO  K(11) = DIFERENCIA
"""

import shutil
import datetime
import re
from pathlib import Path
import openpyxl

# Column indices (1-based, matching template layout)
COL_CUENTA      = 3
COL_NOMBRE      = 4
COL_CORRIENTE   = 5   # MES CORRIENTE  = cur_banco + cur_caja
COL_BANCO       = 6
COL_CAJA        = 7
COL_ANTERIORES  = 8   # MESES ANTERIORES = prev_total
COL_ACUMULADO   = 9   # SUMA ACUMULADO  = corriente + anteriores
COL_PRESUPUESTO = 10
COL_DIFERENCIA  = 11  # PRESUPUESTO - ACUMULADO

TOTAL_LABELS = {"SUMA TOTAL GASTOS", "SUMA TOTAL INGRESOS", "INVERSIONES"}


def _normalizar_codigo(v):
    """Convert any template account code format to 6-digit string, e.g. '629.2' -> '629200'."""
    if v is None:
        return None
    s = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
    s = re.sub(r"\D", "", s)
    if not s.isdigit():
        return None
    return s.ljust(6, "0")[:6]


def _parse_template_code(v):
    """Return normalized code plus interpretation mode for template rows.

    Mode rules:
    - "exact": dotted/comma formats like 602.40 mean the concrete account 602400.
    - "group": dashed formats like 602-400 mean the parent/group row for 6024xx.
    - "auto": fallback for plain numeric codes, preserving the historical rollup logic.
    """
    if v is None:
        return None, "auto"

    raw = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
    codigo = _normalizar_codigo(raw)
    if not codigo:
        return None, "auto"

    if "-" in raw:
        return codigo, "group"
    if "." in raw or "," in raw:
        return codigo, "exact"
    return codigo, "auto"


def _group_prefix(codigo):
    codigo = str(codigo).strip()
    if codigo.endswith("000") and len(codigo) > 3:
        return codigo[:-3]
    if codigo.endswith("00") and len(codigo) > 2:
        return codigo[:-2]
    return codigo


def _has_descendants(codigo, candidate_codes):
    pref = _group_prefix(codigo)
    for candidate in candidate_codes:
        cta = str(candidate).strip()
        if len(cta) != len(str(codigo)):
            continue
        if cta != str(codigo) and cta.startswith(pref):
            return True
    return False


def _resolve_template_mode(raw_code, normalized_code, sheet_codes, candidate_codes):
    raw = str(raw_code or "").strip()
    if not normalized_code:
        return "auto"

    if "-" in raw:
        return "group"

    parts = [p for p in re.split(r"[.,]", raw) if p]
    has_separator = any(sep in raw for sep in (".", ","))

    # A dotted/comma row is exact if the same normalized code also exists as a plain row.
    if has_separator:
        for other_raw, other_norm in sheet_codes:
            if other_norm != normalized_code:
                continue
            other_txt = str(other_raw or "").strip()
            if other_txt == raw:
                continue
            if not any(sep in other_txt for sep in (".", ",", "-")):
                return "exact"

    # Three decimal digits usually map to explicit leaves like 602.401 -> 602401.
    if has_separator and any(len(p) >= 3 for p in parts[1:]):
        return "exact"

    if has_separator:
        return "exact"

    if _has_descendants(normalized_code, candidate_codes):
        return "group"
    return "auto"


def _cell_is_sum_formula(cell):
    value = cell.value
    return isinstance(value, str) and value.strip().upper().startswith("=SUM(")


def _template_row_mode(ws, row, raw_code, normalized_code, sheet_codes, candidate_codes):
    for col in (COL_CORRIENTE, COL_BANCO, COL_CAJA, COL_ANTERIORES, COL_ACUMULADO):
        if _cell_is_sum_formula(ws.cell(row, col)):
            return "group"

    return "exact"


def _rollup(codigo, totales_exactos, mode="auto"):
    """Sum accounts whose 6-char code shares the same meaningful prefix.

    Parent accounts like '600000' (ends in 000) roll up all 600xxx children.
    Leaf accounts like '628030' match exactly.
    """
    codigo = str(codigo).strip()
    if mode == "exact":
        pref = codigo
        exact_match = True
    elif mode == "group":
        exact_match = False
        if codigo.endswith("000") and len(codigo) > 3:
            pref = codigo[:-3]
        elif codigo.endswith("00") and len(codigo) > 2:
            pref = codigo[:-2]
        else:
            pref = codigo
    elif codigo.endswith("000") and len(codigo) > 3:
        pref = codigo[:-3]   # "600000" → "600", "628000" → "628"
        exact_match = False
    elif codigo.endswith("00") and len(codigo) > 2:
        pref = codigo[:-2]   # "629000" → "6290"
        exact_match = False
    else:
        pref = codigo         # "628030" → exact match
        exact_match = True

    cur_banco = cur_caja = prev_total = 0.0
    for cta, vals in totales_exactos.items():
        cta_s = str(cta)
        if len(cta_s) != len(codigo):
            continue
        if exact_match and cta_s != pref:
            continue
        if not exact_match and not cta_s.startswith(pref):
            continue
        cur_banco  += float(vals.get("cur_banco",  0) or 0)
        cur_caja   += float(vals.get("cur_caja",   0) or 0)
        prev_total += float(vals.get("prev_total", 0) or 0)
    return cur_banco, cur_caja, prev_total


def _fill_row(ws, row, cur_banco, cur_caja, prev_total):
    """Write computed values into the data columns of a template row."""
    corriente  = round(cur_banco + cur_caja, 2)
    acumulado  = round(corriente + prev_total, 2)
    try:
        presupuesto = float(ws.cell(row, COL_PRESUPUESTO).value or 0)
    except (TypeError, ValueError):
        presupuesto = 0.0
    diferencia  = round(presupuesto - acumulado, 2)

    ws.cell(row, COL_CORRIENTE).value  = corriente
    ws.cell(row, COL_BANCO).value      = round(cur_banco,  2)
    ws.cell(row, COL_CAJA).value       = round(cur_caja,   2)
    ws.cell(row, COL_ANTERIORES).value = round(prev_total, 2)
    ws.cell(row, COL_ACUMULADO).value  = acumulado
    ws.cell(row, COL_DIFERENCIA).value = diferencia

    return corriente, cur_banco, cur_caja, prev_total, acumulado


def _normalizar_presupuesto(presupuesto_por_cuenta):
    if not presupuesto_por_cuenta:
        return {}

    normalizado = {}
    for raw_code, raw_value in presupuesto_por_cuenta.items():
        codigo = _normalizar_codigo(raw_code)
        if not codigo:
            continue
        try:
            normalizado[codigo] = float(raw_value or 0)
        except (TypeError, ValueError):
            normalizado[codigo] = 0.0
    return normalizado


def exportar_modelo_evolutivo(ruta_archivo, totales_exactos, cuentas_by_section,
                               anio, periodo_str="", nombre_por_cuenta=None,
                               ruta_plantilla=None, presupuesto_por_cuenta=None):
    """
    Generate the report from live totals and save to ruta_archivo.

    The programmatic path is the primary one because it is the only path that
    is guaranteed to stay aligned with the JSON source. The template-based path
    is kept only as a compatibility fallback.
    """
    try:
        return _exportar_programatico(
            ruta_archivo, totales_exactos, cuentas_by_section,
            anio, periodo_str, nombre_por_cuenta, presupuesto_por_cuenta,
            ruta_plantilla
        )
    except Exception:
        if ruta_plantilla and Path(ruta_plantilla).exists():
            return _exportar_desde_plantilla(
                ruta_archivo, totales_exactos, anio, periodo_str, ruta_plantilla
            )
        raise


# ──────────────────────────────────────────────────────────────────────────────
# Template-fill path  (preferred — preserves the original format exactly)
# ──────────────────────────────────────────────────────────────────────────────

def _exportar_desde_plantilla(ruta_archivo, totales_exactos, anio, periodo_str, ruta_plantilla):
    dest = Path(ruta_archivo)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ruta_plantilla, dest)

    wb = openpyxl.load_workbook(dest)
    ws = wb.active

    # Update date cell (template row 4, col G)
    try:
        ws.cell(4, 7).value = datetime.date(anio, 12, 31)
    except Exception:
        pass

    # Exact base totals from JSON by section.
    section_totals = {s: [0.0, 0.0, 0.0, 0.0] for s in ("6", "7", "2")}
    # [corriente, banco, caja, anteriores]
    for cta, vals in totales_exactos.items():
        cta_s = str(cta).strip()
        if not cta_s or cta_s[0] not in section_totals:
            continue
        banco = float(vals.get("cur_banco", 0) or 0)
        caja = float(vals.get("cur_caja", 0) or 0)
        prev = float(vals.get("prev_total", 0) or 0)
        t = section_totals[cta_s[0]]
        t[0] += banco + caja
        t[1] += banco
        t[2] += caja
        t[3] += prev

    current_section = None
    sheet_codes = []
    candidate_codes = set(str(cta).strip() for cta in totales_exactos.keys())

    for row in range(1, ws.max_row + 1):
        code_raw = ws.cell(row, COL_CUENTA).value
        code_norm = _normalizar_codigo(code_raw)
        if code_norm:
            sheet_codes.append((code_raw, code_norm))
            candidate_codes.add(code_norm)

    for row in range(1, ws.max_row + 1):
        code_raw  = ws.cell(row, COL_CUENTA).value
        label_d   = str(ws.cell(row, COL_NOMBRE).value or "").strip()
        code_str  = str(code_raw or "").strip()
        # Section header: C is literal "CUENTAS..." OR a formula (=+C61) with section keyword in D
        cuenta_hdr = (
            code_str.upper().startswith("CUENTAS") or
            (code_str.startswith("=") and
             any(k in label_d.upper() for k in ("GASTOS", "INGRESOS", "INVER")))
        )

        # Section header rows have "CUENTAS" in col C — skip them and update section
        if cuenta_hdr:
            if "GASTOS"     in label_d: current_section = "6"
            elif "INGRESOS" in label_d: current_section = "7"
            elif "INVER"    in label_d: current_section = "2"
            continue

        # Total rows: col D in TOTAL_LABELS, col C does NOT start with "CUENTAS"
        if label_d in TOTAL_LABELS and current_section:
            t = section_totals[current_section]
            corriente  = round(t[0], 2)
            acumulado  = round(t[0] + t[3], 2)
            try:
                presupuesto = float(ws.cell(row, COL_PRESUPUESTO).value or 0)
            except (TypeError, ValueError):
                presupuesto = 0.0
            ws.cell(row, COL_CORRIENTE).value  = corriente
            ws.cell(row, COL_BANCO).value      = round(t[1], 2)
            ws.cell(row, COL_CAJA).value       = round(t[2], 2)
            ws.cell(row, COL_ANTERIORES).value = round(t[3], 2)
            ws.cell(row, COL_ACUMULADO).value  = acumulado
            ws.cell(row, COL_DIFERENCIA).value = round(float(presupuesto) - acumulado, 2)
            continue

        # Data rows: account code in col C
        codigo, _ = _parse_template_code(code_raw)
        if not codigo or not current_section:
            continue
        if not codigo.isdigit() or codigo[0] not in ("6", "7", "2"):
            continue
        if codigo[0] != current_section:
            continue
        mode = _template_row_mode(ws, row, code_raw, codigo, sheet_codes, candidate_codes)

        cur_banco, cur_caja, prev_total = _rollup(codigo, totales_exactos, mode=mode)
        _fill_row(ws, row, cur_banco, cur_caja, prev_total)

    wb.save(dest)
    return True


# ──────────────────────────────────────────────────────────────────────────────
# Fallback: minimal programmatic layout (used only when template is missing)
# ──────────────────────────────────────────────────────────────────────────────

def _exportar_programatico(ruta_archivo, totales_exactos, cuentas_by_section,
                            anio, periodo_str, nombre_por_cuenta,
                            presupuesto_por_cuenta=None, ruta_plantilla=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BURDEOS = "8C0052"
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Evolutivo"

    fill_h = PatternFill(start_color=BURDEOS, end_color=BURDEOS, fill_type="solid")
    fill_t = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    font_h = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center", vertical="center")
    borde  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    SEC_HEADERS = {
        "6": ["CUENTA", "NOMBRE", "MES CORRIENTE", "BANCO", "CAJA",
              "GASTO MESES ANTERIORES", "SUMA DE GASTO ACUMULADO",
              f"PRESUPUESTO {anio}", "DIFERENCIA"],
        "7": ["CUENTA", "NOMBRE", "MES CORRIENTE", "BANCO", "CAJA",
              "INGRESOS MESES ANTERIORES", "SUMA DE INGRESO ACUMULADO",
              f"PRESUPUESTO {anio}", "DIFERENCIA"],
        "2": ["CUENTA", "NOMBRE", "MES CORRIENTE", "BANCO", "CAJA",
              "INVERSIONES MESES ANTER.", "SUMA DE GASTO ACUMULADO",
              f"PRESUPUESTO {anio}", "DIFERENCIA"],
    }
    SEC_TOTALS = {"6": "SUMA TOTAL GASTOS", "7": "SUMA TOTAL INGRESOS", "2": "INVERSIONES"}
    presupuestos = _normalizar_presupuesto(presupuesto_por_cuenta)

    def section_base_totals(sec):
        cur = banco = caja = prev = 0.0
        for cta, vals in totales_exactos.items():
            cta_s = str(cta).strip()
            if not cta_s.startswith(sec):
                continue
            b = float(vals.get("cur_banco", 0) or 0)
            ca = float(vals.get("cur_caja", 0) or 0)
            pr = float(vals.get("prev_total", 0) or 0)
            banco += b
            caja += ca
            cur += b + ca
            prev += pr
        return cur, banco, caja, prev

    def template_layout_rows():
        if not ruta_plantilla or not Path(ruta_plantilla).exists():
            return {}

        wb_tpl = openpyxl.load_workbook(ruta_plantilla, data_only=False)
        ws_tpl = wb_tpl.active
        sheet_codes = []
        candidate_codes = set(str(cta).strip() for cta in totales_exactos.keys())
        for r in range(1, ws_tpl.max_row + 1):
            raw = ws_tpl.cell(r, COL_CUENTA).value
            norm = _normalizar_codigo(raw)
            if norm:
                sheet_codes.append((raw, norm))
                candidate_codes.add(norm)

        rows_by_sec = {"6": [], "7": [], "2": []}
        current_section = None
        for r in range(1, ws_tpl.max_row + 1):
            raw = ws_tpl.cell(r, COL_CUENTA).value
            label = str(ws_tpl.cell(r, COL_NOMBRE).value or "").strip()
            raw_txt = str(raw or "").strip()
            header = (
                raw_txt.upper().startswith("CUENTAS") or
                (raw_txt.startswith("=") and any(k in label.upper() for k in ("GASTOS", "INGRESOS", "INVER")))
            )
            if header:
                if "GASTOS" in label:
                    current_section = "6"
                elif "INGRESOS" in label:
                    current_section = "7"
                elif "INVER" in label:
                    current_section = "2"
                continue
            if label in TOTAL_LABELS:
                continue

            code = _normalizar_codigo(raw)
            if not code or not current_section or code[0] != current_section:
                continue
            rows_by_sec[current_section].append({
                "code": code,
                "raw": raw,
                "name": label,
                "mode": _template_row_mode(ws_tpl, r, raw, code, sheet_codes, candidate_codes),
            })
        return rows_by_sec

    try:
        layout_rows = template_layout_rows()
    except Exception:
        layout_rows = {}

    row = 1
    section_acumulados = {}
    # Title
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)
    c = ws.cell(row, 3, "CUENTAS DE COMUNIDAD DE SHILLONG")
    c.font = font_h; c.fill = fill_h; c.alignment = centro
    row += 1

    ws.cell(row, 6, "FECHA:")
    ws.cell(row, 7, datetime.date(anio, 12, 31))
    row += 3   # match template offset (row 4 in template = data row)

    for sec in ("6", "7", "2"):
        headers  = SEC_HEADERS[sec]
        n_cols   = len(headers)
        tot_lbl  = SEC_TOTALS[sec]
        rows_layout = list(layout_rows.get(sec, []))
        if rows_layout:
            vistos_cuentas = {item["code"] for item in rows_layout}
            extras = [
                str(c).strip()
                for c in list(totales_exactos.keys()) + list(presupuestos.keys())
                if str(c).strip().startswith(sec)
            ]
            for cta in sorted(set(extras)):
                if cta not in vistos_cuentas:
                    rows_layout.append({"code": cta, "raw": cta, "name": "", "mode": "exact"})
                    vistos_cuentas.add(cta)
        else:
            cuentas  = list(cuentas_by_section.get(sec, []))
            vistos_cuentas = {str(c).strip() for c in cuentas}
            extras = [
                str(c).strip()
                for c in list(totales_exactos.keys()) + list(presupuestos.keys())
                if str(c).strip().startswith(sec)
            ]
            for cta in sorted(set(extras)):
                if cta not in vistos_cuentas:
                    cuentas.append(cta)
                    vistos_cuentas.add(cta)
            rows_layout = [{"code": str(c).strip(), "raw": str(c).strip(), "name": "", "mode": "exact"} for c in cuentas]

        # Header row
        for ci, h in enumerate(headers, start=3):
            c = ws.cell(row, ci, h)
            c.fill = fill_h; c.font = font_h; c.border = borde
            c.alignment = centro
            ws.column_dimensions[get_column_letter(ci)].width = 20
        row += 2

        duplicated_codes = {
            item["code"]
            for item in rows_layout
            if sum(1 for other in rows_layout if other["code"] == item["code"]) > 1
        }
        budget_used = set()
        t_presupuesto = 0.0

        def has_child_budget(group_item):
            code = group_item["code"]
            pref = _group_prefix(code)
            for other in rows_layout:
                other_code = other["code"]
                if other_code == code:
                    continue
                if not other_code.startswith(pref):
                    continue
                if round(presupuestos.get(_normalizar_codigo(other_code), 0.0), 2):
                    return True
            return False

        def has_exact_duplicate(group_item):
            code = group_item["code"]
            return any(
                other is not group_item
                and other["code"] == code
                and other.get("mode") == "exact"
                for other in rows_layout
            )

        for item in rows_layout:
            cta = item["code"]
            nombre = item.get("name") or ""
            if callable(nombre_por_cuenta):
                try:
                    nombre = nombre or nombre_por_cuenta(cta)
                except Exception: pass
            elif isinstance(nombre_por_cuenta, dict):
                nombre = nombre or nombre_por_cuenta.get(cta, "")

            b, ca, pr = _rollup(cta, totales_exactos, mode=item.get("mode", "exact"))
            corr = round(b + ca, 2); acum = round(corr + pr, 2)
            budget_code = _normalizar_codigo(cta)
            if item.get("mode") == "group" and has_exact_duplicate(item):
                presupuesto = 0.0
            elif budget_code in duplicated_codes and budget_code in budget_used:
                presupuesto = 0.0
            else:
                presupuesto = round(presupuestos.get(budget_code, 0.0), 2)
                budget_used.add(budget_code)
            diferencia = round(presupuesto - acum, 2)
            if not (item.get("mode") == "group" and has_child_budget(item)):
                t_presupuesto += presupuesto

            vals = [cta, nombre, corr, round(b,2), round(ca,2), round(pr,2), acum, presupuesto, diferencia]
            for ci, v in enumerate(vals, start=3):
                c = ws.cell(row, ci, v)
                c.border = borde
                if ci >= 5:
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
            row += 1

        # Total row
        t_corr, t_banco, t_caja, t_ant = section_base_totals(sec)
        t_acum = round(t_corr + t_ant, 2)
        section_acumulados[sec] = t_acum
        t_diferencia = round(t_presupuesto - t_acum, 2)
        ws.cell(row, COL_NOMBRE, tot_lbl).font = Font(bold=True)
        for ci, v in [(5, round(t_corr,2)), (6, round(t_banco,2)),
                      (7, round(t_caja,2)), (8, round(t_ant,2)), (9, t_acum),
                      (10, round(t_presupuesto, 2)), (11, t_diferencia)]:
            c = ws.cell(row, ci, v)
            c.fill = fill_t; c.number_format = "#,##0.00"
        row += 2

    ingresos_acum = round(section_acumulados.get("7", 0.0), 2)
    gastos_inversiones_acum = round(
        section_acumulados.get("6", 0.0) + section_acumulados.get("2", 0.0), 2
    )
    balance = round(ingresos_acum - gastos_inversiones_acum, 2)

    ws.cell(row, COL_NOMBRE, "BALANCE").font = Font(bold=True)
    ws.cell(row, COL_NOMBRE).fill = fill_t
    row += 1

    for label, value in (
        ("INGRESOS", ingresos_acum),
        ("GASTOS E INVERSIONES", gastos_inversiones_acum),
        ("BALANCE", balance),
    ):
        label_cell = ws.cell(row, COL_NOMBRE, label)
        value_cell = ws.cell(row, COL_ACUMULADO, value)
        label_cell.border = borde
        value_cell.border = borde
        value_cell.number_format = "#,##0.00"
        value_cell.alignment = Alignment(horizontal="right")
        if label == "BALANCE":
            label_cell.font = Font(bold=True)
            value_cell.font = Font(bold=True)
            label_cell.fill = fill_t
            value_cell.fill = fill_t
        row += 1

    dest = Path(ruta_archivo)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return True
