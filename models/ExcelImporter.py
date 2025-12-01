# -*- coding: utf-8 -*-
"""
ExcelImporter.py — SHILLONG CONTABILIDAD v3.6 PRO
Motor lógico para leer e interpretar archivos Excel de movimientos.
"""

import openpyxl
from datetime import datetime

class ExcelImporter:
    """
    Clase encargada de leer un Excel, validar datos y convertirlos
    al formato de diccionario que usa ContabilidadData.
    """

    # Nombres de columnas esperados en el Excel (puedes añadir variaciones)
    COL_MAP = {
        "FECHA": ["fecha", "date", "día", "dia"],
        "DOCUMENTO": ["documento", "doc", "ref", "referencia"],
        "CONCEPTO": ["concepto", "descripción", "descripcion", "detalle"],
        "CUENTA": ["cuenta", "cta", "rubro", "código", "codigo"],
        "DEBE": ["debe", "gasto", "débito", "cargo", "salida"],
        "HABER": ["haber", "ingreso", "crédito", "abono", "entrada"],
        "BANCO": ["banco", "caja", "tesorería", "origen"],
        "ESTADO": ["estado", "status", "situación"]
    }

    def importar(self, ruta_archivo):
        """
        Lee el archivo y devuelve una lista de diccionarios con los movimientos válidos
        y una lista de errores (si los hay).
        Retorna: (movimientos_validos, lista_errores)
        """
        try:
            wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"No se pudo abrir el archivo: {str(e)}"]

        headers = {}
        movimientos = []
        errores = []

        # 1. Detectar cabeceras en la primera fila
        for cell in ws[1]:
            if cell.value:
                val = str(cell.value).lower().strip()
                # Mapear nombre de columna a nuestra clave interna
                for key, variations in self.COL_MAP.items():
                    if val in variations:
                        headers[key] = cell.column - 1 # Guardamos índice (0-based)

        # Validar que existan columnas mínimas
        if "FECHA" not in headers or "CONCEPTO" not in headers:
            return [], ["El Excel no tiene columnas 'Fecha' o 'Concepto' en la fila 1."]

        # 2. Iterar filas de datos (empezando en fila 2)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # --- FECHA ---
                raw_fecha = row[headers["FECHA"]]
                if not raw_fecha:
                    continue # Saltar filas vacías
                
                fecha_str = self._procesar_fecha(raw_fecha)
                if not fecha_str:
                    errores.append(f"Fila {i}: Fecha inválida ({raw_fecha})")
                    continue

                # --- CONCEPTO ---
                concepto = str(row[headers["CONCEPTO"]] or "").strip()
                if not concepto:
                    errores.append(f"Fila {i}: Falta concepto")
                    continue

                # --- IMPORTES ---
                idx_debe = headers.get("DEBE")
                idx_haber = headers.get("HABER")
                
                debe = float(row[idx_debe]) if idx_debe is not None and isinstance(row[idx_debe], (int, float)) else 0.0
                haber = float(row[idx_haber]) if idx_haber is not None and isinstance(row[idx_haber], (int, float)) else 0.0

                if debe == 0 and haber == 0:
                    # Si no hay importes, saltamos o avisamos (opcional)
                    pass

                # --- CUENTA ---
                idx_cta = headers.get("CUENTA")
                cuenta = str(row[idx_cta]).strip() if (idx_cta is not None and row[idx_cta] is not None) else "PENDIENTE"
                # Limpiar ".0" si el excel lo leyó como float
                if cuenta.endswith(".0"): 
                    cuenta = cuenta[:-2]

                # --- BANCO ---
                idx_banco = headers.get("BANCO")
                banco = str(row[idx_banco]).strip() if (idx_banco is not None and row[idx_banco]) else "Caja"

                # --- ESTADO ---
                idx_estado = headers.get("ESTADO")
                estado = str(row[idx_estado]).lower().strip() if (idx_estado is not None and row[idx_estado]) else "pagado"
                if estado not in ["pagado", "pendiente"]:
                    estado = "pagado"

                # --- DOCUMENTO ---
                idx_doc = headers.get("DOCUMENTO")
                doc = str(row[idx_doc]).strip() if (idx_doc is not None and row[idx_doc]) else f"IMP-{i}"

                # Construir el objeto movimiento
                mov = {
                    "fecha": fecha_str,
                    "documento": doc,
                    "concepto": concepto,
                    "cuenta": cuenta,
                    "debe": debe,
                    "haber": haber,
                    "moneda": "INR", # Por defecto
                    "banco": banco,
                    "estado": estado,
                    "saldo": 0.0 # Se recalculará al insertar
                }
                movimientos.append(mov)

            except Exception as e:
                errores.append(f"Fila {i}: Error procesando datos ({str(e)})")

        return movimientos, errores

    def _procesar_fecha(self, valor):
        """Convierte datetime de Excel o string a 'dd/mm/yyyy'"""
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y")
        elif isinstance(valor, str):
            # Intentar formatos comunes
            formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]
            for f in formatos:
                try:
                    dt = datetime.strptime(valor, f)
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    continue
        return None