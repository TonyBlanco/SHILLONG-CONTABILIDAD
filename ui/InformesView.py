# -*- coding: utf-8 -*-
"""
INFORMESVIEW v4.3 — SHILLONG Contabilidad 3.7.7 PRO BI
-------------------------------------------------------
INCLUYE:
- Diario general
- Libro mayor agrupado (vista + Excel profesional)
- Balance de Sumas y Saldos (vista + Excel profesional)
- Resumen mensual por cuentas
BOTONES:
- Generar Informe
- Exportar Vista Actual
- Exportar Libro Mayor Profesional
- Exportar Balance Profesional
COLORES SHILLONG:
- Morado Encabezado: #7030A0
- Verde Total: #E2EFDA
"""

from PySide6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QPushButton, QComboBox,
    QTableWidget, QTableWidgetItem, QFileDialog, QDateEdit, QScrollArea,
    QCheckBox, QMessageBox
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QAction, QTextDocument

import datetime
import json
import re
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
try:
    # programmatic exporter for Modelo Evolutivo
    from models.ExportadorModeloEvolutivo import exportar_modelo_evolutivo
except Exception:
    exportar_modelo_evolutivo = None

# FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
# Soporte de ruta de recursos en EXE (PyInstaller)
try:
    from utils.rutas import ruta_recurso as _ruta_recurso

    def ruta_recurso(p: str) -> Path:
        return Path(_ruta_recurso(p))
except Exception:  # pragma: no cover
    def ruta_recurso(p: str) -> Path:
        return Path(p)

try:
    from models.SaldosMensuales import SaldosMensuales
except Exception:  # pragma: no cover
    SaldosMensuales = None


class InformesView(QWidget):

    def __init__(self, data):
        super().__init__()
        self.data = data
        self.saldos_sistema = SaldosMensuales() if SaldosMensuales else None
        self._build_ui()

    # ================================================================
    # UI PRINCIPAL
    # ================================================================
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(12)

        # TÍTULO
        titulo = QLabel("📊 Informes Contables")
        titulo.setFont(QFont("Segoe UI", 22, QFont.Bold))
        titulo.setStyleSheet("color: #334155;")
        layout.addWidget(titulo)

        # ------------------------------------------------------------
        # Selector tipo informe
        # ------------------------------------------------------------
        h_sel = QHBoxLayout()
        h_sel.addWidget(QLabel("Tipo de Informe:"))

        self.cbo_tipo = QComboBox()
        self.cbo_tipo.addItems([
            "📘 Diario General (Rango)",
            "📒 Libro Mayor (Por Cuenta)",
            "⚖️ Balance de Sumas y Saldos",
            "🧮 Resumen Mensual por Cuentas"
        ])
        # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
        self.cbo_tipo.addItem("➕ Reporte por Cuentas (Formato Excel)")
        # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
        self.cbo_tipo.addItem("Modelo Evolutivo Presupuestario")
        self.cbo_tipo.addItem("💧 Flujo de Caja Mensual")
        self.cbo_tipo.currentIndexChanged.connect(self._cambiar_tipo)
        h_sel.addWidget(self.cbo_tipo)
        h_sel.addStretch()
        layout.addLayout(h_sel)

        # ------------------------------------------------------------
        # Filtros dinámicos
        # ------------------------------------------------------------
        self.filtros = QHBoxLayout()
        layout.addLayout(self.filtros)

        # ------------------------------------------------------------
        # BOTONES SUPERIORES
        # ------------------------------------------------------------
        self.btn_generar = QPushButton("Generar Informe")
        self.btn_generar.setStyleSheet(
            "background:#2563eb; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_generar.clicked.connect(self._generar)

        self.btn_export_vista = QPushButton("Exportar Vista a Excel")
        self.btn_export_vista.setStyleSheet(
            "background:#475569; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_vista.clicked.connect(self._exportar_excel_vista)

        self.btn_export_mayor = QPushButton("📘 Exportar Libro Mayor Profesional (Excel)")
        self.btn_export_mayor.setStyleSheet(
            "background:#7030A0; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_mayor.clicked.connect(self._exportar_libro_mayor)

        self.btn_export_balance = QPushButton("📊 Exportar Balance Profesional (Excel)")
        self.btn_export_balance.setStyleSheet(
            "background:#7030A0; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_balance.clicked.connect(self._exportar_balance)

        # Botón Imprimir visible (atajo directo desde la barra)
        self.btn_imprimir = QPushButton("Imprimir")
        self.btn_imprimir.setStyleSheet("background:#334155; color:white; padding:6px 20px; font-weight:bold;")
        self.btn_imprimir.clicked.connect(self._imprimir)

        self.chk_shillong = QCheckBox("Estilo SHILLONG (Invertir Columnas: Entra=Debe, Sale=Haber)")
        self.chk_shillong.setStyleSheet("color: #1e40af; font-weight: bold; margin-left: 10px;")
        self.chk_estilo_shillong = self.chk_shillong

        h_btns = QHBoxLayout()
        h_btns.addWidget(self.btn_generar)
        h_btns.addWidget(self.btn_export_vista)
        h_btns.addWidget(self.btn_export_mayor)
        h_btns.addWidget(self.btn_export_balance)
        h_btns.addWidget(self.btn_imprimir)
        h_btns.addWidget(self.chk_shillong)
        h_btns.addStretch()
        layout.addLayout(h_btns)

        # ------------------------------------------------------------
        # SCROLL PARA LAS TABLAS
        # ------------------------------------------------------------
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
        # Impresión/preview accesible sin añadir botones nuevos (menú contextual).
        self.scroll_area.setContextMenuPolicy(Qt.ActionsContextMenu)
        act_preview = QAction("Previsualizar impresión", self)
        act_preview.triggered.connect(self._previsualizar_impresion)
        self.scroll_area.addAction(act_preview)
        act_print = QAction("Imprimir…", self)
        act_print.triggered.connect(self._imprimir)
        self.scroll_area.addAction(act_print)

        self.contenedor = QWidget()
        self.contenedor_layout = QVBoxLayout(self.contenedor)

        self.scroll_area.setWidget(self.contenedor)
        layout.addWidget(self.scroll_area)

        self._cambiar_tipo()

    # ================================================================
    # CAMBIO DE TIPO DE INFORME
    # ================================================================
    def _cambiar_tipo(self):
        tipo = self.cbo_tipo.currentIndex()
        self._limpiar_filtros()
        export_invert_prev = bool(getattr(self, "chk_export_invertido_bi", None) and self.chk_export_invertido_bi.isChecked())

        # recrear widgets
        self.fecha_ini = QDateEdit()
        self.fecha_ini.setCalendarPopup(True)
        self.fecha_ini.setDate(QDate.currentDate().addMonths(-1))

        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())

        self.cbo_cuenta = QComboBox()
        self.cbo_cuenta.addItem("Todas")
        for cta, d in self.data.cuentas.items():
            self.cbo_cuenta.addItem(f"{cta} — {d.get('nombre','')}")

        self.cbo_mes = QComboBox()
        self.cbo_mes.addItems([
            "Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
        ])
        self.cbo_mes.setCurrentIndex(datetime.date.today().month - 1)

        self.cbo_anio = QComboBox()
        for a in range(2020, 2036):
            self.cbo_anio.addItem(str(a))
        self.cbo_anio.setCurrentText(str(datetime.date.today().year))

        self.chk_export_invertido_bi = QCheckBox("Exportar con columnas invertidas (ShillongStyle)")
        self.chk_export_invertido_bi.setChecked(export_invert_prev)

        # Mostrar filtros
        if tipo == 0:  # DIARIO
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)

        elif tipo == 1:
            self.filtros.addWidget(QLabel("Cuenta:"))
            self.filtros.addWidget(self.cbo_cuenta)
            self.filtros.addWidget(QLabel("Ordenar por fecha:"))
            self.cbo_orden_mayor = QComboBox()
            self.cbo_orden_mayor.addItems([
                "Fecha ↑ (más antiguo primero)",
                "Fecha ↓ (más reciente primero)",
                "Sin ordenar (orden de entrada)",
            ])
            self.filtros.addWidget(self.cbo_orden_mayor)

        elif tipo == 2:
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)
            lab = QLabel("Balance profesional SHILLONG agrupado por cuentas para el rango seleccionado.")
            lab.setStyleSheet("color:#475569; font-style:italic;")
            self.filtros.addWidget(lab)

        elif tipo == 3:
            self.filtros.addWidget(QLabel("Mes:"))
            self.filtros.addWidget(self.cbo_mes)
            self.filtros.addWidget(QLabel("Año:"))
            self.filtros.addWidget(self.cbo_anio)

        # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
        elif tipo == 4:
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)
            self.filtros.addWidget(QLabel("Cuenta:"))
            self.filtros.addWidget(self.cbo_cuenta)
            self.filtros.addWidget(self.chk_export_invertido_bi)

        # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
        elif tipo == 5:
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)
            self.filtros.addWidget(self.chk_export_invertido_bi)

        elif tipo == 6:
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)
            lab = QLabel("Entradas y salidas agrupadas por mes para el rango seleccionado.")
            lab.setStyleSheet("color:#475569; font-style:italic;")
            self.filtros.addWidget(lab)

        # visibilidad de botones
        self.btn_export_mayor.setVisible(tipo == 1)
        self.btn_export_balance.setVisible(tipo == 2)

    def _limpiar_filtros(self):
        while self.filtros.count():
            item = self.filtros.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    # ================================================================
    # GENERAR INFORME
    # ================================================================
    def _generar(self):
        tipo = self.cbo_tipo.currentIndex()
        self._limpiar_vista()

        if tipo == 0:
            self._mostrar_diario()
        elif tipo == 1:
            self._mostrar_libro_mayor_agrupado()
        elif tipo == 2:
            self._mostrar_sumas_saldos()
        elif tipo == 3:
            self._mostrar_resumen_mensual()
        # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
        elif tipo == 4:
            self._mostrar_reporte_por_cuentas()
        # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
        elif tipo == 5:
            self._mostrar_reporte_modelo_sisters()
        elif tipo == 6:
            self._mostrar_flujo_caja_mensual()

    def _limpiar_vista(self):
        for i in reversed(range(self.contenedor_layout.count())):
            item = self.contenedor_layout.takeAt(i)
            if item.widget():
                item.widget().deleteLater()

    def _to_float(self, value):
        try:
            return float(str(value).replace(",", "."))
        except (ValueError, TypeError):
            return 0.0

    def _slug_texto(self, texto):
        limpio = re.sub(r"[^\w\s-]", "", str(texto or "").strip(), flags=re.UNICODE)
        limpio = re.sub(r"\s+", "_", limpio)
        return limpio[:80] or "reporte"

    def _carpeta_reportes(self):
        carpeta = Path("reportes")
        carpeta.mkdir(exist_ok=True)
        return carpeta

    def _nombre_archivo_sugerido(self):
        tipo = self.cbo_tipo.currentIndex()
        ini = self.fecha_ini.date().toString("yyyyMMdd") if hasattr(self, "fecha_ini") else ""
        fin = self.fecha_fin.date().toString("yyyyMMdd") if hasattr(self, "fecha_fin") else ""

        if tipo == 0:
            return f"Diario_General_{ini}_{fin}.xlsx"
        if tipo == 1:
            cuenta = self._cuenta_desde_combo(self.cbo_cuenta.currentText()) or "Todas"
            return f"Libro_Mayor_{self._slug_texto(cuenta)}.xlsx"
        if tipo == 2:
            return f"Balance_Sumas_Saldos_{ini}_{fin}.xlsx"
        if tipo == 3:
            anio = self.cbo_anio.currentText()
            mes = f"{self.cbo_mes.currentIndex() + 1:02d}"
            return f"Resumen_Mensual_{anio}_{mes}.xlsx"
        if tipo == 4:
            cuenta = self._cuenta_desde_combo(self.cbo_cuenta.currentText()) or "Todas"
            return f"Reporte_por_Cuentas_{self._slug_texto(cuenta)}_{ini}_{fin}.xlsx"
        if tipo == 5:
            return f"Modelo_Evolutivo_{ini}_{fin}.xlsx"
        if tipo == 6:
            return f"CashFlow_Mensual_{ini}_{fin}.xlsx"
        return f"Reporte_{ini}_{fin}.xlsx"

    def _ruta_exporte_por_defecto(self):
        return str(self._carpeta_reportes() / self._nombre_archivo_sugerido())

    def _cargar_bancos_ordenados(self):
        try:
            with open("data/bancos.json", "r", encoding="utf-8") as f:
                bancos = [b.get("nombre", "").strip() for b in json.load(f).get("banks", [])]
                bancos = [b for b in bancos if b]
                if bancos:
                    return bancos
        except (IOError, json.JSONDecodeError, AttributeError):
            pass
        return []

    def _obtener_saldos_iniciales_rango(self, fecha_inicio, bancos):
        saldos = {b: 0.0 for b in bancos}
        if not self.saldos_sistema:
            return saldos
        mes = fecha_inicio.month
        anio = fecha_inicio.year
        for banco in bancos:
            saldo = self.saldos_sistema.obtener_saldo_inicial(mes, anio, banco)
            saldos[banco] = float(saldo or 0.0)
        return saldos

    # ================================================================
    # DIARIO GENERAL
    # ================================================================
    def _mostrar_diario(self):
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()
        invertir = self.chk_shillong.isChecked()
        
        datos = self.data.get_movimientos_rango(ini, fin)
        
        # Agrupar por banco (exacto) y preparar orden fijo del catálogo
        por_banco = defaultdict(list)
        for m in datos:
            b = str(m.get("banco", "Caja") or "Caja").strip()
            por_banco[b].append(m)

        bancos_orden = self._cargar_bancos_ordenados()
        if not bancos_orden:
            bancos_orden = sorted(por_banco.keys())
        extras = sorted([b for b in por_banco.keys() if b not in bancos_orden])
        bancos_orden.extend(extras)

        saldos_iniciales = self._obtener_saldos_iniciales_rango(ini, bancos_orden)
            
        header_labels = ["Fecha","Documento","Concepto","Cuenta","Entra (Debe)" if invertir else "Debe","Sale (Haber)" if invertir else "Haber","Saldo","Banco"]
        total_debe_general = 0.0
        total_haber_general = 0.0
        total_saldo_general = 0.0
        
        for banco in bancos_orden:
            self.contenedor_layout.addWidget(QLabel(f"🏦 BANCO: {banco}"))
            tabla = QTableWidget(0, 8)
            tabla.setHorizontalHeaderLabels(header_labels)
            
            saldo_inicial = float(saldos_iniciales.get(banco, 0.0))
            saldo_acum = saldo_inicial
            total_debe_banco = 0.0
            total_haber_banco = 0.0

            # Fila inicial independiente por banco/caja
            r = tabla.rowCount()
            tabla.insertRow(r)
            fila_inicial = [
                ini.strftime("%d/%m/%Y"),
                "",
                "SALDO INICIAL",
                "",
                "",
                "",
                round(saldo_inicial, 2),
                banco
            ]
            for c, val in enumerate(fila_inicial):
                it = QTableWidgetItem(str(val))
                if c >= 4:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla.setItem(r, c, it)

            for m in por_banco.get(banco, []):
                r = tabla.rowCount()
                tabla.insertRow(r)
                
                d = self._to_float(m.get("debe", 0) or 0)
                h = self._to_float(m.get("haber", 0) or 0)
                
                if invertir:
                    val_debe = h
                    val_haber = d
                    saldo_acum += (val_debe - val_haber)
                else:
                    val_debe = d
                    val_haber = h
                    saldo_acum += (val_haber - val_debe)

                total_debe_banco += val_debe
                total_haber_banco += val_haber
                
                fila = [
                    m.get("fecha",""),
                    m.get("documento",""),
                    m.get("concepto",""),
                    m.get("cuenta",""),
                    val_debe,
                    val_haber,
                    round(saldo_acum, 2),
                    banco
                ]
                for c, val in enumerate(fila):
                    it = QTableWidgetItem(str(val))
                    if c >= 4:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    tabla.setItem(r,c,it)
            
            # Fila Total Banco
            r = tabla.rowCount()
            tabla.insertRow(r)
            it_tot = QTableWidgetItem(f"TOTAL {banco}")
            it_tot.setFont(QFont("Segoe UI", 9, QFont.Bold))
            tabla.setItem(r, 2, it_tot)
            tabla.setItem(r, 4, QTableWidgetItem(str(round(total_debe_banco, 2))))
            tabla.setItem(r, 5, QTableWidgetItem(str(round(total_haber_banco, 2))))
            tabla.setItem(r, 6, QTableWidgetItem(str(round(saldo_acum, 2))))
            tabla.item(r, 4).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tabla.item(r, 5).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tabla.item(r, 6).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tabla.item(r, 4).setFont(QFont("Segoe UI", 9, QFont.Bold))
            tabla.item(r, 5).setFont(QFont("Segoe UI", 9, QFont.Bold))
            tabla.item(r, 6).setFont(QFont("Segoe UI", 9, QFont.Bold))
            
            self.contenedor_layout.addWidget(tabla)
            self.contenedor_layout.addSpacing(20)

            total_debe_general += total_debe_banco
            total_haber_general += total_haber_banco
            total_saldo_general += saldo_acum

        # Suma total global (caja + todos los bancos)
        self.contenedor_layout.addWidget(QLabel("🔢 TOTAL GENERAL TESORERÍA"))
        resumen = QTableWidget(1, 4)
        resumen.setHorizontalHeaderLabels([
            "Resumen",
            "Entra (Debe)" if invertir else "Debe",
            "Sale (Haber)" if invertir else "Haber",
            "Saldo Final"
        ])
        resumen.setItem(0, 0, QTableWidgetItem("TOTAL CAJA + BANCOS"))
        resumen.setItem(0, 1, QTableWidgetItem(str(round(total_debe_general, 2))))
        resumen.setItem(0, 2, QTableWidgetItem(str(round(total_haber_general, 2))))
        resumen.setItem(0, 3, QTableWidgetItem(str(round(total_saldo_general, 2))))
        for c in (1, 2, 3):
            if resumen.item(0, c):
                resumen.item(0, c).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                resumen.item(0, c).setFont(QFont("Segoe UI", 9, QFont.Bold))
        self.contenedor_layout.addWidget(resumen)

    # ================================================================
    # LIBRO MAYOR AGRUPADO
    # ================================================================
    def _mostrar_libro_mayor_agrupado(self):

        texto = self.cbo_cuenta.currentText()
        if texto == "Todas":
            cuentas = sorted(self.data.cuentas.keys())
        else:
            cuentas = [texto.split(" — ")[0]]

        for cta in cuentas:
            movs = self.data.movimientos_por_cuenta(cta)
            if not movs:
                continue

            # Ordenar por fecha según selección del combo
            orden = getattr(self, "cbo_orden_mayor", None)
            orden_idx = orden.currentIndex() if orden else 0
            if orden_idx in (0, 1):
                def _parse_f(f):
                    try:
                        return datetime.date.fromisoformat(f)
                    except Exception:
                        return datetime.date.min
                movs = sorted(movs, key=lambda m: _parse_f(m.get("fecha", "")), reverse=(orden_idx == 1))

            nombre = self.data.cuentas[cta].get("nombre","")

            header = QLabel(f"{cta} — {nombre}")
            header.setStyleSheet(
                "background:#7030A0; color:white; font-size:16px;"
                "padding:6px; font-weight:bold;"
            )
            self.contenedor_layout.addWidget(header)

            tabla = QTableWidget(0, 6)
            tabla.setHorizontalHeaderLabels(
                ["Fecha","Documento","Desglose","Debe","Haber","Saldo"]
            )

            total_debe=0
            total_haber=0
            saldo_acum=0

            for m in movs:
                r = tabla.rowCount()
                tabla.insertRow(r)

                concepto = m.get("concepto","").strip()
                if concepto:
                    des = concepto
                else:
                    des = m.get("nombre_cuenta","")

                debe = float(m.get("debe",0))
                haber = float(m.get("haber",0))

                saldo_acum += (haber - debe)
                total_debe += debe
                total_haber += haber

                fila = [
                    m.get("fecha",""),
                    m.get("documento",""),
                    des,
                    debe,
                    haber,
                    saldo_acum
                ]

                for c,val in enumerate(fila):
                    it = QTableWidgetItem(str(val))
                    if c>=3:
                        it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                    tabla.setItem(r,c,it)

            r = tabla.rowCount()
            tabla.insertRow(r)
            tabla.setItem(r,2, QTableWidgetItem("TOTAL"))
            tabla.setItem(r,3, QTableWidgetItem(str(total_debe)))
            tabla.setItem(r,4, QTableWidgetItem(str(total_haber)))
            tabla.setItem(r,5, QTableWidgetItem(str(total_haber-total_debe)))

            self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # SUMAS & SALDOS — VISTA SHILLONG
    # ================================================================
    def _mostrar_sumas_saldos(self):
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()

        tabla = QTableWidget(0,5)
        tabla.setHorizontalHeaderLabels(
            ["Cuenta","Nombre","Debe","Haber","Saldo"]
        )

        resumen = defaultdict(lambda: {"nombre":"", "debe":0, "haber":0})

        for m in self._movimientos_en_rango(ini, fin):
            cta = str(m.get("cuenta",""))
            resumen[cta]["nombre"] = m.get("nombre_cuenta","")
            resumen[cta]["debe"] += float(m.get("debe",0))
            resumen[cta]["haber"] += float(m.get("haber",0))

        total_debe=0
        total_haber=0

        for cta in sorted(resumen.keys()):
            d = resumen[cta]
            saldo = d["haber"]-d["debe"]

            total_debe += d["debe"]
            total_haber+= d["haber"]

            r = tabla.rowCount()
            tabla.insertRow(r)
            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila):
                it = QTableWidgetItem(str(val))
                if c>=2:
                    it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                tabla.setItem(r,c,it)

        r = tabla.rowCount()
        tabla.insertRow(r)
        tabla.setItem(r,1, QTableWidgetItem("TOTAL GENERAL"))
        tabla.setItem(r,2, QTableWidgetItem(str(total_debe)))
        tabla.setItem(r,3, QTableWidgetItem(str(total_haber)))
        tabla.setItem(r,4, QTableWidgetItem(str(total_haber-total_debe)))

        self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # RESUMEN MENSUAL
    # ================================================================
    def _mostrar_resumen_mensual(self):

        mes = self.cbo_mes.currentIndex()+1
        anio = int(self.cbo_anio.currentText())

        movs = self.data.movimientos_por_mes(mes, anio)

        resumen = defaultdict(lambda: {"nombre":"", "debe":0, "haber":0})

        for m in movs:
            cta=str(m.get("cuenta",""))
            resumen[cta]["nombre"]=m.get("nombre_cuenta","")
            resumen[cta]["debe"]+=float(m.get("debe",0))
            resumen[cta]["haber"]+=float(m.get("haber",0))

        tabla = QTableWidget(0,5)
        tabla.setHorizontalHeaderLabels(
            ["Cuenta","Nombre","Debe","Haber","Saldo"]
        )

        for cta in sorted(resumen.keys()):
            d=resumen[cta]
            saldo=d["haber"]-d["debe"]

            r=tabla.rowCount()
            tabla.insertRow(r)
            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila):
                it=QTableWidgetItem(str(val))
                if c>=2:
                    it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                tabla.setItem(r,c,it)

        self.contenedor_layout.addWidget(tabla)

    def _mostrar_flujo_caja_mensual(self):
        resumen, detalle_bancos = self._datos_flujo_caja_mensual()

        if not resumen:
            lab = QLabel("No hay movimientos para el rango seleccionado.")
            lab.setStyleSheet("color:#475569; font-style:italic;")
            self.contenedor_layout.addWidget(lab)
            return

        tabla = QTableWidget(0, 7)
        tabla.setHorizontalHeaderLabels([
            "Mes",
            "Entra Banco",
            "Entra Caja",
            "Sale Banco",
            "Sale Caja",
            "Flujo Neto",
            "Acumulado",
        ])

        acumulado = 0.0
        total_entra_banco = 0.0
        total_entra_caja = 0.0
        total_sale_banco = 0.0
        total_sale_caja = 0.0

        for fila_data in resumen:
            entra_banco = fila_data["entra_banco"]
            entra_caja = fila_data["entra_caja"]
            sale_banco = fila_data["sale_banco"]
            sale_caja = fila_data["sale_caja"]
            flujo = fila_data["flujo"]
            acumulado = fila_data["acumulado"]
            total_entra_banco += entra_banco
            total_entra_caja += entra_caja
            total_sale_banco += sale_banco
            total_sale_caja += sale_caja

            r = tabla.rowCount()
            tabla.insertRow(r)
            fila = [
                fila_data["mes"],
                entra_banco,
                entra_caja,
                sale_banco,
                sale_caja,
                flujo,
                acumulado,
            ]
            for c, val in enumerate(fila):
                it = QTableWidgetItem(str(val))
                if c >= 1:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla.setItem(r, c, it)

        r = tabla.rowCount()
        tabla.insertRow(r)
        tabla.setItem(r, 0, QTableWidgetItem("TOTAL"))
        tabla.setItem(r, 1, QTableWidgetItem(str(round(total_entra_banco, 2))))
        tabla.setItem(r, 2, QTableWidgetItem(str(round(total_entra_caja, 2))))
        tabla.setItem(r, 3, QTableWidgetItem(str(round(total_sale_banco, 2))))
        tabla.setItem(r, 4, QTableWidgetItem(str(round(total_sale_caja, 2))))
        tabla.setItem(
            r,
            5,
            QTableWidgetItem(
                str(
                    round(
                        (total_entra_banco + total_entra_caja)
                        - (total_sale_banco + total_sale_caja),
                        2,
                    )
                )
            ),
        )
        tabla.setItem(r, 6, QTableWidgetItem(str(acumulado)))
        for c in range(1, 7):
            if tabla.item(r, c):
                tabla.item(r, c).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla.item(r, c).setFont(QFont("Segoe UI", 9, QFont.Bold))

        self.contenedor_layout.addWidget(tabla)

        self.contenedor_layout.addWidget(QLabel("Desglose por Banco"))
        tabla_bancos = QTableWidget(0, 6)
        tabla_bancos.setHorizontalHeaderLabels([
            "Mes",
            "Banco",
            "Entra",
            "Sale",
            "Flujo Neto",
            "Acumulado",
        ])

        acumulado_por_banco = defaultdict(float)
        for fila_data in detalle_bancos:
            banco = fila_data["banco"]
            flujo = round(fila_data["entra"] - fila_data["sale"], 2)
            acumulado_por_banco[banco] = round(acumulado_por_banco[banco] + flujo, 2)

            r = tabla_bancos.rowCount()
            tabla_bancos.insertRow(r)
            fila = [
                fila_data["mes"],
                banco,
                fila_data["entra"],
                fila_data["sale"],
                flujo,
                acumulado_por_banco[banco],
            ]
            for c, val in enumerate(fila):
                it = QTableWidgetItem(str(val))
                if c >= 2:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla_bancos.setItem(r, c, it)

        self.contenedor_layout.addWidget(tabla_bancos)

    # ================================================================
    # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
    # ================================================================
    def _parsear_fecha_movimiento(self, fecha_str):
        """
        Retorna datetime.date o None.
        Formatos soportados: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY.
        """
        if not fecha_str:
            return None
        s = str(fecha_str).strip()
        try:
            if "/" in s:
                return datetime.datetime.strptime(s, "%d/%m/%Y").date()
            if "-" in s:
                try:
                    return datetime.datetime.strptime(s, "%Y-%m-%d").date()
                except ValueError:
                    return datetime.datetime.strptime(s, "%d-%m-%Y").date()
        except ValueError:
            return None
        return None

    def _cuenta_desde_combo(self, texto):
        if texto == "Todas":
            return None
        if " — " in texto:
            return texto.split(" — ", 1)[0].strip()
        if " - " in texto:
            return texto.split(" - ", 1)[0].strip()
        return str(texto).strip() or None

    def _movimientos_en_rango(self, fecha_inicio, fecha_fin):
        movimientos = []
        for m in getattr(self.data, "movimientos", []):
            fecha_obj = self._parsear_fecha_movimiento(m.get("fecha", ""))
            if fecha_obj is None:
                continue
            if fecha_inicio <= fecha_obj <= fecha_fin:
                movimientos.append(m)
        return movimientos

    def _obtener_datos_reporte_por_cuentas(self):
        """
        Agrupa movimientos por cuenta contable usando:
        - self.data.movimientos
        - self.data.obtener_nombre_cuenta(cuenta)
        Respeta rango de fechas seleccionado y cuenta (Todas o específica).
        """
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()

        texto_cuenta = self.cbo_cuenta.currentText()
        cuenta_filtro = self._cuenta_desde_combo(texto_cuenta)

        agrupado = defaultdict(list)  # {cuenta: [(fecha_obj, mov_dict), ...]}
        for m in getattr(self.data, "movimientos", []):
            cuenta = str(m.get("cuenta", "")).strip()
            if not cuenta:
                continue
            if cuenta_filtro and cuenta != str(cuenta_filtro):
                continue

            fecha_obj = self._parsear_fecha_movimiento(m.get("fecha", ""))
            if fecha_obj is None:
                continue
            if not (ini <= fecha_obj <= fin):
                continue

            agrupado[cuenta].append((fecha_obj, m))

        salida = []
        for cuenta in sorted(agrupado.keys()):
            movs = agrupado[cuenta]
            movs.sort(key=lambda t: (t[0], str(t[1].get("documento", ""))))

            if hasattr(self.data, "obtener_nombre_cuenta"):
                nombre = self.data.obtener_nombre_cuenta(cuenta)
            else:
                nombre = self.data.cuentas.get(cuenta, {}).get("nombre", "")

            salida.append((cuenta, nombre, movs))
        return salida

    def _mostrar_reporte_por_cuentas(self):
        """
        Columnas:
        Cuenta | Fecha | Concepto | Debe | Haber | Estado | Documento
        """
        headers = ["Cuenta", "Fecha", "Concepto", "Debe", "Haber", "Estado", "Documento"]

        datos = self._obtener_datos_reporte_por_cuentas()
        if not datos:
            lab = QLabel("No hay movimientos para el rango/cuenta seleccionado.")
            lab.setStyleSheet("color:#475569; font-style:italic;")
            self.contenedor_layout.addWidget(lab)
            return

        for cuenta, nombre, movs in datos:
            header = QLabel(f"{cuenta} - {nombre}")
            header.setStyleSheet(
                "background:#7030A0; color:white; font-size:16px;"
                "padding:6px; font-weight:bold;"
            )
            self.contenedor_layout.addWidget(header)

            tabla = QTableWidget(0, len(headers))
            tabla.setHorizontalHeaderLabels(headers)
            tabla.horizontalHeader().setSectionsMovable(True)  # reordenar solo visual

            total_debe = 0.0
            total_haber = 0.0

            for fecha_obj, m in movs:
                debe = float(m.get("debe", 0) or 0)
                haber = float(m.get("haber", 0) or 0)
                total_debe += debe
                total_haber += haber

                r = tabla.rowCount()
                tabla.insertRow(r)
                fila = [
                    cuenta,
                    m.get("fecha", fecha_obj.strftime("%d/%m/%Y")),
                    m.get("concepto", ""),
                    debe,
                    haber,
                    m.get("estado", ""),
                    m.get("documento", ""),
                ]

                for c, val in enumerate(fila):
                    it = QTableWidgetItem(str(val))
                    if headers[c] in ("Debe", "Haber"):
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    tabla.setItem(r, c, it)

            # Total por cuenta
            r = tabla.rowCount()
            tabla.insertRow(r)
            tabla.setItem(r, 2, QTableWidgetItem("TOTAL"))
            tabla.setItem(r, 3, QTableWidgetItem(str(total_debe)))
            tabla.setItem(r, 4, QTableWidgetItem(str(total_haber)))
            if tabla.item(r, 3):
                tabla.item(r, 3).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if tabla.item(r, 4):
                tabla.item(r, 4).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
    # ================================================================
    def _ruta_modelo_sisters_layout(self):
        """
        Retorna la ruta fija del layout oficial.
        Regla: nunca usar datos de una plantilla cargada por el usuario.
        """
        # Prefer an explicit blank template first (no sample data)
        tried = []
        blank = Path(ruta_recurso("ui/Modelo evolutivo presupuestario 26 - blank.xlsx"))
        tried.append(str(blank))
        if blank.exists():
            return str(blank)

        # Next, prefer the main template in ui/
        principal = Path(ruta_recurso("ui/Modelo evolutivo presupuestario 26.xlsx"))
        tried.append(str(principal))
        if principal.exists():
            return str(principal)

        # Then try an alternative template that exists in the repo
        alternative = Path(ruta_recurso("ui/Modelo por cuentas.xlsx"))
        tried.append(str(alternative))
        if alternative.exists():
            return str(alternative)

        # If none found, raise a clear error (caller should handle it)
        raise FileNotFoundError("Ninguna plantilla de modelo encontrada. Se buscaron: " + ", ".join(tried))

    def _celda_es_formula(self, cell):
        v = cell.value
        if v is None:
            return False
        if isinstance(v, str) and v.startswith("="):
            return True
        return getattr(cell, "data_type", None) == "f"

    def _es_caja(self, banco):
        s = str(banco or "").strip().lower()
        return ("caja" in s) or (s == "cash")

    def _modo_invertido_bi(self):
        return bool(
            (getattr(self, "chk_export_invertido_bi", None) and self.chk_export_invertido_bi.isChecked())
            or (getattr(self, "chk_shillong", None) and self.chk_shillong.isChecked())
        )

    def _normalizar_codigo_template(self, valor_celda, codigos_disponibles):
        """
        Convierte códigos del layout (ej: 602.10, 602,10, 629.2, 651, 668)
        al código contable interno (ej: 602100, 629200, 651000, 668000).
        """
        if valor_celda is None:
            return None

        txt = str(valor_celda).strip()
        if not txt:
            return None

        # Solo dígitos para comparar
        digitos = re.sub(r"\D", "", txt)
        if not digitos:
            return None

        candidatos = []
        candidatos.append(digitos)
        candidatos.append(digitos + "0")
        candidatos.append(digitos + "00")

        # Normalizaciones típicas del plan (3/4/5 dígitos -> 6 dígitos)
        if len(digitos) == 3:
            candidatos.append(digitos + "000")
        elif len(digitos) == 4:
            candidatos.append(digitos + "00")
        elif len(digitos) == 5:
            candidatos.append(digitos + "0")

        # Priorizar coincidencia exacta en catálogo/movimientos
        vistos = set()
        for c in candidatos:
            if c in vistos:
                continue
            vistos.add(c)
            if c in codigos_disponibles:
                return c

        # Si no hay match, usar el más largo (6 dígitos si es posible)
        if len(digitos) < 6:
            return (digitos + ("0" * (6 - len(digitos))))[:6]
        return digitos

    def _parsear_codigo_template(self, valor_celda, codigos_disponibles):
        codigo = self._normalizar_codigo_template(valor_celda, codigos_disponibles)
        if not codigo:
            return None, "auto"
        return codigo, "auto"

    def _prefijo_grupo_codigo(self, codigo):
        codigo = str(codigo).strip()
        if codigo.endswith("000") and len(codigo) > 3:
            return codigo[:-3]
        if codigo.endswith("00") and len(codigo) > 2:
            return codigo[:-2]
        return codigo

    def _tiene_descendientes_codigo(self, codigo, candidate_codes):
        pref = self._prefijo_grupo_codigo(codigo)
        for candidate in candidate_codes:
            cta = str(candidate).strip()
            if len(cta) != len(str(codigo)):
                continue
            if cta != str(codigo) and cta.startswith(pref):
                return True
        return False

    def _resolver_modo_codigo_template(self, valor_celda, codigo, sheet_codes, candidate_codes):
        txt = str(valor_celda or "").strip()
        if not codigo:
            return "auto"

        if "-" in txt:
            return "group"

        partes = [p for p in re.split(r"[.,]", txt) if p]
        tiene_sep = any(sep in txt for sep in (".", ","))

        if tiene_sep:
            for other_raw, other_norm in sheet_codes:
                if other_norm != codigo:
                    continue
                other_txt = str(other_raw or "").strip()
                if other_txt == txt:
                    continue
                if not any(sep in other_txt for sep in (".", ",", "-")):
                    return "exact"

        if tiene_sep and any(len(p) >= 3 for p in partes[1:]):
            return "exact"

        if self._tiene_descendientes_codigo(codigo, candidate_codes):
            return "group"

        if tiene_sep:
            return "exact"
        return "auto"

    def _totales_sisters_por_cuenta(self, invertido):
        """
        Retorna dict:
        {cuenta: {"cur_banco":x, "cur_caja":y, "prev_total":z}}
        """
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()

        year_start = datetime.date(ini.year, 1, 1)
        prev_end = ini - datetime.timedelta(days=1)

        tot = defaultdict(lambda: {"cur_banco": 0.0, "cur_caja": 0.0, "prev_total": 0.0})

        for m in getattr(self.data, "movimientos", []):
            cuenta = str(m.get("cuenta", "")).strip()
            if not cuenta or not cuenta[0].isdigit():
                continue
            if cuenta[0] not in ("6", "7", "2"):
                continue

            fecha_obj = self._parsear_fecha_movimiento(m.get("fecha", ""))
            if fecha_obj is None:
                continue

            debe = float(m.get("debe", 0) or 0)
            haber = float(m.get("haber", 0) or 0)

            # Para el modelo evolutivo siempre se usan los importes contables reales
            # del asiento: ingresos desde HABER, gastos/inversiones desde DEBE.
            # El modo invertido de Shillong afecta a la lectura visual de otros
            # informes, no a esta base de cálculo.
            if cuenta[0] == "7":
                importe = haber
            else:
                importe = debe

            if not importe:
                continue

            if ini <= fecha_obj <= fin:
                if self._es_caja(m.get("banco", "")):
                    tot[cuenta]["cur_caja"] += importe
                else:
                    tot[cuenta]["cur_banco"] += importe
            elif year_start <= fecha_obj <= prev_end:
                tot[cuenta]["prev_total"] += importe

        return tot

    def _rollup_sisters(self, cuenta_codigo, totales_exactos, mode="auto"):
        """
        Suma subcuentas si el código termina en 00/000 (padre), manteniendo largo fijo.
        """
        codigo = str(cuenta_codigo).strip()
        if not codigo:
            return {"cur_banco": 0.0, "cur_caja": 0.0, "prev_total": 0.0}

        if mode == "exact":
            pref = codigo
            exact_match = True
        elif mode == "group":
            exact_match = False
            if codigo.endswith("000") and len(codigo) > 3:
                pref = codigo[:-3]
            elif codigo.endswith("00") and len(codigo) > 2:
                pref = codigo[:-2]
            else:
                pref = codigo
        elif codigo.endswith("000") and len(codigo) > 3:
            pref = codigo[:-3]
            exact_match = False
        elif codigo.endswith("00") and len(codigo) > 2:
            pref = codigo[:-2]
            exact_match = False
        else:
            pref = codigo
            exact_match = True

        cur_banco = 0.0
        cur_caja = 0.0
        prev_total = 0.0
        for cta, vals in totales_exactos.items():
            cta_s = str(cta)
            if len(cta_s) != len(codigo):
                continue
            if exact_match and cta_s != pref:
                continue
            if not exact_match and not cta_s.startswith(pref):
                continue
            cur_banco += float(vals.get("cur_banco", 0) or 0)
            cur_caja += float(vals.get("cur_caja", 0) or 0)
            prev_total += float(vals.get("prev_total", 0) or 0)

        return {"cur_banco": cur_banco, "cur_caja": cur_caja, "prev_total": prev_total}

    def _headers_modelo_sisters(self, seccion, anio):
        if seccion == "6":
            return [
                "CUENTA",
                "NOMBRE",
                "GASTOS",
                "BANCO",
                "CAJA",
                "GASTO MESES ANTERIORES",
                "SUMA DE GASTO ACUMULADO",
                f"PRESUPUESTO {anio}",
                "DIFERENCIA",
            ]
        if seccion == "7":
            return [
                "CUENTA",
                "NOMBRE",
                "INGRESOS",
                "BANCO",
                "CAJA",
                "INGRESOS MESES ANTERIORES",
                "SUMA DE INGRESO ACUMULADO",
                f"PRESUPUESTO {anio}",
                "DIFERENCIA",
            ]
        return [
            "CUENTA",
            "NOMBRE",
            "INVERSIONES",
            "BANCO",
            "CAJA",
            "INVERSIONES MESES ANTER.",
            "SUMA DE GASTO ACUMULADO",
            f"PRESUPUESTO {anio}",
            "DIFERENCIA",
        ]

    def _mostrar_reporte_modelo_sisters(self):
        # Regla: este informe SIEMPRE usa datos de `self.data.movimientos` (shillong_2026),
        # nunca valores numéricos embebidos en una plantilla.
        invertido = self._modo_invertido_bi()
        totales_exactos = self._totales_sisters_por_cuenta(invertido=invertido)
        anio = self.fecha_ini.date().toPython().year

        def cuentas_seccion(prefijo):
            # Prioridad: respetar el orden del layout oficial (sin leer datos numéricos).
            cuentas_layout = []
            try:
                wb = openpyxl.load_workbook(self._ruta_modelo_sisters_layout(), data_only=False)
                ws = wb[wb.sheetnames[0]]
                for r in range(1, ws.max_row + 1):
                    v = ws.cell(r, 3).value  # C = código cuenta
                    if v is None:
                        continue
                    if isinstance(v, (int, float)):
                        cta_s = str(int(v))
                    else:
                        cta_s = str(v).strip()
                    if cta_s.isdigit() and cta_s.startswith(prefijo):
                        cuentas_layout.append(cta_s)
            except Exception:
                cuentas_layout = []

            cuentas_plan = []
            for cta in getattr(self.data, "cuentas", {}).keys():
                cta_s = str(cta).strip()
                if cta_s.startswith(prefijo):
                    cuentas_plan.append(cta_s)

            cuentas_mov = []
            for cta in totales_exactos.keys():
                cta_s = str(cta).strip()
                if cta_s.startswith(prefijo):
                    cuentas_mov.append(cta_s)

            # Orden final: primero el layout, luego cualquier cuenta extra.
            ordenadas = []
            vistos = set()
            for c in cuentas_layout:
                if c not in vistos:
                    ordenadas.append(c)
                    vistos.add(c)
            for c in sorted(set(cuentas_plan + cuentas_mov)):
                if c not in vistos:
                    ordenadas.append(c)
                    vistos.add(c)
            return ordenadas

        def render_seccion(seccion, titulo_total):
            header = QLabel(titulo_total.replace("SUMA TOTAL ", ""))
            header.setStyleSheet(
                "background:#7030A0; color:white; font-size:16px;"
                "padding:6px; font-weight:bold;"
            )
            self.contenedor_layout.addWidget(header)

            headers = self._headers_modelo_sisters(seccion, anio)
            tabla = QTableWidget(0, len(headers))
            tabla.setHorizontalHeaderLabels(headers)

            total_cur = 0.0
            total_banco = 0.0
            total_caja = 0.0
            total_prev = 0.0

            for cuenta in cuentas_seccion(seccion):
                nombre = ""
                if hasattr(self.data, "obtener_nombre_cuenta"):
                    nombre = self.data.obtener_nombre_cuenta(cuenta)
                vals = self._rollup_sisters(cuenta, totales_exactos)
                cur_banco = float(vals["cur_banco"])
                cur_caja = float(vals["cur_caja"])
                cur_total = cur_banco + cur_caja
                prev_total = float(vals["prev_total"])
                acum = cur_total + prev_total

                presupuesto_num = 0.0
                diferencia_num = presupuesto_num - acum

                r = tabla.rowCount()
                tabla.insertRow(r)
                fila = [
                    cuenta,
                    nombre,
                    cur_total,
                    cur_banco,
                    cur_caja,
                    prev_total,
                    acum,
                    presupuesto_num,
                    diferencia_num,
                ]

                for c, val in enumerate(fila):
                    it = QTableWidgetItem(str(val))
                    if c >= 2:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    tabla.setItem(r, c, it)

            # Evitar doble conteo: sumar solo base (código exacto)
            for cta, vals in totales_exactos.items():
                cta_s = str(cta)
                if cta_s.startswith(seccion):
                    b = float(vals.get("cur_banco", 0) or 0)
                    ca = float(vals.get("cur_caja", 0) or 0)
                    pr = float(vals.get("prev_total", 0) or 0)
                    total_banco += b
                    total_caja += ca
                    total_cur += (b + ca)
                    total_prev += pr

            r = tabla.rowCount()
            tabla.insertRow(r)
            tabla.setItem(r, 1, QTableWidgetItem(titulo_total))
            tabla.setItem(r, 2, QTableWidgetItem(str(total_cur)))
            tabla.setItem(r, 3, QTableWidgetItem(str(total_banco)))
            tabla.setItem(r, 4, QTableWidgetItem(str(total_caja)))
            tabla.setItem(r, 5, QTableWidgetItem(str(total_prev)))
            tabla.setItem(r, 6, QTableWidgetItem(str(total_cur + total_prev)))
            for c in range(2, 7):
                if tabla.item(r, c):
                    tabla.item(r, c).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            self.contenedor_layout.addWidget(tabla)

        render_seccion("6", "SUMA TOTAL GASTOS")
        render_seccion("7", "SUMA TOTAL INGRESOS")
        render_seccion("2", "INVERSIONES")

    # ================================================================
    # EXPORTAR VISTA ACTUAL
    # ================================================================
    def _exportar_excel_vista(self):

        ruta,_ = QFileDialog.getSaveFileName(
            self, "Exportar Informe", self._ruta_exporte_por_defecto(), "Excel (*.xlsx)"
        )
        if not ruta:
            return

        # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
        if self.cbo_tipo.currentIndex() == 4:
            self._exportar_excel_reporte_por_cuentas(ruta)
            return

        # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
        if self.cbo_tipo.currentIndex() == 5:
            self._exportar_excel_modelo_sisters(ruta)
            return

        if self.cbo_tipo.currentIndex() == 6:
            self._exportar_excel_flujo_caja(ruta)
            return

        wb=openpyxl.Workbook()
        ws=wb.active

        row=1

        # Vuelca todas las tablas y encabezados de la vista actual
        for i in range(self.contenedor_layout.count()):
            w = self.contenedor_layout.itemAt(i).widget()

            if isinstance(w, QLabel):
                ws.cell(row=row,column=1,value=w.text()).font=Font(bold=True)
                row+=2

            if isinstance(w, QTableWidget):
                tabla = w

                # Encabezados
                for c in range(tabla.columnCount()):
                    ws.cell(row=row,column=c+1,
                        value=tabla.horizontalHeaderItem(c).text())
                row+=1

                # Datos
                for r2 in range(tabla.rowCount()):
                    for c2 in range(tabla.columnCount()):
                        it=tabla.item(r2,c2)
                        ws.cell(row=row,column=c2+1,
                            value=it.text() if it else "")
                    row+=1

                row+=2

        wb.save(ruta)

    # ================================================================
    # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
    # EXPORTACIÓN EXCEL con formato contable
    # ================================================================
    def _exportar_excel_reporte_por_cuentas(self, ruta):
        datos = self._obtener_datos_reporte_por_cuentas()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Cuentas"

        headers = ["Cuenta", "Fecha", "Concepto", "Debe", "Haber", "Estado", "Documento"]
        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        row = 1
        ini = self.fecha_ini.date().toString("dd/MM/yyyy")
        fin = self.fecha_fin.date().toString("dd/MM/yyyy")
        texto_cuenta = self.cbo_cuenta.currentText()
        titulo = f"Reporte por Cuentas — {ini} a {fin} — {texto_cuenta}"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws.cell(row=row, column=1, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = morado
        cell.alignment = Alignment(horizontal="center")
        row += 2

        if not datos:
            ws.cell(row=row, column=1, value="Sin datos para el rango/cuenta seleccionado.")
            wb.save(ruta)
            return

        col_widths = [len(h) for h in headers]

        for cuenta, nombre, movs in datos:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=f"{cuenta} - {nombre}")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = morado
            row += 2

            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = morado
                cell.border = borde
            row += 1

            total_debe = 0.0
            total_haber = 0.0
            invertir = self._modo_invertido_bi()

            for _, m in movs:
                d_orig = float(m.get("debe", 0) or 0)
                h_orig = float(m.get("haber", 0) or 0)
                
                if invertir:
                    debe = h_orig
                    haber = d_orig
                else:
                    debe = d_orig
                    haber = h_orig
                    
                total_debe += debe
                total_haber += haber

                fila = [
                    cuenta,
                    m.get("fecha", ""),
                    m.get("concepto", ""),
                    debe,
                    haber,
                    m.get("estado", ""),
                    m.get("documento", ""),
                ]

                for c, val in enumerate(fila, start=1):
                    cell = ws.cell(row=row, column=c, value=val)
                    cell.border = borde
                    if c in (4, 5):
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#,##0.00"
                    col_widths[c - 1] = max(col_widths[c - 1], len(str(val or "")))
                row += 1

            # Total por cuenta
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = borde
                cell.fill = verde
            ws.cell(row=row, column=3, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=4, value=total_debe).number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=total_haber).number_format = "#,##0.00"
            row += 3

        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 12), 60)

        wb.save(ruta)

    def _datos_flujo_caja_mensual(self):
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()
        invertir = self.chk_shillong.isChecked()

        resumen = defaultdict(
            lambda: {
                "entra_banco": 0.0,
                "entra_caja": 0.0,
                "sale_banco": 0.0,
                "sale_caja": 0.0,
            }
        )
        detalle_bancos = defaultdict(lambda: {"entra": 0.0, "sale": 0.0})

        for m in self._movimientos_en_rango(ini, fin):
            fecha_obj = self._parsear_fecha_movimiento(m.get("fecha", ""))
            if fecha_obj is None:
                continue

            debe = float(m.get("debe", 0) or 0)
            haber = float(m.get("haber", 0) or 0)
            if invertir:
                entra = debe
                sale = haber
            else:
                entra = haber
                sale = debe

            clave = (fecha_obj.year, fecha_obj.month)
            banco_nombre = str(m.get("banco", "") or "Caja").strip() or "Caja"
            if self._es_caja(m.get("banco", "")):
                resumen[clave]["entra_caja"] += entra
                resumen[clave]["sale_caja"] += sale
            else:
                resumen[clave]["entra_banco"] += entra
                resumen[clave]["sale_banco"] += sale
            detalle_bancos[(clave, banco_nombre)]["entra"] += entra
            detalle_bancos[(clave, banco_nombre)]["sale"] += sale

        filas = []
        acumulado = 0.0
        for year, month in sorted(resumen.keys()):
            entra_banco = round(resumen[(year, month)]["entra_banco"], 2)
            entra_caja = round(resumen[(year, month)]["entra_caja"], 2)
            sale_banco = round(resumen[(year, month)]["sale_banco"], 2)
            sale_caja = round(resumen[(year, month)]["sale_caja"], 2)
            flujo = round((entra_banco + entra_caja) - (sale_banco + sale_caja), 2)
            acumulado = round(acumulado + flujo, 2)
            filas.append(
                {
                    "mes": self._texto_mes(year, month),
                    "entra_banco": entra_banco,
                    "entra_caja": entra_caja,
                    "sale_banco": sale_banco,
                    "sale_caja": sale_caja,
                    "flujo": flujo,
                    "acumulado": acumulado,
                }
            )

        filas_bancos = []
        for (year_month, banco), valores in sorted(detalle_bancos.items(), key=lambda item: (item[0][0][0], item[0][0][1], item[0][1])):
            year, month = year_month
            filas_bancos.append(
                {
                    "mes": self._texto_mes(year, month),
                    "banco": banco,
                    "entra": round(valores["entra"], 2),
                    "sale": round(valores["sale"], 2),
                }
            )

        return filas, filas_bancos

    def _texto_mes(self, year, month):
        nombres = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        return f"{nombres[month - 1]} {year}"

    def _exportar_excel_flujo_caja(self, ruta):
        filas, filas_bancos = self._datos_flujo_caja_mensual()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        morado = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        azul_claro = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        headers = [
            "Mes",
            "Entra Banco",
            "Entra Caja",
            "Sale Banco",
            "Sale Caja",
            "Flujo Neto",
            "Acumulado",
        ]
        col_widths = [14, 16, 16, 16, 16, 16, 16]

        ini = self.fecha_ini.date().toString("dd/MM/yyyy")
        fin = self.fecha_fin.date().toString("dd/MM/yyyy")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        titulo = ws.cell(row=1, column=1, value="Cash Flow Mensual")
        titulo.font = Font(bold=True, color="FFFFFF", size=14)
        titulo.fill = morado
        titulo.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        subtitulo = ws.cell(row=2, column=1, value=f"Periodo: {ini} a {fin}")
        subtitulo.font = Font(italic=True, color="334155")
        subtitulo.fill = azul_claro
        subtitulo.alignment = Alignment(horizontal="center", vertical="center")

        row = 4
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = morado
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        total_entra_banco = 0.0
        total_entra_caja = 0.0
        total_sale_banco = 0.0
        total_sale_caja = 0.0
        total_flujo = 0.0
        acumulado_final = 0.0

        for fila in filas:
            valores = [
                fila["mes"],
                fila["entra_banco"],
                fila["entra_caja"],
                fila["sale_banco"],
                fila["sale_caja"],
                fila["flujo"],
                fila["acumulado"],
            ]
            for c, val in enumerate(valores, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = borde
                if c >= 2:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")
            total_entra_banco += fila["entra_banco"]
            total_entra_caja += fila["entra_caja"]
            total_sale_banco += fila["sale_banco"]
            total_sale_caja += fila["sale_caja"]
            total_flujo += fila["flujo"]
            acumulado_final = fila["acumulado"]
            row += 1

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = borde
            cell.fill = verde

        totales = [
            "TOTAL",
            round(total_entra_banco, 2),
            round(total_entra_caja, 2),
            round(total_sale_banco, 2),
            round(total_sale_caja, 2),
            round(total_flujo, 2),
            round(acumulado_final, 2),
        ]
        for c, val in enumerate(totales, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = Font(bold=True)
            cell.border = borde
            cell.fill = verde
            if c >= 2:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A5"
        row += 3
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        subtitulo_bancos = ws.cell(row=row, column=1, value="Desglose por Banco")
        subtitulo_bancos.font = Font(bold=True, color="FFFFFF")
        subtitulo_bancos.fill = morado
        subtitulo_bancos.alignment = Alignment(horizontal="center", vertical="center")
        row += 2

        headers_bancos = ["Mes", "Banco", "Entra", "Sale", "Flujo Neto", "Acumulado"]
        for c, h in enumerate(headers_bancos, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = morado
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        acumulado_por_banco = defaultdict(float)
        for fila in filas_bancos:
            flujo = round(fila["entra"] - fila["sale"], 2)
            acumulado_por_banco[fila["banco"]] = round(acumulado_por_banco[fila["banco"]] + flujo, 2)
            valores = [
                fila["mes"],
                fila["banco"],
                fila["entra"],
                fila["sale"],
                flujo,
                acumulado_por_banco[fila["banco"]],
            ]
            for c, val in enumerate(valores, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = borde
                if c >= 3:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        for idx, width in enumerate([14, 24, 16, 16, 16, 16], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(ws.column_dimensions[get_column_letter(idx)].width or 0, width)

        wb.save(ruta)

    # ================================================================
    # FEATURE v3.8.0 — Reporte por Cuentas (Modelo Sisters)
    # EXPORTACIÓN XLSX basada en plantilla
    # ================================================================
    def _exportar_excel_modelo_sisters(self, ruta):
        # Programmatic export: prefer generating the report programmatically
        # using the calculated totals (avoids relying on potentially stale template data).
        invertido = self._modo_invertido_bi()
        totales_exactos = self._totales_sisters_por_cuenta(invertido=invertido)

        # Build ordered lists of accounts per section (respecting layout order if available)
        def cuentas_seccion(prefijo):
            cuentas_layout = []
            try:
                wb = openpyxl.load_workbook(self._ruta_modelo_sisters_layout(), data_only=False)
                ws = wb[wb.sheetnames[0]]
                for r in range(1, ws.max_row + 1):
                    v = ws.cell(r, 3).value  # column C typically contains account codes
                    if v is None:
                        continue
                    if isinstance(v, (int, float)):
                        cta_s = str(int(v))
                    else:
                        cta_s = str(v).strip()
                    if cta_s.isdigit() and cta_s.startswith(prefijo):
                        cuentas_layout.append(cta_s)
            except Exception:
                cuentas_layout = []

            cuentas_plan = []
            for cta in getattr(self.data, "cuentas", {}).keys():
                cta_s = str(cta).strip()
                if cta_s.startswith(prefijo):
                    cuentas_plan.append(cta_s)

            cuentas_mov = []
            for cta in totales_exactos.keys():
                cta_s = str(cta).strip()
                if cta_s.startswith(prefijo):
                    cuentas_mov.append(cta_s)

            ordenadas = []
            vistos = set()
            for c in cuentas_layout:
                if c not in vistos:
                    ordenadas.append(c)
                    vistos.add(c)
            for c in sorted(set(cuentas_plan + cuentas_mov)):
                if c not in vistos:
                    ordenadas.append(c)
                    vistos.add(c)
            return ordenadas

        cuentas_by_section = {
            "6": cuentas_seccion("6"),
            "7": cuentas_seccion("7"),
            "2": cuentas_seccion("2"),
        }

        anio = self.fecha_ini.date().toPython().year
        periodo_str = f"{self.fecha_ini.date().toString('dd/MM/yyyy')} - {self.fecha_fin.date().toString('dd/MM/yyyy')}"

        # Try programmatic exporter if available
        if exportar_modelo_evolutivo:
            try:
                nombre_por_cuenta = getattr(self.data, "obtener_nombre_cuenta", None)
                plantilla = self._ruta_modelo_sisters_layout()
                exportar_modelo_evolutivo(
                    ruta, totales_exactos, cuentas_by_section,
                    anio, periodo_str, nombre_por_cuenta,
                    ruta_plantilla=plantilla
                )
                return
            except Exception:
                pass

        # Fallback: use template-based fill (legacy behaviour)
        plantilla = self._ruta_modelo_sisters_layout()
        try:
            wb = openpyxl.load_workbook(plantilla, data_only=False)
            ws = wb[wb.sheetnames[0]]
        except Exception as e:
            QMessageBox.critical(self, "Exportación", f"No se pudo cargar la plantilla:\n{plantilla}\n\n{e}")
            return

        # legacy: fill dynamic cells in the template (kept for compatibility)
        codigos_disponibles = set(str(k).strip() for k in getattr(self.data, "cuentas", {}).keys())
        codigos_disponibles.update(str(k).strip() for k in totales_exactos.keys())
        sheet_codes = []
        candidate_codes = set(codigos_disponibles)
        for r in range(1, ws.max_row + 1):
            raw_code = ws.cell(r, 3).value
            norm_code = self._normalizar_codigo_template(raw_code, codigos_disponibles)
            if norm_code:
                sheet_codes.append((raw_code, norm_code))
                candidate_codes.add(norm_code)

        # columnas por defecto (compatibilidad con versiones antiguas del layout)
        def find_col(ws, keyword_list, fallback=None):
            try:
                for hr in range(1, min(40, ws.max_row) + 1):
                    vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else "" for c in range(1, ws.max_column + 1)]
                    for key in vals:
                        for kw in keyword_list:
                            if kw in key:
                                return vals.index(key) + 1
            except Exception:
                pass
            return fallback

        code_col = find_col(ws, ["CUENTA"], fallback=3)
        banco_col = find_col(ws, ["BANCO"], fallback=6)
        caja_col = find_col(ws, ["CAJA"], fallback=7)
        prev_col = find_col(ws, ["MESES", "MESES ANTERIORES", "MESES ANTER"], fallback=8)
        total_col = find_col(ws, ["MES CORRIENTE", "GASTOS", "INGRESOS", "INVERSIONES", "SUMA DE", "SUMA"], fallback=5)
        presupuesto_col = find_col(ws, ["PRESUPUESTO"], fallback=None)
        diferencia_col = find_col(ws, ["DIFERENCIA"], fallback=None)

        # Fill template dynamically (same logic as before)
        for r in range(1, ws.max_row + 1):
            try:
                v = ws.cell(r, code_col).value if code_col else ws.cell(r, 3).value
            except Exception:
                v = ws.cell(r, 3).value
            if v is None:
                continue

            cuenta, _ = self._parsear_codigo_template(v, codigos_disponibles)
            if not cuenta or not cuenta.isdigit() or cuenta[0] not in ("6", "7", "2"):
                continue
            mode = self._resolver_modo_codigo_template(v, cuenta, sheet_codes, candidate_codes)

            vals = self._rollup_sisters(cuenta, totales_exactos, mode=mode)
            cur_banco = float(vals.get("cur_banco", 0) or 0)
            cur_caja = float(vals.get("cur_caja", 0) or 0)
            cur_total = cur_banco + cur_caja
            prev_total = float(vals.get("prev_total", 0) or 0)

            if banco_col:
                ws.cell(r, banco_col).value = round(cur_banco, 2); ws.cell(r, banco_col).number_format = "#,##0.00"
            if caja_col:
                ws.cell(r, caja_col).value = round(cur_caja, 2); ws.cell(r, caja_col).number_format = "#,##0.00"
            if prev_col:
                ws.cell(r, prev_col).value = round(prev_total, 2); ws.cell(r, prev_col).number_format = "#,##0.00"
            if total_col:
                ws.cell(r, total_col).value = round(cur_total, 2); ws.cell(r, total_col).number_format = "#,##0.00"
            if presupuesto_col:
                if not self._celda_es_formula(ws.cell(r, presupuesto_col)):
                    ws.cell(r, presupuesto_col).value = 0.0; ws.cell(r, presupuesto_col).number_format = "#,##0.00"
            if diferencia_col:
                try:
                    presupuesto_val = float(ws.cell(r, presupuesto_col).value or 0) if presupuesto_col else 0.0
                except Exception:
                    presupuesto_val = 0.0
                diferencia_val = round(presupuesto_val - (cur_total + prev_total), 2)
                if not self._celda_es_formula(ws.cell(r, diferencia_col)):
                    ws.cell(r, diferencia_col).value = diferencia_val; ws.cell(r, diferencia_col).number_format = "#,##0.00"

        try:
            wb.save(ruta)
        except Exception as e:
            QMessageBox.critical(self, "Exportación", f"No se pudo guardar el archivo:\n{ruta}\n\n{e}")

    # ================================================================
    # FEATURE v3.8.0 — Reporte por Cuentas (Informes BI)
    # IMPRESIÓN / PREVIEW (menú contextual)
    # ================================================================
    def _html_vista_actual(self):
        html = """
        <html><head><meta charset="utf-8">
        <style>
        body { font-family: Segoe UI, Arial, sans-serif; font-size: 10pt; }
        .title { background:#7030A0; color:white; padding:8px; font-weight:bold; }
        table { border-collapse: collapse; width: 100%; margin-top: 6px; }
        th { background:#7030A0; color:white; padding:6px; border:1px solid #000; text-align:left; }
        td { padding:5px; border:1px solid #000; }
        td.num { text-align: right; }
        </style></head><body>
        """
        for i in range(self.contenedor_layout.count()):
            w = self.contenedor_layout.itemAt(i).widget()
            if isinstance(w, QLabel):
                html += f"<div class='title'>{w.text()}</div>"
            if isinstance(w, QTableWidget):
                tabla = w
                html += "<table><thead><tr>"
                for c in range(tabla.columnCount()):
                    h = tabla.horizontalHeaderItem(c).text() if tabla.horizontalHeaderItem(c) else ""
                    html += f"<th>{h}</th>"
                html += "</tr></thead><tbody>"
                for r in range(tabla.rowCount()):
                    html += "<tr>"
                    for c in range(tabla.columnCount()):
                        it = tabla.item(r, c)
                        v = it.text() if it else ""
                        h = tabla.horizontalHeaderItem(c).text() if tabla.horizontalHeaderItem(c) else ""
                        cls = "num" if h in ("Debe", "Haber", "Saldo") else ""
                        html += f"<td class='{cls}'>{v}</td>"
                    html += "</tr>"
                html += "</tbody></table>"
        html += "</body></html>"
        return html

    def _seleccion_impresion_cash_flow(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Impresión Cash Flow")
        msg.setText("Selecciona el formato para imprimir el cash flow.")
        btn_global = msg.addButton("Global", QMessageBox.AcceptRole)
        btn_bancos = msg.addButton("Con todos los bancos", QMessageBox.AcceptRole)
        btn_cancel = msg.addButton(QMessageBox.Cancel)
        msg.exec()

        clicked = msg.clickedButton()
        if clicked == btn_global:
            return "global"
        if clicked == btn_bancos:
            return "bancos"
        return None

    def _html_flujo_caja(self, modo):
        resumen, detalle_bancos = self._datos_flujo_caja_mensual()
        if not resumen:
            return "<html><body><p>Sin datos para el rango seleccionado.</p></body></html>"

        def render_table(headers, rows):
            html = "<table><thead><tr>"
            for h in headers:
                html += f"<th>{h}</th>"
            html += "</tr></thead><tbody>"
            for row in rows:
                html += "<tr>"
                for idx, val in enumerate(row):
                    cls = "num" if idx >= 2 or (modo == "global" and idx >= 1) else ""
                    html += f"<td class='{cls}'>{val}</td>"
                html += "</tr>"
            html += "</tbody></table>"
            return html

        html = """
        <html><head><meta charset="utf-8">
        <style>
        body { font-family: Segoe UI, Arial, sans-serif; font-size: 10pt; }
        .title { background:#7030A0; color:white; padding:8px; font-weight:bold; margin-bottom:8px; }
        .subtitle { background:#DCE6F1; color:#334155; padding:6px; margin-bottom:8px; }
        table { border-collapse: collapse; width: 100%; margin-top: 6px; margin-bottom: 16px; }
        th { background:#7030A0; color:white; padding:6px; border:1px solid #000; text-align:left; }
        td { padding:5px; border:1px solid #000; }
        td.num { text-align: right; }
        </style></head><body>
        """
        periodo = f"{self.fecha_ini.date().toString('dd/MM/yyyy')} a {self.fecha_fin.date().toString('dd/MM/yyyy')}"
        html += "<div class='title'>Cash Flow Mensual</div>"
        html += f"<div class='subtitle'>Periodo: {periodo}</div>"

        if modo == "global":
            headers = ["Mes", "Entra Banco", "Entra Caja", "Sale Banco", "Sale Caja", "Flujo Neto", "Acumulado"]
            rows = []
            for fila in resumen:
                rows.append([
                    fila["mes"],
                    f'{fila["entra_banco"]:.2f}',
                    f'{fila["entra_caja"]:.2f}',
                    f'{fila["sale_banco"]:.2f}',
                    f'{fila["sale_caja"]:.2f}',
                    f'{fila["flujo"]:.2f}',
                    f'{fila["acumulado"]:.2f}',
                ])
            html += render_table(headers, rows)
        else:
            headers = ["Mes", "Banco", "Entra", "Sale", "Flujo Neto", "Acumulado"]
            rows = []
            acumulado_por_banco = defaultdict(float)
            for fila in detalle_bancos:
                flujo = round(fila["entra"] - fila["sale"], 2)
                acumulado_por_banco[fila["banco"]] = round(acumulado_por_banco[fila["banco"]] + flujo, 2)
                rows.append([
                    fila["mes"],
                    fila["banco"],
                    f'{fila["entra"]:.2f}',
                    f'{fila["sale"]:.2f}',
                    f'{flujo:.2f}',
                    f'{acumulado_por_banco[fila["banco"]]:.2f}',
                ])
            html += render_table(headers, rows)

        html += "</body></html>"
        return html

    def _regenerar_si_modelo_sisters(self):
        if self.cbo_tipo.currentIndex() == 5:
            self._limpiar_vista()
            self._mostrar_reporte_modelo_sisters()

    def _previsualizar_impresion(self):
        from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
        self._regenerar_si_modelo_sisters()
        printer = QPrinter(QPrinter.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        if self.cbo_tipo.currentIndex() == 6:
            modo = self._seleccion_impresion_cash_flow()
            if not modo:
                return
            html = self._html_flujo_caja(modo)
        else:
            html = self._html_vista_actual()
        doc = QTextDocument()
        doc.setHtml(html)
        preview.paintRequested.connect(lambda p: doc.print_(p))
        preview.exec()

    def _imprimir(self):
        from PySide6.QtPrintSupport import QPrinter, QPrintDialog
        self._regenerar_si_modelo_sisters()
        printer = QPrinter(QPrinter.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec():
            if self.cbo_tipo.currentIndex() == 6:
                modo = self._seleccion_impresion_cash_flow()
                if not modo:
                    return
                html = self._html_flujo_caja(modo)
            else:
                html = self._html_vista_actual()
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print_(printer)

    # ================================================================
    # EXPORTAR LIBRO MAYOR PROFESIONAL SHILLONG
    # ================================================================
    def _exportar_libro_mayor(self):
        ruta,_= QFileDialog.getSaveFileName(
            self,"Exportar Libro Mayor SHILLONG",
            self._ruta_exporte_por_defecto(),"Excel (*.xlsx)"
        )
        if not ruta:
            return

        invertir = self.chk_shillong.isChecked()
        wb=openpyxl.Workbook()
        ws=wb.active

        morado=PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        borde=Border(
            left=Side(style="thin",color="000000"),
            right=Side(style="thin",color="000000"),
            top=Side(style="thin",color="000000"),
            bottom=Side(style="thin",color="000000")
        )

        row=1
        cuentas=sorted(self.data.cuentas.keys())

        for cta in cuentas:
            movs=self.data.movimientos_por_cuenta(cta)
            if not movs:
                continue

            # Ordenar por fecha igual que en la vista (si el combo existe)
            orden = getattr(self, "cbo_orden_mayor", None)
            orden_idx = orden.currentIndex() if orden else 0
            if orden_idx in (0, 1):
                def _parse_f(f):
                    try:
                        return datetime.date.fromisoformat(f)
                    except Exception:
                        return datetime.date.min
                movs = sorted(movs, key=lambda m: _parse_f(m.get("fecha", "")), reverse=(orden_idx == 1))

            nombre=self.data.cuentas[cta].get("nombre","")

            # ENCABEZADO CUENTA
            cell=ws.cell(row=row,column=1,value=f"{cta} — {nombre}")
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=morado
            row+=2

            headers=["Fecha","Documento","Desglose","Entra" if invertir else "Debe","Sale" if invertir else "Haber","Saldo"]
            for c,h in enumerate(headers, start=1):
                cell=ws.cell(row=row,column=c,value=h)
                cell.font=Font(bold=True,color="FFFFFF")
                cell.fill=morado
                cell.border=borde
            row+=1

            saldo_acum=0
            total_debe=0 # Realmente 'Entra' si invertido
            total_haber=0 # Realmente 'Sale' si invertido

            for m in movs:
                concepto=m.get("concepto","").strip()
                if concepto:
                    des=concepto
                else:
                    des=m.get("nombre_cuenta","")

                d_orig=float(m.get("debe",0))
                h_orig=float(m.get("haber",0))

                if invertir:
                    debe = h_orig
                    haber = d_orig
                    saldo_acum += (debe - haber)
                else:
                    debe = d_orig
                    haber = h_orig
                    saldo_acum += (haber - debe)

                total_debe+=debe
                total_haber+=haber

                fila=[
                    m.get("fecha",""),
                    m.get("documento",""),
                    des,
                    debe,
                    haber,
                    saldo_acum
                ]

                for c,val in enumerate(fila,start=1):
                    cell=ws.cell(row=row,column=c,value=val)
                    cell.border=borde
                    if c>=4:
                        cell.alignment=Alignment(horizontal="right")
                        cell.number_format = "#,##0.00"
                row+=1

            # TOTAL
            for c in range(1,7):
                cell=ws.cell(row=row,column=c)
                cell.border=borde
                cell.fill=verde

            ws.cell(row=row,column=3,value="TOTAL").font=Font(bold=True)
            ws.cell(row=row,column=4,value=total_debe).number_format = "#,##0.00"
            ws.cell(row=row,column=5,value=total_haber).number_format = "#,##0.00"
            ws.cell(row=row,column=6,value=saldo_acum).number_format = "#,##0.00"

            row+=3

        wb.save(ruta)

    # ================================================================
    # EXPORTAR BALANCE SHILLONG (SUMAS & SALDOS)
    # ================================================================
    def _exportar_balance(self):
        ruta,_= QFileDialog.getSaveFileName(
            self,"Exportar Balance SHILLONG",
            self._ruta_exporte_por_defecto(),"Excel (*.xlsx)"
        )
        if not ruta:
            return

        invertir = self.chk_shillong.isChecked()
        wb=openpyxl.Workbook()
        ws=wb.active

        morado=PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        borde=Border(
            left=Side(style="thin",color="000000"),
            right=Side(style="thin",color="000000"),
            top=Side(style="thin",color="000000"),
            bottom=Side(style="thin",color="000000")
        )

        headers=["Cuenta","Nombre","Entra" if invertir else "Debe","Sale" if invertir else "Haber","Saldo"]
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()

        row=1
        for c,h in enumerate(headers,start=1):
            cell=ws.cell(row=row,column=c,value=h)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=morado
            cell.border=borde
        row+=1

        resumen=defaultdict(lambda:{"nombre":"", "debe":0, "haber":0})

        for m in self._movimientos_en_rango(ini, fin):
            cta=str(m.get("cuenta",""))
            resumen[cta]["nombre"]=m.get("nombre_cuenta","")
            d_orig = float(m.get("debe",0))
            h_orig = float(m.get("haber",0))
            
            if invertir:
                resumen[cta]["debe"] += h_orig
                resumen[cta]["haber"] += d_orig
            else:
                resumen[cta]["debe"] += d_orig
                resumen[cta]["haber"] += h_orig

        total_debe=0
        total_haber=0

        for cta in sorted(resumen.keys()):
            d=resumen[cta]
            if invertir:
                saldo = d["debe"] - d["haber"]
            else:
                saldo = d["haber"] - d["debe"]

            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila,start=1):
                cell=ws.cell(row=row,column=c,value=val)
                cell.border=borde
                if c>=3:
                    cell.alignment=Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"

            total_debe+=d["debe"]
            total_haber+=d["haber"]
            row+=1

        # TOTAL GENERAL
        for c in range(1,6):
            cell=ws.cell(row=row,column=c)
            cell.border=borde
            cell.fill=verde

        ws.cell(row=row,column=2,value="TOTAL GENERAL").font=Font(bold=True)
        ws.cell(row=row,column=3,value=total_debe).number_format = "#,##0.00"
        ws.cell(row=row,column=4,value=total_haber).number_format = "#,##0.00"
        ws.cell(row=row,column=5,value=(total_debe-total_haber if invertir else total_haber-total_debe)).number_format = "#,##0.00"

        wb.save(ruta)
