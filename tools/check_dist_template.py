#!/usr/bin/env python3
import openpyxl
from pathlib import Path

p = Path('dist/SHILLONG_v3_PRO/_internal/ui/Modelo evolutivo presupuestario 26.xlsx')
if not p.exists():
    print('TEMPLATE NOT FOUND:', p)
    raise SystemExit(2)

wb = openpyxl.load_workbook(p, data_only=False)
out_rows = []
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print('Sheet:', sheetname, 'max_row=', ws.max_row, 'max_col=', ws.max_column)
    # find header row
    header_row = None
    max_check = min(40, ws.max_row)
    for hr in range(1, max_check+1):
        vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else '' for c in range(1, ws.max_column+1)]
        if any('CUENTA' in v for v in vals):
            header_row = hr
            break
    print('  header_row=', header_row)
    start_data = (header_row + 2) if header_row else 8

    merged = list(ws.merged_cells.ranges)
    if merged:
        print('  merged ranges count:', len(merged))
    # scan rows for non-formula non-empty cells in data area
    issues = []
    for r in range(start_data, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(r, c)
            v = getattr(cell, 'value', None)
            if v is None:
                continue
            is_formula = (isinstance(v, str) and v.startswith('=')) or (getattr(cell, 'data_type', None) == 'f')
            if not is_formula:
                issues.append((r, c, v, type(v).__name__))
    print('  non-formula non-empty cells found:', len(issues))
    for idx, it in enumerate(issues[:200]):
        r,c,v,t = it
        print(f'    R{r} C{c}: ({t}) {v}')
    if len(issues) > 200:
        print('    ... plus', len(issues)-200, 'more')

print('Done')
