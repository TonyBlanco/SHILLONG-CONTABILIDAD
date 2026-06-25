#!/usr/bin/env python3
"""Strip sample data from the Modelo evolutivo template and write cleaned copy to dist/_internal/ui/"""
import openpyxl, sys
from pathlib import Path

src = Path('ui/Modelo evolutivo presupuestario 26.xlsx')
blank_src = Path('ui/Modelo evolutivo presupuestario 26 - blank.xlsx')
dst = Path('dist/SHILLONG_v3_PRO/_internal/ui/Modelo evolutivo presupuestario 26.xlsx')
if blank_src.exists():
    # Prefer explicit blank template if provided
    wb = openpyxl.load_workbook(blank_src, data_only=False)
    ws = wb[wb.sheetnames[0]]
else:
    # If blank missing, generate it from source by clearing non-formula sample data
    if not src.exists():
        print('Source template not found:', src)
        sys.exit(2)
    wb = openpyxl.load_workbook(src, data_only=False)
    ws = wb[wb.sheetnames[0]]
    # find header row
    header_row = None
    max_check = min(40, ws.max_row)
    for hr in range(1, max_check + 1):
        vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else "" for c in range(1, ws.max_column + 1)]
        if any("CUENTA" in v for v in vals):
            header_row = hr
            break
    if header_row is None:
        header_row = 6
    start_data = header_row + 2
    try:
        merged = list(ws.merged_cells.ranges)
        for m in merged:
            ws.unmerge_cells(str(m))
    except Exception:
        pass

    # Clear non-formula cells individually (do not skip whole rows that contain formulas)
    cleared_local = 0
    max_col = ws.max_column
    for r in range(start_data, ws.max_row + 1):
        code_cell = ws.cell(r, 3)
        if code_cell.value is None:
            continue
        for c in range(3, min(10, max_col + 1)):
            cell = ws.cell(r, c)
            if cell is None:
                continue
            v = getattr(cell, 'value', None)
            if v is None:
                continue
            # Preserve formulas (string starting with '=' or data_type 'f')
            if isinstance(v, str) and v.startswith('='):
                continue
            if getattr(cell, 'data_type', None) == 'f':
                continue
            try:
                cell.value = None
                cleared_local += 1
            except Exception:
                pass
    # save the generated blank template for future builds
    try:
        wb.save(blank_src)
        print(f'Generated blank template: {blank_src} (cleared {cleared_local} cells)')
    except Exception as e:
        print('Warning: could not save blank template:', e)

# Unmerge merged cells to allow clearing cell values safely
try:
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))
except Exception:
    pass

# Find header row similar to InformesView logic
header_row = None
header_map = {}
max_check = min(40, ws.max_row)
for hr in range(1, max_check+1):
    vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else "" for c in range(1, ws.max_column+1)]
    if any('CUENTA' in v for v in vals):
        header_row = hr
        for c, v in enumerate(vals, start=1):
            if not v:
                continue
            header_map[v] = c
        break

if header_row is None:
    print('Header row not found, aborting.')
    sys.exit(3)

start_data = header_row + 2
cleared = 0
max_col = ws.max_column
for r in range(start_data, ws.max_row+1):
    code_cell = ws.cell(r, 3)
    if code_cell.value is None:
        continue
    # Clear non-formula values per-cell (preserve formulas)
    for c in range(3, min(10, max_col+1)):
        cell = ws.cell(r, c)
        if cell is None:
            continue
        v = getattr(cell, 'value', None)
        if v is None:
            continue
        if isinstance(v, str) and v.startswith('='):
            continue
        if getattr(cell, 'data_type', None) == 'f':
            continue
        cell.value = None
        cleared += 1

# Ensure destination folder exists
dst.parent.mkdir(parents=True, exist_ok=True)
wb.save(dst)
print(f'Cleaned template saved to {dst} (cleared {cleared} cells)')
