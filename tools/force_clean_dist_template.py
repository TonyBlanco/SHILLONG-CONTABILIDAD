#!/usr/bin/env python3
import openpyxl
from pathlib import Path
import shutil

p = Path('dist/SHILLONG_v3_PRO/_internal/ui/Modelo evolutivo presupuestario 26.xlsx')
if not p.exists():
    print('TEMPLATE NOT FOUND:', p)
    raise SystemExit(2)

backup = p.with_suffix('.pre_clean.xlsx')
shutil.copy2(p, backup)
print('Backup saved to', backup)

wb = openpyxl.load_workbook(p, data_only=False)
cleared = 0
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    # find header row
    header_row = None
    max_check = min(40, ws.max_row)
    for hr in range(1, max_check+1):
        vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else '' for c in range(1, ws.max_column+1)]
        if any('CUENTA' in v for v in vals):
            header_row = hr
            break
    start_data = (header_row + 2) if header_row else 8
    print(f'Processing sheet {sheetname}, header_row={header_row}, start_data={start_data}, max_row={ws.max_row}, max_col={ws.max_column}')
    # clear non-formula cells from column 3 onwards (keep account code/name)
    for r in range(start_data, ws.max_row+1):
        for c in range(3, ws.max_column+1):
            cell = ws.cell(r, c)
            v = getattr(cell, 'value', None)
            if v is None:
                continue
            is_formula = (isinstance(v, str) and v.startswith('=')) or (getattr(cell, 'data_type', None) == 'f')
            if not is_formula:
                cell.value = None
                cleared += 1

wb.save(p)
print('Saved cleaned template to', p)
print('Cleared cells:', cleared)

# copy cleaned template to ui/ as blank template
ui_dst = Path('ui') / 'Modelo evolutivo presupuestario 26 - blank.xlsx'
ui_dst.parent.mkdir(parents=True, exist_ok=True)
shutil.copy2(p, ui_dst)
print('Copied cleaned template to', ui_dst)
