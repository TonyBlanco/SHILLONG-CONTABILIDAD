#!/usr/bin/env python3
"""Create a blank template preserving headers and formulas but removing sample data.
Writes: ui/Modelo evolutivo presupuestario 26 - blank.xlsx
Also writes copy to dist/_internal/ui path if dist exists.
"""
import openpyxl
from pathlib import Path
import sys

src = Path('ui/Modelo evolutivo presupuestario 26.xlsx')
blank = Path('ui/Modelo evolutivo presupuestario 26 - blank.xlsx')
copy_to_dist = Path('dist/SHILLONG_v3_PRO/_internal/ui/Modelo evolutivo presupuestario 26.xlsx')

if not src.exists():
    print('Source template not found:', src)
    sys.exit(2)

wb = openpyxl.load_workbook(src, data_only=False)
ws = wb[wb.sheetnames[0]]

# Find header row
header_row = None
max_check = min(40, ws.max_row)
for hr in range(1, max_check+1):
    vals = [str(ws.cell(hr, c).value).strip().upper() if ws.cell(hr, c).value is not None else "" for c in range(1, ws.max_column+1)]
    if any('CUENTA' in v for v in vals):
        header_row = hr
        break

if header_row is None:
    print('Header row not found; using default start row 8')
    header_row = 6

start_data = header_row + 2

# Unmerge to allow writes
try:
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))
except Exception:
    pass

cleared = 0
for r in range(start_data, ws.max_row+1):
    # Determine if row has code
    code = ws.cell(r, 3).value
    if code is None:
        continue
    # Clear non-formula values per-cell (preserve formulas)
    for c in range(3, ws.max_column+1):
        cell = ws.cell(r, c)
        v = getattr(cell, 'value', None)
        if v is None:
            continue
        # Preserve formulas: string starting with '=' or data_type 'f'
        if isinstance(v, str) and v.startswith('='):
            continue
        if getattr(cell, 'data_type', None) == 'f':
            continue
        cell.value = None
        cleared += 1

# Save blank template
wb.save(blank)
print(f'Blank template created: {blank} (cleared {cleared} cells)')

# Also copy into dist path if it exists
if copy_to_dist.parent.exists():
    # create a fresh workbook from blank and save to dist path
    wb2 = openpyxl.load_workbook(blank, data_only=False)
    copy_to_dist.parent.mkdir(parents=True, exist_ok=True)
    wb2.save(copy_to_dist)
    print(f'Copied blank template to {copy_to_dist}')

print('Done')
