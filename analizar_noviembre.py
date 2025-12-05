# -*- coding: utf-8 -*-
"""
Script de análisis del Excel de Noviembre (review)
Extrae información crítica para configurar el sistema
"""

import openpyxl
from datetime import datetime
from collections import defaultdict

def analizar_libro_noviembre(ruta_excel):
    """
    Analiza el Excel de Noviembre y extrae:
    1. Estructura de columnas
    2. Saldos iniciales por banco
    3. Total de movimientos
    4. Saldos finales por banco
    5. Totales Debe/Haber
    """
    
    print("="*70)
    print("📊 ANÁLISIS DEL LIBRO DE NOVIEMBRE 2024")
    print("="*70)
    
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        ws = wb.active
        print(f"✅ Archivo cargado: {ws.title}")
    except Exception as e:
        print(f"❌ Error al abrir archivo: {e}")
        return None
    
    # ============================================================
    # 1. ANALIZAR ESTRUCTURA DE COLUMNAS
    # ============================================================
    print("\n" + "="*70)
    print("1️⃣  ESTRUCTURA DE COLUMNAS")
    print("="*70)
    
    headers = {}
    row1 = list(ws[1])
    
    for idx, cell in enumerate(row1, 1):
        if cell.value:
            col_name = str(cell.value).strip()
            headers[col_name] = idx
            print(f"   Columna {chr(64+idx)}: {col_name}")
    
    # Mapeo de nombres comunes
    col_map = {
        'fecha': ['Fecha', 'Date', 'FECHA'],
        'cuenta': ['Cuenta', 'Account', 'CUENTA', 'Cta'],
        'concepto': ['Concepto', 'Concept', 'CONCEPTO', 'Descripción'],
        'debe': ['Debe', 'Debit', 'DEBE', 'Gasto'],
        'haber': ['Haber', 'Credit', 'HABER', 'Ingreso'],
        'saldo': ['Saldo', 'Balance', 'SALDO'],
        'banco': ['Banco', 'Bank', 'BANCO', 'Caja'],
        'documento': ['Documento', 'Doc', 'DOCUMENTO', 'Ref']
    }
    
    # Detectar columnas automáticamente
    detected_cols = {}
    for key, variations in col_map.items():
        for var in variations:
            if var in headers:
                detected_cols[key] = headers[var]
                break
    
    print("\n📋 Columnas detectadas:")
    for key, col_idx in detected_cols.items():
        print(f"   {key.upper()}: Columna {chr(64+col_idx)}")
    
    # ============================================================
    # 2. BUSCAR SALDO INICIAL (primera fila de datos)
    # ============================================================
    print("\n" + "="*70)
    print("2️⃣  SALDOS INICIALES (Fila con 'Saldo inicial' o primera fila)")
    print("="*70)
    
    saldos_iniciales = {}
    fila_inicio_datos = 2
    
    # Buscar fila "Saldo inicial"
    for row_idx in range(2, min(10, ws.max_row + 1)):
        concepto_cell = ws.cell(row=row_idx, column=detected_cols.get('concepto', 3))
        if concepto_cell.value and 'saldo inicial' in str(concepto_cell.value).lower():
            # Extraer saldo inicial
            if 'saldo' in detected_cols:
                saldo_val = ws.cell(row=row_idx, column=detected_cols['saldo']).value
                banco_val = ws.cell(row=row_idx, column=detected_cols.get('banco', 8)).value or "Caja"
                
                try:
                    saldos_iniciales[banco_val] = float(saldo_val or 0)
                except:
                    saldos_iniciales[banco_val] = 0.0
                
                print(f"   ✅ {banco_val}: {saldos_iniciales[banco_val]:,.2f} INR")
                fila_inicio_datos = row_idx + 1
            break
    
    if not saldos_iniciales:
        print("   ⚠️  No se encontró fila 'Saldo inicial', usando valores por defecto")
        # Tomar el primer saldo como referencia
        if 'saldo' in detected_cols:
            primer_saldo = ws.cell(row=2, column=detected_cols['saldo']).value
            try:
                saldos_iniciales['Caja'] = float(primer_saldo or 0)
            except:
                saldos_iniciales['Caja'] = 0.0
    
    # ============================================================
    # 3. ANALIZAR MOVIMIENTOS
    # ============================================================
    print("\n" + "="*70)
    print("3️⃣  ANÁLISIS DE MOVIMIENTOS")
    print("="*70)
    
    movimientos = []
    total_debe = 0.0
    total_haber = 0.0
    movimientos_por_banco = defaultdict(lambda: {'debe': 0, 'haber': 0, 'count': 0})
    
    for row_idx in range(fila_inicio_datos, ws.max_row + 1):
        # Saltar filas vacías o de totales
        concepto = ws.cell(row=row_idx, column=detected_cols.get('concepto', 3)).value
        if not concepto:
            continue
        if 'total' in str(concepto).lower():
            break
        
        # Extraer datos
        try:
            fecha = ws.cell(row=row_idx, column=detected_cols.get('fecha', 1)).value
            cuenta = ws.cell(row=row_idx, column=detected_cols.get('cuenta', 2)).value
            debe = ws.cell(row=row_idx, column=detected_cols.get('debe', 4)).value or 0
            haber = ws.cell(row=row_idx, column=detected_cols.get('haber', 5)).value or 0
            saldo = ws.cell(row=row_idx, column=detected_cols.get('saldo', 7)).value or 0
            banco = ws.cell(row=row_idx, column=detected_cols.get('banco', 8)).value or "Caja"
            doc = ws.cell(row=row_idx, column=detected_cols.get('documento', 9)).value or ""
            
            # Convertir fecha si es datetime
            if isinstance(fecha, datetime):
                fecha_str = fecha.strftime("%d/%m/%Y")
            else:
                fecha_str = str(fecha)
            
            # Convertir números
            debe = float(debe) if debe else 0.0
            haber = float(haber) if haber else 0.0
            saldo = float(saldo) if saldo else 0.0
            
            mov = {
                'fecha': fecha_str,
                'documento': str(doc),
                'concepto': str(concepto),
                'cuenta': str(cuenta) if cuenta else "",
                'debe': debe,
                'haber': haber,
                'saldo': saldo,
                'banco': str(banco)
            }
            
            movimientos.append(mov)
            total_debe += debe
            total_haber += haber
            
            # Acumular por banco
            movimientos_por_banco[str(banco)]['debe'] += debe
            movimientos_por_banco[str(banco)]['haber'] += haber
            movimientos_por_banco[str(banco)]['count'] += 1
            
        except Exception as e:
            print(f"   ⚠️  Error en fila {row_idx}: {e}")
            continue
    
    print(f"\n📊 Total de movimientos procesados: {len(movimientos)}")
    print(f"   💰 Total DEBE (Ingresos en Excel): {total_debe:,.2f} INR")
    print(f"   💸 Total HABER (Gastos en Excel): {total_haber:,.2f} INR")
    print(f"   📈 Diferencia (Debe - Haber): {total_debe - total_haber:,.2f} INR")
    
    # ============================================================
    # 4. RESUMEN POR BANCO
    # ============================================================
    print("\n" + "="*70)
    print("4️⃣  RESUMEN POR BANCO")
    print("="*70)
    
    for banco, datos in sorted(movimientos_por_banco.items()):
        print(f"\n🏦 {banco}:")
        print(f"   Movimientos: {datos['count']}")
        print(f"   Debe (Ingresos): {datos['debe']:,.2f} INR")
        print(f"   Haber (Gastos): {datos['haber']:,.2f} INR")
        print(f"   Diferencia: {datos['debe'] - datos['haber']:,.2f} INR")
        
        # Calcular saldo final
        saldo_inicial = saldos_iniciales.get(banco, 0.0)
        saldo_final = saldo_inicial + datos['debe'] - datos['haber']
        print(f"   Saldo inicial: {saldo_inicial:,.2f} INR")
        print(f"   ➡️  Saldo final: {saldo_final:,.2f} INR")
    
    # ============================================================
    # 5. VALIDAR CONVENCIÓN DEBE/HABER
    # ============================================================
    print("\n" + "="*70)
    print("5️⃣  VALIDACIÓN DE CONVENCIÓN CONTABLE")
    print("="*70)
    
    # Analizar primeros 10 movimientos
    print("\n📝 Muestra de primeros movimientos:")
    for i, mov in enumerate(movimientos[:10], 1):
        print(f"\n   {i}. {mov['concepto'][:40]}")
        print(f"      Debe: {mov['debe']:,.2f} | Haber: {mov['haber']:,.2f}")
        print(f"      Cuenta: {mov['cuenta']} | Banco: {mov['banco']}")
    
    # Detectar convención
    gastos_en_haber = sum(1 for m in movimientos if m['haber'] > 0 and m['debe'] == 0)
    ingresos_en_debe = sum(1 for m in movimientos if m['debe'] > 0 and m['haber'] == 0)
    
    print(f"\n🔍 Análisis de convención:")
    print(f"   Movimientos con HABER>0 y DEBE=0: {gastos_en_haber}")
    print(f"   Movimientos con DEBE>0 y HABER=0: {ingresos_en_debe}")
    
    if gastos_en_haber > ingresos_en_debe:
        print("\n   ⚠️  CONVENCIÓN DETECTADA: INVERTIDA")
        print("   En este Excel:")
        print("   - DEBE = Ingresos (dinero entra)")
        print("   - HABER = Gastos (dinero sale)")
        print("\n   ✅ La app invertirá automáticamente al exportar")
    else:
        print("\n   ✅ CONVENCIÓN ESTÁNDAR")
        print("   - DEBE = Gastos")
        print("   - HABER = Ingresos")
    
    # ============================================================
    # 6. REPORTE FINAL
    # ============================================================
    print("\n" + "="*70)
    print("6️⃣  REPORTE FINAL - DATOS PARA CONFIGURACIÓN")
    print("="*70)
    
    print("\n📋 Saldos iniciales a configurar en la app:")
    print("```json")
    print("{")
    for banco, saldo in saldos_iniciales.items():
        print(f'  "{banco}": {saldo},')
    print("}")
    print("```")
    
    print("\n📊 Totales de Noviembre:")
    print(f"   Total ingresos (Debe): {total_debe:,.2f} INR")
    print(f"   Total gastos (Haber): {total_haber:,.2f} INR")
    print(f"   Saldo neto del mes: {total_debe - total_haber:,.2f} INR")
    
    print("\n✅ Análisis completado")
    print("="*70)
    
    return {
        'saldos_iniciales': saldos_iniciales,
        'movimientos': movimientos,
        'total_debe': total_debe,
        'total_haber': total_haber,
        'movimientos_por_banco': dict(movimientos_por_banco)
    }


if __name__ == "__main__":
    import os
    from pathlib import Path
    
    # Intentar encontrar el archivo automáticamente
    posibles_rutas = [
        r"D:\ShillongV3\Libro_Noviembre-review.xlsx",
        "Libro_Noviembre-review.xlsx",
        "Libro_Noviembrereview.xlsx",
        "../Libro_Noviembre-review.xlsx",
        "../../Libro_Noviembre-review.xlsx",
        Path.home() / "Downloads" / "Libro_Noviembre-review.xlsx",
        Path.home() / "Desktop" / "Libro_Noviembre-review.xlsx",
    ]
    
    ruta = None
    for path in posibles_rutas:
        if Path(path).exists():
            ruta = path
            print(f"✅ Archivo encontrado en: {path}\n")
            break
    
    if not ruta:
        print("❌ No se encontró el archivo automáticamente.")
        print("\nIngresa la ruta completa del archivo Excel:")
        print("Ejemplo Windows: C:\\Users\\TuUsuario\\Downloads\\Libro_Noviembrereview.xlsx")
        print("Ejemplo Mac/Linux: /Users/tuusuario/Downloads/Libro_Noviembrereview.xlsx")
        ruta = input("\nRuta: ").strip().strip('"').strip("'")
        
        if not Path(ruta).exists():
            print(f"\n❌ ERROR: No se encuentra el archivo en: {ruta}")
            print("\nVerifica que:")
            print("  1. La ruta esté correcta")
            print("  2. El archivo exista")
            print("  3. Tengas permisos de lectura")
            exit(1)
    
    resultado = analizar_libro_noviembre(ruta)
    
    if resultado:
        print("\n🎯 Próximos pasos:")
        print("   1. Verificar que los saldos iniciales coincidan con tus registros")
        print("   2. Confirmar la convención Debe/Haber")
        print("   3. Validar que los totales cuadren con tu contabilidad")
        print("   4. Proceder a implementar el sistema de saldos mensuales")