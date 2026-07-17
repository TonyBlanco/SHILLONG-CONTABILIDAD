import os
import sys
import unittest
import tempfile
from pathlib import Path
import openpyxl

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from PySide6.QtWidgets import QApplication

from models.ExportadorModeloEvolutivo import exportar_modelo_evolutivo, _parse_template_code, _resolve_template_mode, _rollup
from ui.InformesView import InformesView


class _MockData:
    def __init__(self):
        self.cuentas = {}
        self.movimientos = []


class TestModeloEvolutivoTemplateCodes(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def test_parse_template_code_distinguishes_group_and_exact(self):
        self.assertEqual(_parse_template_code("602-400"), ("602400", "group"))
        self.assertEqual(_parse_template_code("602.40"), ("602400", "exact"))

    def test_resolve_template_mode_uses_sheet_context(self):
        sheet_codes = [
            ("602400", "602400"),
            ("602.40", "602400"),
            ("602.401", "602401"),
            ("650.10", "650100"),
            ("650.11", "650110"),
            ("758.00", "758000"),
            ("758.1", "758100"),
            ("759", "759000"),
            ("759.10", "759100"),
        ]
        candidate_codes = {code for _, code in sheet_codes}

        self.assertEqual(_resolve_template_mode("602.40", "602400", sheet_codes, candidate_codes), "exact")
        self.assertEqual(_resolve_template_mode("650.10", "650100", sheet_codes, candidate_codes), "exact")
        self.assertEqual(_resolve_template_mode("758.00", "758000", sheet_codes, candidate_codes), "exact")
        self.assertEqual(_resolve_template_mode("759.10", "759100", sheet_codes, candidate_codes), "exact")

    def test_rollup_group_and_exact_do_not_collide(self):
        totales = {
            "602400": {"cur_banco": 7651.0, "cur_caja": 0.0, "prev_total": 0.0},
            "602401": {"cur_banco": 120.0, "cur_caja": 0.0, "prev_total": 0.0},
            "602403": {"cur_banco": 80.0, "cur_caja": 0.0, "prev_total": 0.0},
        }

        self.assertEqual(_rollup("602400", totales, mode="exact"), (7651.0, 0.0, 0.0))
        self.assertEqual(_rollup("602400", totales, mode="group"), (7851.0, 0.0, 0.0))

    def test_informes_view_uses_same_group_vs_exact_rule(self):
        view = InformesView(_MockData())
        codigos = {"602400", "602401", "602403", "650100", "650110", "758000", "758100", "759000", "759100"}
        totales = {
            "602400": {"cur_banco": 7651.0, "cur_caja": 0.0, "prev_total": 0.0},
            "602401": {"cur_banco": 120.0, "cur_caja": 0.0, "prev_total": 0.0},
            "602403": {"cur_banco": 80.0, "cur_caja": 0.0, "prev_total": 0.0},
            "650110": {"cur_banco": 2995.0, "cur_caja": 0.0, "prev_total": 0.0},
            "758100": {"cur_banco": 297852.0, "cur_caja": 0.0, "prev_total": 0.0},
            "759100": {"cur_banco": 6000.0, "cur_caja": 0.0, "prev_total": 0.0},
        }
        sheet_codes = [
            ("602400", "602400"),
            ("602.40", "602400"),
            ("602.401", "602401"),
            ("650.10", "650100"),
            ("650.11", "650110"),
            ("758.00", "758000"),
            ("758.1", "758100"),
            ("759", "759000"),
            ("759.10", "759100"),
        ]

        cuenta_grupo, mode_grupo = view._parsear_codigo_template("602-400", codigos)
        cuenta_exacta, mode_exacto = view._parsear_codigo_template("602.40", codigos)
        _, mode_65010 = view._parsear_codigo_template("650.10", codigos)
        _, mode_75800 = view._parsear_codigo_template("758.00", codigos)
        _, mode_75910 = view._parsear_codigo_template("759.10", codigos)

        mode_grupo = view._resolver_modo_codigo_template("602-400", cuenta_grupo, sheet_codes, codigos)
        mode_exacto = view._resolver_modo_codigo_template("602.40", cuenta_exacta, sheet_codes, codigos)
        mode_65010 = view._resolver_modo_codigo_template("650.10", "650100", sheet_codes, codigos)
        mode_75800 = view._resolver_modo_codigo_template("758.00", "758000", sheet_codes, codigos)
        mode_75910 = view._resolver_modo_codigo_template("759.10", "759100", sheet_codes, codigos)

        self.assertEqual(
            view._rollup_sisters(cuenta_grupo, totales, mode=mode_grupo),
            {"cur_banco": 7851.0, "cur_caja": 0.0, "prev_total": 0.0},
        )
        self.assertEqual(
            view._rollup_sisters(cuenta_exacta, totales, mode=mode_exacto),
            {"cur_banco": 7651.0, "cur_caja": 0.0, "prev_total": 0.0},
        )
        self.assertEqual(mode_65010, "exact")
        self.assertEqual(mode_75800, "exact")
        self.assertEqual(mode_75910, "exact")

    def test_programmatic_export_respects_template_sum_rows(self):
        totales = {
            "602400": {"cur_banco": 0.0, "cur_caja": 2202.0, "prev_total": 0.0},
            "602401": {"cur_banco": 0.0, "cur_caja": 4630.0, "prev_total": 0.0},
            "602402": {"cur_banco": 0.0, "cur_caja": 819.0, "prev_total": 0.0},
            "650300": {"cur_banco": 0.0, "cur_caja": 2000.0, "prev_total": 0.0},
            "650320": {"cur_banco": 0.0, "cur_caja": 1520.0, "prev_total": 0.0},
            "650330": {"cur_banco": 0.0, "cur_caja": 470.0, "prev_total": 0.0},
            "650500": {"cur_banco": 99800.0, "cur_caja": 78000.0, "prev_total": 0.0},
            "650510": {"cur_banco": 0.0, "cur_caja": 15000.0, "prev_total": 0.0},
            "650530": {"cur_banco": 0.0, "cur_caja": 200.0, "prev_total": 0.0},
            "650560": {"cur_banco": 0.0, "cur_caja": 4000.0, "prev_total": 0.0},
            "650570": {"cur_banco": 0.0, "cur_caja": 1215.0, "prev_total": 0.0},
            "758000": {"cur_banco": 46043.0, "cur_caja": 29300.0, "prev_total": 0.0},
            "758100": {"cur_banco": 0.0, "cur_caja": 297852.0, "prev_total": 0.0},
            "759000": {"cur_banco": 0.0, "cur_caja": 0.0, "prev_total": 0.0},
            "759100": {"cur_banco": 6000.0, "cur_caja": 0.0, "prev_total": 0.0},
        }
        cuentas_by_section = {
            "6": sorted(c for c in totales if c.startswith("6")),
            "7": sorted(c for c in totales if c.startswith("7")),
            "2": [],
        }
        presupuestos = {
            "602400": 35000.0,
            "628000": 60000.0,
            "628030": 60000.0,
            "650100": 65000.0,
            "650110": 7000.0,
            "650140": 58000.0,
            "650300": 20000.0,
            "650500": 240000.0,
            "758000": 35000.0,
        }

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "modelo.xlsx"
            exportar_modelo_evolutivo(
                out, totales, cuentas_by_section, 2026,
                nombre_por_cuenta={}, ruta_plantilla="ui/Modelo evolutivo presupuestario 26 - blank.xlsx",
                presupuesto_por_cuenta=presupuestos,
            )
            ws = openpyxl.load_workbook(out, data_only=True).active
            rows = {
                (str(ws.cell(r, 3).value), str(ws.cell(r, 4).value)): {
                    "corriente": ws.cell(r, 5).value,
                    "presupuesto": ws.cell(r, 10).value,
                    "diferencia": ws.cell(r, 11).value,
                }
                for r in range(1, ws.max_row + 1)
            }
            total_gastos_budget = next(
                ws.cell(r, 10).value
                for r in range(1, ws.max_row + 1)
                if ws.cell(r, 4).value == "SUMA TOTAL GASTOS"
            )

        self.assertEqual(rows[("602400", "C. de mat. de limp., lavandería, peluquería y aseo")]["corriente"], 7651)
        self.assertEqual(rows[("602400", "Material de limpieza")]["corriente"], 2202)
        self.assertEqual(rows[("602400", "Material de limpieza")]["presupuesto"], 35000)
        self.assertEqual(rows[("602400", "Material de limpieza")]["diferencia"], 32798)
        self.assertEqual(rows[("650300", "Atenciones comunitarias")]["corriente"], 2000)
        self.assertEqual(rows[("650500", "Formacion profesional")]["corriente"], 177800)
        self.assertEqual(rows[("758000", "Donativos")]["corriente"], 75343)
        self.assertEqual(rows[("759000", "Sueldos y salarios (Prestacion de servicios externos)")]["corriente"], 0)
        self.assertEqual(rows[("628000", "Suministros")]["presupuesto"], 60000)
        self.assertEqual(rows[("628030", "Suministro electricidad")]["presupuesto"], 60000)
        self.assertEqual(rows[("650100", "Asistencia sanitaria varia")]["presupuesto"], 65000)
        self.assertEqual(rows[("650110", "Odontología-consultas-prótesis")]["presupuesto"], 7000)
        self.assertEqual(rows[("650140", "Consultas, análisis, exploraciones")]["presupuesto"], 58000)
        self.assertEqual(total_gastos_budget, 420000)


if __name__ == "__main__":
    unittest.main()
