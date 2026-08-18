# -*- coding: utf-8 -*-
"""
pytest — RESUMEN DE TESORERÍA (InformesView, tipo 6)
====================================================
Calcula de forma INDEPENDIENTE los saldos acumulados por mes (Caja + bancos)
a partir de data/test_clean.json y verifica:
  1) la estructura y valores de la tabla que renderiza el informe,
  2) la fila TOTAL,
  3) la exportación a Excel.

Ejecutar desde la raíz del proyecto:
    python -m pytest tests/test_tesoreria_view.py -v
"""

import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")  # antes de importar Qt
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)  # la app usa rutas relativas "data/..."

import pytest

pytest.importorskip("PySide6")
from PySide6.QtWidgets import QApplication, QTableWidget, QTabWidget

import openpyxl

from models.ContabilidadData import ContabilidadData
from ui.InformesView import InformesView

MESES_CORTOS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
AÑO = 2025


# ============================================================
# FIXTURES
# ============================================================
@pytest.fixture(scope="module")
def qapp():
    app = QApplication.instance() or QApplication([])
    yield app


@pytest.fixture(scope="module")
def data():
    return ContabilidadData("data/test_clean.json")


@pytest.fixture(scope="module")
def view(qapp, data):
    """InformesView con el informe de tesorería generado para 2025."""
    v = InformesView(data)
    v.cbo_tipo.setCurrentIndex(6)                 # 💶 Resumen de Tesorería
    v.cbo_anio.setCurrentText(str(AÑO))
    v._generar()
    return v


# ============================================================
# CÁLCULO INDEPENDIENTE (no usa el código del informe)
# ============================================================
def _bancos_config():
    with open(ROOT / "data/bancos.json", encoding="utf-8") as f:
        return [b["nombre"] for b in json.load(f).get("banks", [])]


def _saldos_acumulados_esperados(movimientos, año):
    """
    saldo(m) = saldo_inicial + Σ (haber − debe) de movimientos PAGADOS
    con fecha ≤ fin del mes m del año indicado.

    Los meses POSTERIORES al último mes con movimientos reales quedan en None
    (celda vacía): el informe no arrastra el saldo a meses sin data.
    """
    # Saldo inicial real: primer mes del año con saldo registrado (por banco).
    saldos_iniciales = {}
    try:
        with open(ROOT / "data/saldos_mensuales.json", encoding="utf-8") as f:
            sd = json.load(f)
        saldos = sd.get("saldos", {})
        for mes in range(1, 13):
            mes_data = saldos.get(f"{año}-{mes:02d}")
            if not isinstance(mes_data, dict):
                continue
            for banco, valores in mes_data.items():
                if banco in saldos_iniciales or not isinstance(valores, dict):
                    continue
                saldos_iniciales[banco] = float(valores.get("inicial", 0.0) or 0.0)
    except (OSError, json.JSONDecodeError):
        pass

    # Último mes del año con movimientos reales (corte del informe)
    ultimo_mes = 0
    for m in movimientos:
        f_raw = str(m.get("fecha", ""))
        try:
            if "/" in f_raw:
                d, mm, a = map(int, f_raw.split("/"))
            elif "-" in f_raw:
                p = f_raw.split("-")
                if int(p[0]) > 1000:
                    a, mm, d = int(p[0]), int(p[1]), int(p[2])
                else:
                    d, mm, a = int(p[0]), int(p[1]), int(p[2])
            else:
                continue
        except (ValueError, IndexError):
            continue
        if a != año or not (1 <= mm <= 12):
            continue
        try:
            h = float(str(m.get("haber", 0)).replace(",", "."))
            dv = float(str(m.get("debe", 0)).replace(",", "."))
        except (ValueError, TypeError):
            continue
        if str(m.get("estado", "")).lower() != "pagado":
            continue
        if h - dv and mm > ultimo_mes:
            ultimo_mes = mm

    def _cortar(serie):
        if ultimo_mes <= 0:
            return [None] * 12
        return serie[:ultimo_mes] + [None] * (12 - ultimo_mes)

    # El saldo inicial es la base de los meses hasta el último con data
    acum = {b: _cortar([saldos_iniciales.get(b, 0.0)] * 12) for b in _bancos_config()}

    for m in movimientos:
        f_raw = str(m.get("fecha", ""))
        try:
            if "/" in f_raw:
                d, mm, a = map(int, f_raw.split("/"))
            elif "-" in f_raw:
                p = f_raw.split("-")
                if int(p[0]) > 1000:
                    a, mm, d = int(p[0]), int(p[1]), int(p[2])
                else:
                    d, mm, a = int(p[0]), int(p[1]), int(p[2])
            else:
                continue
        except (ValueError, IndexError):
            continue

        if a != año or not (1 <= mm <= 12):
            continue

        try:
            h = float(str(m.get("haber", 0)).replace(",", "."))
            dv = float(str(m.get("debe", 0)).replace(",", "."))
        except (ValueError, TypeError):
            continue

        if str(m.get("estado", "")).lower() != "pagado":
            continue

        neto = h - dv
        if not neto:
            continue

        banco = str(m.get("banco", "") or "Caja")
        if banco not in acum:
            acum[banco] = _cortar([0.0] * 12)
        for idx in range(mm - 1, ultimo_mes):
            if acum[banco][idx] is not None:
                acum[banco][idx] += neto

    return acum


def _obtener_tabla(view):
    for i in range(view.contenedor_layout.count()):
        w = view.contenedor_layout.itemAt(i).widget()
        if isinstance(w, QTableWidget):
            return w
    raise AssertionError("No se encontró la tabla del Resumen de Tesorería")


def _fila_por_nombre(tabla):
    return {tabla.item(r, 0).text(): r for r in range(tabla.rowCount())}


def _celda_float(tabla, r, c):
    """Devuelve el valor numérico de una celda, o None si está vacía."""
    it = tabla.item(r, c)
    if it is None or not it.text().strip():
        return None
    return float(it.text().replace(",", ""))


# ============================================================
# DATOS DE CONFIGURACIÓN
# ============================================================
def test_cuentas_y_bancos_nuevos():
    """Las cuentas 556/557 y las opciones Cambio Euros/Contrapartida existen."""
    with open(ROOT / "data/plan_contable_v3.json", encoding="utf-8") as f:
        plan = json.load(f)
    assert plan["556"]["nombre"] == "Cambio Euros"
    assert plan["557"]["nombre"] == "Contrapartida"
    bancos = _bancos_config()
    assert "Cambio Euros" in bancos
    assert "Contrapartida" in bancos


# ============================================================
# TABLA DEL INFORME
# ============================================================
def test_tabla_estructura(view):
    tabla = _obtener_tabla(view)

    assert tabla.columnCount() == 13
    assert tabla.horizontalHeaderItem(0).text() == "BANCO / TESORERÍA"
    for c, mes in enumerate(MESES_CORTOS, start=1):
        assert tabla.horizontalHeaderItem(c).text() == mes

    filas = _fila_por_nombre(tabla)
    assert filas["Caja"] == 0                       # Caja primero
    assert "Cambio Euros" in filas
    assert "Contrapartida" in filas
    assert filas["TOTAL"] == tabla.rowCount() - 1   # TOTAL al final


def test_tabla_valores_acumulados(view, data):
    """Cada celda de cada banco coincide con el cálculo independiente.
    Las celdas vacías (None = mes sin data real) deben coincidir también."""
    esperado = _saldos_acumulados_esperados(data.movimientos, AÑO)
    tabla = _obtener_tabla(view)

    for r in range(tabla.rowCount() - 1):           # excluye TOTAL
        nombre = tabla.item(r, 0).text()
        vals = esperado.get(nombre, [None] * 12)
        for c in range(12):
            celda = _celda_float(tabla, r, c + 1)
            if vals[c] is None:
                assert celda is None, (nombre, MESES_CORTOS[c], celda)
            else:
                assert celda == pytest.approx(round(vals[c], 2), abs=0.01), (
                    nombre, MESES_CORTOS[c], celda, vals[c])


def test_tabla_fila_total(view, data):
    """La fila TOTAL es la suma de los bancos en cada mes (None = vacío)."""
    esperado = _saldos_acumulados_esperados(data.movimientos, AÑO)
    tabla = _obtener_tabla(view)
    r_total = _fila_por_nombre(tabla)["TOTAL"]

    totales = [None] * 12
    for r in range(tabla.rowCount() - 1):
        vals = esperado.get(tabla.item(r, 0).text(), [None] * 12)
        for c in range(12):
            v = vals[c]
            if v is None:
                continue
            totales[c] = (totales[c] or 0.0) + v

    for c in range(12):
        celda = _celda_float(tabla, r_total, c + 1)
        if totales[c] is None:
            assert celda is None, MESES_CORTOS[c]
        else:
            assert celda == pytest.approx(round(totales[c], 2), abs=0.01), MESES_CORTOS[c]


def test_valores_concretos_noviembre_y_diciembre_vacio(view):
    """
    Último mes con data (test_clean.json solo tiene noviembre 2025):
    saldo inicial real de nov. 2025 (saldos_mensuales.json) + movimientos.
    Caja: 248655.01 − 72708.00 = 175947.01 · Union Bank: 56278.81 − 9237.90 = 47040.91
    TOTAL: 175947.01 + 47040.91 + 1572.30 + 5109.14 = 229669.36
    Diciembre queda VACÍO (sin movimientos reales en el año).
    """
    tabla = _obtener_tabla(view)
    filas = _fila_por_nombre(tabla)

    assert _celda_float(tabla, filas["Caja"], 11) == pytest.approx(175947.01, abs=0.01)
    assert _celda_float(tabla, filas["Union Bank"], 11) == pytest.approx(47040.91, abs=0.01)
    assert _celda_float(tabla, filas["Federal Bank"], 11) == pytest.approx(1572.30, abs=0.01)
    assert _celda_float(tabla, filas["TOTAL"], 11) == pytest.approx(229669.36, abs=0.01)

    # Meses posteriores al último con data → celdas vacías
    assert _celda_float(tabla, filas["Caja"], 12) is None
    assert _celda_float(tabla, filas["TOTAL"], 12) is None


# ============================================================
# AGRUPACIÓN POR CUENTA REAL (cuenta_banco)
# ============================================================
def test_agrupacion_por_cuenta_banco():
    """
    Movimientos de la MISMA cuenta real (cuenta_banco) se agrupan en UNA sola
    fila aunque el nombre del banco difiera (p.ej. 'Union Bank' vs
    'Union Bank, sr Elisa' = 5725). Los que no traen cuenta_banco se resuelven
    por la etiqueta ya vista con esa cuenta.
    """
    from models.tesoreria import datos_tesoreria
    movs = [
        {"fecha": "15/01/2026", "documento": "A", "concepto": "Interes", "cuenta": "769000",
         "debe": 0.0, "haber": 208.0, "estado": "pagado",
         "banco": "Union Bank", "cuenta_banco": "5725"},
        {"fecha": "16/01/2026", "documento": "B", "concepto": "Salario", "cuenta": "5725",
         "debe": 0.0, "haber": 10000.0, "estado": "pagado",
         "banco": "Union Bank, sr Elisa", "cuenta_banco": "5725"},
        # Sin cuenta_banco: se resuelve por la etiqueta ya mapeada a 5725
        {"fecha": "20/01/2026", "documento": "C", "concepto": "Otro", "cuenta": "5725",
         "debe": 0.0, "haber": 500.0, "estado": "pagado",
         "banco": "Union Bank, sr Elisa"},
    ]
    orden, acum = datos_tesoreria(movs, 2026)

    union = [b for b in orden if "Union" in b]
    assert len(union) == 1, f"El mismo banco aparece repetido: {orden}"
    assert acum[union[0]][0] == pytest.approx(10708.0, abs=0.01)  # 208 + 10000 + 500
    # Solo hay movimientos en enero: los meses posteriores quedan VACÍOS (None)
    assert acum[union[0]][1] is None
    assert len(acum[union[0]]) == 12


def test_no_filas_vacias_de_bancos_sin_datos():
    """Los bancos de configuración sin movimientos NI saldo inicial no aparecen
    como filas vacías (evita filas que confunden)."""
    from models.tesoreria import datos_tesoreria
    movs = [
        {"fecha": "10/01/2026", "documento": "A", "concepto": "Compra", "cuenta": "603000",
         "debe": 500.0, "haber": 0.0, "estado": "pagado",
         "banco": "Caja", "cuenta_banco": "570"},
    ]
    orden, acum = datos_tesoreria(movs, 2026)
    assert "Caja" in orden
    assert "Federal Bank" not in orden
    assert "SBI" not in orden
    assert "Otro" not in orden
    assert "Cambio Euros" in orden
    assert "Contrapartida" in orden
    assert orden[-1] not in ("Caja",)


# ============================================================
# PESTAÑA PROPIA EN EL HUB DE CIERRES
# ============================================================
def test_pestaña_propia_en_cierr_hub(qapp, data):
    """El hub de cierres tiene una pestaña 'Resumen de Tesorería' con los mismos valores."""
    from ui.CierresHub import CierresHub
    hub = CierresHub(data)

    tabs = hub.findChild(QTabWidget)
    nombres = [tabs.tabText(i) for i in range(tabs.count())]
    assert "Resumen de Tesorería" in nombres

    widget = tabs.widget(nombres.index("Resumen de Tesorería"))
    assert widget is hub.tab_tesoreria

    # Mismos acumulados que en Informes BI (saldo real de nov. 2025 + movimientos)
    widget.cbo_anio.setCurrentText("2025")
    filas = _fila_por_nombre(widget.tabla)
    assert _celda_float(widget.tabla, filas["Caja"], 11) == pytest.approx(175947.01, abs=0.01)
    assert _celda_float(widget.tabla, filas["Union Bank"], 11) == pytest.approx(47040.91, abs=0.01)
    assert _celda_float(widget.tabla, filas["TOTAL"], 11) == pytest.approx(229669.36, abs=0.01)
    # Diciembre vacío (misma regla de corte en la pestaña del hub)
    assert _celda_float(widget.tabla, filas["Caja"], 12) is None
    assert _celda_float(widget.tabla, filas["TOTAL"], 12) is None

    # La exportación usa la misma función compartida (cubierta en test_exportacion_excel)


# ============================================================
# EXPORTACIÓN A EXCEL
# ============================================================
def test_exportacion_excel(view, tmp_path):
    ruta = tmp_path / "tesoreria_2025.xlsx"
    view._exportar_excel_tesoreria(str(ruta))
    assert ruta.exists()

    wb = openpyxl.load_workbook(str(ruta))
    ws = wb["Tesorería"]

    assert ws.cell(1, 1).value == f"RESUMEN DE TESORERÍA {AÑO} — SALDO ACUMULADO POR MES"
    assert ws.cell(3, 1).value == "BANCO / TESORERÍA"
    assert ws.cell(3, 2).value == "Enero"
    assert ws.cell(3, 13).value == "Diciembre"

    # Fila Caja = fila 4 (1 título, 2 en blanco, 3 cabecera)
    assert ws.cell(4, 1).value == "Caja"
    # Noviembre (col 12) = último mes con data real; Diciembre (col 13) vacío
    assert ws.cell(4, 12).value == pytest.approx(175947.01, abs=0.01)
    assert ws.cell(4, 13).value is None

    # Última fila = TOTAL con el acumulado del último mes con data
    assert ws.cell(ws.max_row, 1).value == "TOTAL"
    assert ws.cell(ws.max_row, 12).value == pytest.approx(229669.36, abs=0.01)
    assert ws.cell(ws.max_row, 13).value is None
