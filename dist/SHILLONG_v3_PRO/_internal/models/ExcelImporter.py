# -*- coding: utf-8 -*-
"""
ExcelImporter.py — SHILLONG CONTABILIDAD v3.6 PRO
Motor lógico para leer e interpretar archivos Excel de movimientos.
"""

import openpyxl
from datetime import datetime

class ExcelImporter:
    """
    Clase encargada de leer un Excel, validar datos y convertirlos
    al formato de diccionario que usa ContabilidadData.
    """

    # Nombres de columnas esperados en el Excel (puedes añadir variaciones)
    COL_MAP = {
        "FECHA": ["fecha", "date", "día", "dia", "f. valor"],
        "DOCUMENTO": ["documento", "doc", "ref", "referencia", "cheque", "n°", "número", "numero"],
        "CONCEPTO": ["concepto", "descripción", "descripcion", "detalle", "transacción", "leyenda", "narration"],
        "CUENTA": ["cuenta", "cta", "rubro", "código", "codigo", "partida"],
        "DEBE": ["debe", "gasto", "débito", "cargo", "salida", "debit", "dr"],
        "HABER": ["haber", "ingreso", "crédito", "abono", "entrada", "credit", "cr"],
        "BANCO": ["banco", "caja", "tesorería", "origen"],
        "ESTADO": ["estado", "status", "situación"]
    }

    def importar(self, ruta_archivo):
        """
        Lee el archivo y devuelve una lista de diccionarios con los movimientos válidos
        y una lista de errores (si los hay).
        Retorna: (movimientos_validos, lista_errores)
        """
        try:
            wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"No se pudo abrir el archivo: {str(e)}"]

        headers = {}
        movimientos = []
        errores = []
        
        # 1. Detectar cabeceras buscando en las primeras 15 filas
        header_row_idx = -1
        
        # Iteramos filas para encontrar la que tiene "Fecha" y "Concepto"
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            temp_headers = {}
            row_values = [str(r).lower().strip() if r is not None else "" for r in row]
            
            # Revisar cada celda de esta fila
            for col_idx, val in enumerate(row_values):
                if not val: continue
                
                # Mapear nombre de columna a nuestra clave interna
                for key, variations in self.COL_MAP.items():
                    if val in variations:
                        temp_headers[key] = col_idx # Guardamos índice (0-based)
            
            # Criterio de aceptación: Debe tener al menos FECHA y CONCEPTO
            if "FECHA" in temp_headers and "CONCEPTO" in temp_headers:
                headers = temp_headers
                header_row_idx = i
                break
        
        if header_row_idx == -1:
            return [], ["No se encontraron las columnas 'Fecha' y 'Concepto' en las primeras 15 filas.\nAsegúrese de que el Excel tenga cabeceras claras."]

        # 2. Contexto de Cuenta (para reportes agrupados por cuenta)
        # Si encontramos una celda que diga "Cuenta: X", la usaremos para los siguientes registros si no tienen cuenta.
        context_account = None

        # 3. Iterar filas de datos (empezando justo después de la fila de headers)
        start_data_row = header_row_idx + 1
        for i, row in enumerate(ws.iter_rows(min_row=start_data_row, values_only=True), start=start_data_row):
            try:
                # --- DETECCIÓN DE CONTEXTO ---
                # Verificar si es una fila de cabecera de grupo (ej: "Cuenta: 60001")
                # Unimos toda la fila en texto para buscar patrones
                row_str = " ".join([str(c).strip() for c in row if c]).lower()
                
                if "cuenta:" in row_str or "rubro:" in row_str or "código:" in row_str:
                    # Intentar extraer el código o nombre, limpiando un poco
                    # Asumimos que lo que sigue es la cuenta
                    clean_acc = row_str.replace("cuenta:", "").replace("rubro:", "").replace("código:", "").strip().upper()
                    if len(clean_acc) > 3: # Validación mínima
                        context_account = clean_acc
                        continue # Es una fila de título, no de datos

                # --- FECHA ---
                raw_fecha = row[headers["FECHA"]]
                
                if not raw_fecha:
                    continue # Saltar filas vacías
                
                # REGLA ANTI-FOOTER
                str_val = str(raw_fecha).lower()
                keywords_fin = ["preparado", "nombre:", "cargo:", "aprobado", "firma", "total", "saldo final"]
                if any(k in str_val for k in keywords_fin):
                    continue 

                fecha_str = self._procesar_fecha(raw_fecha)
                if not fecha_str:
                    # Si falla la fecha, podría ser basura o título intermedio no capturado
                    continue

                # --- CONCEPTO ---
                concepto = str(row[headers["CONCEPTO"]] or "").strip()
                if not concepto:
                    errores.append(f"Fila {i}: Falta concepto")
                    continue

                # --- IMPORTES ---
                idx_debe = headers.get("DEBE")
                idx_haber = headers.get("HABER")
                
                debe = float(row[idx_debe]) if idx_debe is not None and isinstance(row[idx_debe], (int, float)) else 0.0
                haber = float(row[idx_haber]) if idx_haber is not None and isinstance(row[idx_haber], (int, float)) else 0.0

                if debe == 0 and haber == 0:
                    pass

                # --- CUENTA ---
                idx_cta = headers.get("CUENTA")
                cuenta = None
                
                if idx_cta is not None and row[idx_cta] is not None:
                    cuenta = str(row[idx_cta]).strip()
                    if cuenta.endswith(".0"): cuenta = cuenta[:-2]
                
                # Si no hay cuenta en la fila, usar la del contexto
                if not cuenta or cuenta == "":
                    cuenta = context_account if context_account else "PENDIENTE"

                # --- BANCO ---
                idx_banco = headers.get("BANCO")
                banco = str(row[idx_banco]).strip() if (idx_banco is not None and row[idx_banco]) else "Caja"

                # --- ESTADO ---
                idx_estado = headers.get("ESTADO")
                estado = str(row[idx_estado]).lower().strip() if (idx_estado is not None and row[idx_estado]) else "pagado"
                if estado not in ["pagado", "pendiente"]:
                    estado = "pagado"

                # --- DOCUMENTO ---
                idx_doc = headers.get("DOCUMENTO")
                doc = str(row[idx_doc]).strip() if (idx_doc is not None and row[idx_doc]) else f"IMP-{i}"

                # Construir el objeto movimiento
                mov = {
                    "fecha": fecha_str,
                    "documento": doc,
                    "concepto": concepto,
                    "cuenta": str(cuenta).upper(), # Normalizar a mayúsculas
                    "debe": debe,
                    "haber": haber,
                    "moneda": "INR",
                    "banco": banco,
                    "estado": estado,
                    "saldo": 0.0,
                }
                movimientos.append(mov)

            except Exception as e:
                errores.append(f"Fila {i}: Error procesando datos ({str(e)})")

        return movimientos, errores

    def _procesar_fecha(self, valor):
        """Convierte datetime de Excel o string a 'dd/mm/yyyy'"""
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y")
        elif isinstance(valor, str):
            # Intentar formatos comunes
            formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]
            for f in formatos:
                try:
                    dt = datetime.strptime(valor, f)
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    continue
        return None