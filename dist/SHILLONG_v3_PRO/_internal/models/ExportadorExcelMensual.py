# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

class ExportadorExcelMensual:
    """
    Clase para exportar reportes en Excel nativo (.xlsx) con estilos profesionales
    usando openpyxl.
    """

    # Colores corporativos
    COLOR_HEADER_BG = "2563eb"  # Azul brillante SHILLONG
    COLOR_HEADER_FG = "FFFFFF"  # Blanco
    COLOR_TOTAL_BG = "E0F2FE"   # Azul muy claro para totales
    COLOR_BORDER = "94A3B8"     # Gris para bordes

    # Colores Semáforo
    COLOR_BAD_BG = "FEE2E2"     # Rojo muy claro fondo
    COLOR_BAD_FG = "DC2626"     # Rojo texto
    COLOR_GOOD_BG = "DCFCE7"    # Verde muy claro fondo
    COLOR_GOOD_FG = "16A34A"    # Verde texto

    @staticmethod
    def _estilizar_hoja(ws):
        """Aplica estilos generales a la hoja"""
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

    @staticmethod
    def _crear_estilo_header():
        return {
            "font": Font(bold=True, color=ExportadorExcelMensual.COLOR_HEADER_FG, size=11),
            "fill": PatternFill("solid", fgColor=ExportadorExcelMensual.COLOR_HEADER_BG),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(bottom=Side(style="thin", color=ExportadorExcelMensual.COLOR_BORDER))
        }

    @staticmethod
    def _crear_estilo_moneda():
        return "_,_(* #,##0.00_);_,_(* (#,##0.00);_,_(* \"-\"??_);_(@_)"

    # --- MÉTODOS EXISTENTES (GENERAL, AGRUPADO, EVOLUTIVO) ---

    @staticmethod
    def exportar_general(filename, datos, periodo):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Movimientos {periodo}"
        ExportadorExcelMensual._estilizar_hoja(ws)

        ws["A1"] = f"REPORTE MENSUAL DE MOVIMIENTOS - {periodo}"
        ws["A1"].font = Font(size=16, bold=True, color="1E3A8A")
        ws.merge_cells("A1:G1")

        headers = ["Fecha", "Documento", "Concepto", "Cuenta", "Debe", "Haber", "Saldo", "Banco", "Categoría"]
        header_style = ExportadorExcelMensual._crear_estilo_header()

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_style["font"]
            cell.fill = header_style["fill"]
            cell.alignment = header_style["alignment"]
            cell.border = header_style["border"]

        row_num = 4
        for d in datos:
            ws.cell(row=row_num, column=1, value=d.get("fecha"))
            ws.cell(row=row_num, column=2, value=d.get("documento"))
            ws.cell(row=row_num, column=3, value=d.get("concepto"))
            ws.cell(row=row_num, column=4, value=d.get("cuenta"))
            
            c_debe = ws.cell(row=row_num, column=5, value=d.get("debe", 0))
            c_debe.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            
            c_haber = ws.cell(row=row_num, column=6, value=d.get("haber", 0))
            c_haber.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            
            c_saldo = ws.cell(row=row_num, column=7, value=d.get("saldo", 0))
            c_saldo.number_format = ExportadorExcelMensual._crear_estilo_moneda()

            ws.cell(row=row_num, column=8, value=d.get("banco"))
            ws.cell(row=row_num, column=9, value=d.get("categoria", ""))
            row_num += 1

        last_row = row_num
        ws.cell(row=last_row, column=4, value="TOTALES:").font = Font(bold=True)
        for col_let in ['E', 'F']:
            cell = ws[f"{col_let}{last_row}"]
            cell.value = f"=SUM({col_let}4:{col_let}{last_row-1})"
            cell.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=ExportadorExcelMensual.COLOR_TOTAL_BG)

        wb.save(filename)

    @staticmethod
    def exportar_agrupado(filename, grupos, periodo, tipo="Categoría"):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Por {tipo}"
        ExportadorExcelMensual._estilizar_hoja(ws)
        
        ws["A1"] = f"REPORTE AGRUPADO POR {tipo.upper()} - {periodo}"
        ws["A1"].font = Font(size=16, bold=True, color="1E3A8A")
        ws.merge_cells("A1:F1")

        row_num = 3
        for nombre_grupo, items in grupos.items():
            ws.merge_cells(f"A{row_num}:F{row_num}")
            cell = ws.cell(row=row_num, column=1, value=f"{tipo}: {nombre_grupo}")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="1E40AF")
            row_num += 1

            headers = ["Fecha", "Concepto", "Documento", "Debe", "Haber", "Saldo Línea"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=c, value=h)
                cell.font = Font(bold=True, size=9)
                cell.border = Border(bottom=Side(style="thin"))
                cell.fill = PatternFill("solid", fgColor="F1F5F9")
            row_num += 1

            total_debe = 0
            total_haber = 0
            for item in items:
                d = float(item.get("debe", 0))
                h = float(item.get("haber", 0))
                total_debe += d
                total_haber += h
                
                ws.cell(row=row_num, column=1, value=item.get("fecha"))
                ws.cell(row=row_num, column=2, value=item.get("concepto"))
                ws.cell(row=row_num, column=3, value=item.get("documento"))
                ws.cell(row=row_num, column=4, value=d).number_format = ExportadorExcelMensual._crear_estilo_moneda()
                ws.cell(row=row_num, column=5, value=h).number_format = ExportadorExcelMensual._crear_estilo_moneda()
                ws.cell(row=row_num, column=6, value=h-d).number_format = ExportadorExcelMensual._crear_estilo_moneda()
                row_num += 1

            ws.cell(row=row_num, column=3, value="SUBTOTAL:").font = Font(bold=True)
            ws.cell(row=row_num, column=4, value=total_debe).font = Font(bold=True, color="DC2626")
            ws.cell(row=row_num, column=5, value=total_haber).font = Font(bold=True, color="16A34A")
            ws.cell(row=row_num, column=6, value=total_haber-total_debe).font = Font(bold=True)
            row_num += 2

        wb.save(filename)

    @staticmethod
    def exportar_evolutivo_anual(filename, datos_matriz, año):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Evolutivo {año}"
        
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Cuenta", "Nombre", "Ene", "Feb", "Mar", "Abr", "May", "Jun", 
                   "Jul", "Ago", "Sep", "Oct", "Nov", "Dic", "TOTAL"]
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            if i > 2: ws.column_dimensions[get_column_letter(i)].width = 12

        row = 2
        for cuenta, valores in datos_matriz.items():
            nombre_cuenta, mensuales, total = valores
            ws.cell(row=row, column=1, value=cuenta)
            ws.cell(row=row, column=2, value=nombre_cuenta)
            for i, monto in enumerate(mensuales):
                c = ws.cell(row=row, column=i+3, value=monto)
                c.number_format = "#,##0.00"
                if monto == 0: c.font = Font(color="D1D5DB")
            
            c_total = ws.cell(row=row, column=15, value=total)
            c_total.font = Font(bold=True)
            c_total.number_format = "#,##0.00"
            c_total.fill = PatternFill("solid", fgColor="F3F4F6")
            
            row += 1
            
        wb.save(filename)

    # === NUEVO: CONTROL PRESUPUESTARIO (SEMÁFORO) ===
    @staticmethod
    def exportar_presupuesto_vs_real(filename, datos, año):
        """
        Genera el informe comparativo con semáforo.
        datos: Lista de tuplas/dicts [(cuenta, nombre, presupuestado, real, desviacion, porcentaje)...]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"Presupuesto {año}"
        
        # Anchos
        ws.column_dimensions['A'].width = 12 # Cuenta
        ws.column_dimensions['B'].width = 40 # Nombre
        ws.column_dimensions['C'].width = 18 # Presupuesto
        ws.column_dimensions['D'].width = 18 # Real
        ws.column_dimensions['E'].width = 18 # Desviación
        ws.column_dimensions['F'].width = 12 # %

        # Título
        ws["A1"] = f"CONTROL PRESUPUESTARIO {año}"
        ws["A1"].font = Font(size=16, bold=True, color="1E3A8A")
        ws.merge_cells("A1:F1")

        # Cabeceras
        headers = ["Cuenta", "Nombre Cuenta", "Presupuesto", "Realizado", "Desviación", "%"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E40AF")
            cell.alignment = Alignment(horizontal="center")

        row = 4
        for item in datos:
            # item = {cuenta, nombre, ppt, real, diff, pct}
            ws.cell(row=row, column=1, value=item["cuenta"])
            ws.cell(row=row, column=2, value=item["nombre"])
            
            c_ppt = ws.cell(row=row, column=3, value=item["ppt"])
            c_ppt.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            
            c_real = ws.cell(row=row, column=4, value=item["real"])
            c_real.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            c_real.font = Font(bold=True)

            # Desviación (Semáforo)
            diff = item["diff"]
            c_diff = ws.cell(row=row, column=5, value=diff)
            c_diff.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            
            # Lógica Semáforo:
            # Si gasté de MÁS (diff negativo en contabilidad de gastos, o Real > Ppt) -> ROJO
            # Aquí asumimos que ambos son positivos. Si Real > Ppt -> Malo.
            if item["real"] > item["ppt"]:
                c_diff.font = Font(color=ExportadorExcelMensual.COLOR_BAD_FG, bold=True)
                c_diff.fill = PatternFill("solid", fgColor=ExportadorExcelMensual.COLOR_BAD_BG)
            elif item["real"] < item["ppt"] and item["ppt"] > 0:
                c_diff.font = Font(color=ExportadorExcelMensual.COLOR_GOOD_FG, bold=True)
                c_diff.fill = PatternFill("solid", fgColor=ExportadorExcelMensual.COLOR_GOOD_BG)

            c_pct = ws.cell(row=row, column=6, value=item["pct"]/100 if item["ppt"] != 0 else 0)
            c_pct.number_format = "0.0%"
            
            row += 1

        wb.save(filename)

    # === NUEVO: RANKING TOP 10 (PARETO) ===
    @staticmethod
    def exportar_ranking(filename, datos, año, total_anual):
        """
        Genera el Top de gastos.
        datos: Lista de tuplas (ranking, cuenta, nombre, importe, porcentaje_del_total)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"Top Gastos {año}"
        
        ws.column_dimensions['A'].width = 8   # Rank
        ws.column_dimensions['B'].width = 12  # Cuenta
        ws.column_dimensions['C'].width = 40  # Nombre
        ws.column_dimensions['D'].width = 20  # Importe
        ws.column_dimensions['E'].width = 15  # % Total
        ws.column_dimensions['F'].width = 15  # % Acumulado

        ws["A1"] = f"TOP GASTOS (PARETO) - AÑO {año}"
        ws["A1"].font = Font(size=16, bold=True, color="D97706") # Color ámbar/oro
        ws.merge_cells("A1:F1")

        headers = ["Ranking", "Cuenta", "Nombre Cuenta", "Importe Anual", "% del Gasto Total", "% Acumulado"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B45309") # Naranja oscuro
            cell.alignment = Alignment(horizontal="center")

        row = 4
        acumulado = 0
        for item in datos:
            # item = {rank, cuenta, nombre, importe, pct}
            rank = item["rank"]
            importe = item["importe"]
            pct = importe / total_anual if total_anual > 0 else 0
            acumulado += pct

            # Columna Ranking
            c_rank = ws.cell(row=row, column=1, value=f"#{rank}")
            c_rank.alignment = Alignment(horizontal="center")
            c_rank.font = Font(bold=True)
            if rank == 1: 
                c_rank.fill = PatternFill("solid", fgColor="FFD700") # Oro
            elif rank == 2:
                c_rank.fill = PatternFill("solid", fgColor="C0C0C0") # Plata
            elif rank == 3:
                c_rank.fill = PatternFill("solid", fgColor="CD7F32") # Bronce

            ws.cell(row=row, column=2, value=item["cuenta"])
            ws.cell(row=row, column=3, value=item["nombre"])
            
            c_imp = ws.cell(row=row, column=4, value=importe)
            c_imp.number_format = ExportadorExcelMensual._crear_estilo_moneda()
            c_imp.font = Font(bold=True) if rank <= 3 else Font(bold=False)

            ws.cell(row=row, column=5, value=pct).number_format = "0.00%"
            ws.cell(row=row, column=6, value=acumulado).number_format = "0.00%"

            row += 1

        # Pie de página con el total analizado
        ws.cell(row=row+1, column=3, value="TOTAL GASTOS AÑO:").font = Font(bold=True)
        ws.cell(row=row+1, column=4, value=total_anual).number_format = ExportadorExcelMensual._crear_estilo_moneda()
        
        wb.save(filename)