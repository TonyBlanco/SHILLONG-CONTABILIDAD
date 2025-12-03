# -*- coding: utf-8 -*-
"""
INFORMESVIEW v4.3 — SHILLONG Contabilidad 3.7.7 PRO BI
-------------------------------------------------------
INCLUYE:
- Diario general
- Libro mayor agrupado (vista + Excel profesional)
- Balance de Sumas y Saldos (vista + Excel profesional)
- Resumen mensual por cuentas
BOTONES:
- Generar Informe
- Exportar Vista Actual
- Exportar Libro Mayor Profesional
- Exportar Balance Profesional
COLORES SHILLONG:
- Morado Encabezado: #7030A0
- Verde Total: #E2EFDA
"""

from PySide6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QPushButton, QComboBox,
    QTableWidget, QTableWidgetItem, QFileDialog, QDateEdit, QScrollArea
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont

import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class InformesView(QWidget):

    def __init__(self, data):
        super().__init__()
        self.data = data
        self._build_ui()

    # ================================================================
    # UI PRINCIPAL
    # ================================================================
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(12)

        # TÍTULO
        titulo = QLabel("📊 Informes Contables")
        titulo.setFont(QFont("Segoe UI", 22, QFont.Bold))
        titulo.setStyleSheet("color: #334155;")
        layout.addWidget(titulo)

        # ------------------------------------------------------------
        # Selector tipo informe
        # ------------------------------------------------------------
        h_sel = QHBoxLayout()
        h_sel.addWidget(QLabel("Tipo de Informe:"))

        self.cbo_tipo = QComboBox()
        self.cbo_tipo.addItems([
            "📘 Diario General (Rango)",
            "📒 Libro Mayor (Por Cuenta)",
            "⚖️ Balance de Sumas y Saldos",
            "🧮 Resumen Mensual por Cuentas"
        ])
        self.cbo_tipo.currentIndexChanged.connect(self._cambiar_tipo)
        h_sel.addWidget(self.cbo_tipo)
        h_sel.addStretch()
        layout.addLayout(h_sel)

        # ------------------------------------------------------------
        # Filtros dinámicos
        # ------------------------------------------------------------
        self.filtros = QHBoxLayout()
        layout.addLayout(self.filtros)

        # ------------------------------------------------------------
        # BOTONES SUPERIORES
        # ------------------------------------------------------------
        self.btn_generar = QPushButton("Generar Informe")
        self.btn_generar.setStyleSheet(
            "background:#2563eb; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_generar.clicked.connect(self._generar)

        self.btn_export_vista = QPushButton("Exportar Vista a Excel")
        self.btn_export_vista.setStyleSheet(
            "background:#475569; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_vista.clicked.connect(self._exportar_excel_vista)

        self.btn_export_mayor = QPushButton("📘 Exportar Libro Mayor Profesional (Excel)")
        self.btn_export_mayor.setStyleSheet(
            "background:#7030A0; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_mayor.clicked.connect(self._exportar_libro_mayor)

        self.btn_export_balance = QPushButton("📊 Exportar Balance Profesional (Excel)")
        self.btn_export_balance.setStyleSheet(
            "background:#7030A0; color:white; padding:6px 20px; font-weight:bold;"
        )
        self.btn_export_balance.clicked.connect(self._exportar_balance)

        h_btns = QHBoxLayout()
        h_btns.addWidget(self.btn_generar)
        h_btns.addWidget(self.btn_export_vista)
        h_btns.addWidget(self.btn_export_mayor)
        h_btns.addWidget(self.btn_export_balance)
        h_btns.addStretch()
        layout.addLayout(h_btns)

        # ------------------------------------------------------------
        # SCROLL PARA LAS TABLAS
        # ------------------------------------------------------------
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        self.contenedor = QWidget()
        self.contenedor_layout = QVBoxLayout(self.contenedor)

        self.scroll_area.setWidget(self.contenedor)
        layout.addWidget(self.scroll_area)

        self._cambiar_tipo()

    # ================================================================
    # CAMBIO DE TIPO DE INFORME
    # ================================================================
    def _cambiar_tipo(self):
        tipo = self.cbo_tipo.currentIndex()
        self._limpiar_filtros()

        # recrear widgets
        self.fecha_ini = QDateEdit()
        self.fecha_ini.setCalendarPopup(True)
        self.fecha_ini.setDate(QDate.currentDate().addMonths(-1))

        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())

        self.cbo_cuenta = QComboBox()
        self.cbo_cuenta.addItem("Todas")
        for cta, d in self.data.cuentas.items():
            self.cbo_cuenta.addItem(f"{cta} — {d.get('nombre','')}")

        self.cbo_mes = QComboBox()
        self.cbo_mes.addItems([
            "Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
        ])
        self.cbo_mes.setCurrentIndex(datetime.date.today().month - 1)

        self.cbo_anio = QComboBox()
        for a in range(2020, 2036):
            self.cbo_anio.addItem(str(a))
        self.cbo_anio.setCurrentText(str(datetime.date.today().year))

        # Mostrar filtros
        if tipo == 0:  # DIARIO
            self.filtros.addWidget(QLabel("Fecha inicio:"))
            self.filtros.addWidget(self.fecha_ini)
            self.filtros.addWidget(QLabel("Fecha fin:"))
            self.filtros.addWidget(self.fecha_fin)

        elif tipo == 1:
            self.filtros.addWidget(QLabel("Cuenta:"))
            self.filtros.addWidget(self.cbo_cuenta)

        elif tipo == 2:
            lab = QLabel("Balance profesional SHILLONG agrupado por cuentas.")
            lab.setStyleSheet("color:#475569; font-style:italic;")
            self.filtros.addWidget(lab)

        elif tipo == 3:
            self.filtros.addWidget(QLabel("Mes:"))
            self.filtros.addWidget(self.cbo_mes)
            self.filtros.addWidget(QLabel("Año:"))
            self.filtros.addWidget(self.cbo_anio)

        # visibilidad de botones
        self.btn_export_mayor.setVisible(tipo == 1)
        self.btn_export_balance.setVisible(tipo == 2)

    def _limpiar_filtros(self):
        while self.filtros.count():
            item = self.filtros.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    # ================================================================
    # GENERAR INFORME
    # ================================================================
    def _generar(self):
        tipo = self.cbo_tipo.currentIndex()
        self._limpiar_vista()

        if tipo == 0:
            self._mostrar_diario()
        elif tipo == 1:
            self._mostrar_libro_mayor_agrupado()
        elif tipo == 2:
            self._mostrar_sumas_saldos()
        elif tipo == 3:
            self._mostrar_resumen_mensual()

    def _limpiar_vista(self):
        for i in reversed(range(self.contenedor_layout.count())):
            item = self.contenedor_layout.takeAt(i)
            if item.widget():
                item.widget().deleteLater()

    # ================================================================
    # DIARIO GENERAL
    # ================================================================
    def _mostrar_diario(self):
        ini = self.fecha_ini.date().toPython()
        fin = self.fecha_fin.date().toPython()
        datos = self.data.get_movimientos_rango(ini, fin)
        tabla = QTableWidget(0, 8)
        tabla.setHorizontalHeaderLabels(
            ["Fecha","Documento","Concepto","Cuenta","Debe","Haber","Saldo","Banco"]
        )

        for m in datos:
            r = tabla.rowCount()
            tabla.insertRow(r)
            fila = [
                m.get("fecha",""),
                m.get("documento",""),
                m.get("concepto",""),
                m.get("cuenta",""),
                m.get("debe",0),
                m.get("haber",0),
                m.get("saldo",0),
                m.get("banco","")
            ]
            for c, val in enumerate(fila):
                it = QTableWidgetItem(str(val))
                if c >= 4:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla.setItem(r,c,it)

        self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # LIBRO MAYOR AGRUPADO
    # ================================================================
    def _mostrar_libro_mayor_agrupado(self):

        texto = self.cbo_cuenta.currentText()
        if texto == "Todas":
            cuentas = sorted(self.data.cuentas.keys())
        else:
            cuentas = [texto.split(" — ")[0]]

        for cta in cuentas:
            movs = self.data.movimientos_por_cuenta(cta)
            if not movs:
                continue

            nombre = self.data.cuentas[cta].get("nombre","")

            header = QLabel(f"{cta} — {nombre}")
            header.setStyleSheet(
                "background:#7030A0; color:white; font-size:16px;"
                "padding:6px; font-weight:bold;"
            )
            self.contenedor_layout.addWidget(header)

            tabla = QTableWidget(0, 6)
            tabla.setHorizontalHeaderLabels(
                ["Fecha","Documento","Desglose","Debe","Haber","Saldo"]
            )

            total_debe=0
            total_haber=0
            saldo_acum=0

            for m in movs:
                r = tabla.rowCount()
                tabla.insertRow(r)

                concepto = m.get("concepto","").strip()
                if concepto:
                    des = concepto
                else:
                    des = m.get("nombre_cuenta","")

                debe = float(m.get("debe",0))
                haber = float(m.get("haber",0))

                saldo_acum += (haber - debe)
                total_debe += debe
                total_haber += haber

                fila = [
                    m.get("fecha",""),
                    m.get("documento",""),
                    des,
                    debe,
                    haber,
                    saldo_acum
                ]

                for c,val in enumerate(fila):
                    it = QTableWidgetItem(str(val))
                    if c>=3:
                        it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                    tabla.setItem(r,c,it)

            r = tabla.rowCount()
            tabla.insertRow(r)
            tabla.setItem(r,2, QTableWidgetItem("TOTAL"))
            tabla.setItem(r,3, QTableWidgetItem(str(total_debe)))
            tabla.setItem(r,4, QTableWidgetItem(str(total_haber)))
            tabla.setItem(r,5, QTableWidgetItem(str(total_haber-total_debe)))

            self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # SUMAS & SALDOS — VISTA SHILLONG
    # ================================================================
    def _mostrar_sumas_saldos(self):

        tabla = QTableWidget(0,5)
        tabla.setHorizontalHeaderLabels(
            ["Cuenta","Nombre","Debe","Haber","Saldo"]
        )

        resumen = defaultdict(lambda: {"nombre":"", "debe":0, "haber":0})

        for m in self.data.movimientos:
            cta = str(m.get("cuenta",""))
            resumen[cta]["nombre"] = m.get("nombre_cuenta","")
            resumen[cta]["debe"] += float(m.get("debe",0))
            resumen[cta]["haber"] += float(m.get("haber",0))

        total_debe=0
        total_haber=0

        for cta in sorted(resumen.keys()):
            d = resumen[cta]
            saldo = d["haber"]-d["debe"]

            total_debe += d["debe"]
            total_haber+= d["haber"]

            r = tabla.rowCount()
            tabla.insertRow(r)
            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila):
                it = QTableWidgetItem(str(val))
                if c>=2:
                    it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                tabla.setItem(r,c,it)

        r = tabla.rowCount()
        tabla.insertRow(r)
        tabla.setItem(r,1, QTableWidgetItem("TOTAL GENERAL"))
        tabla.setItem(r,2, QTableWidgetItem(str(total_debe)))
        tabla.setItem(r,3, QTableWidgetItem(str(total_haber)))
        tabla.setItem(r,4, QTableWidgetItem(str(total_haber-total_debe)))

        self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # RESUMEN MENSUAL
    # ================================================================
    def _mostrar_resumen_mensual(self):

        mes = self.cbo_mes.currentIndex()+1
        anio = int(self.cbo_anio.currentText())

        movs = self.data.movimientos_por_mes(mes, anio)

        resumen = defaultdict(lambda: {"nombre":"", "debe":0, "haber":0})

        for m in movs:
            cta=str(m.get("cuenta",""))
            resumen[cta]["nombre"]=m.get("nombre_cuenta","")
            resumen[cta]["debe"]+=float(m.get("debe",0))
            resumen[cta]["haber"]+=float(m.get("haber",0))

        tabla = QTableWidget(0,5)
        tabla.setHorizontalHeaderLabels(
            ["Cuenta","Nombre","Debe","Haber","Saldo"]
        )

        for cta in sorted(resumen.keys()):
            d=resumen[cta]
            saldo=d["haber"]-d["debe"]

            r=tabla.rowCount()
            tabla.insertRow(r)
            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila):
                it=QTableWidgetItem(str(val))
                if c>=2:
                    it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                tabla.setItem(r,c,it)

        self.contenedor_layout.addWidget(tabla)

    # ================================================================
    # EXPORTAR VISTA ACTUAL
    # ================================================================
    def _exportar_excel_vista(self):

        ruta,_ = QFileDialog.getSaveFileName(
            self, "Exportar Vista", "Vista.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb=openpyxl.Workbook()
        ws=wb.active

        row=1

        # Vuelca todas las tablas y encabezados de la vista actual
        for i in range(self.contenedor_layout.count()):
            w = self.contenedor_layout.itemAt(i).widget()

            if isinstance(w, QLabel):
                ws.cell(row=row,column=1,value=w.text()).font=Font(bold=True)
                row+=2

            if isinstance(w, QTableWidget):
                tabla = w

                # Encabezados
                for c in range(tabla.columnCount()):
                    ws.cell(row=row,column=c+1,
                        value=tabla.horizontalHeaderItem(c).text())
                row+=1

                # Datos
                for r2 in range(tabla.rowCount()):
                    for c2 in range(tabla.columnCount()):
                        it=tabla.item(r2,c2)
                        ws.cell(row=row,column=c2+1,
                            value=it.text() if it else "")
                    row+=1

                row+=2

        wb.save(ruta)

    # ================================================================
    # EXPORTAR LIBRO MAYOR PROFESIONAL SHILLONG
    # ================================================================
    def _exportar_libro_mayor(self):

        ruta,_= QFileDialog.getSaveFileName(
            self,"Exportar Libro Mayor SHILLONG",
            "LibroMayor.xlsx","Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb=openpyxl.Workbook()
        ws=wb.active

        morado=PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        borde=Border(
            left=Side(style="thin",color="000000"),
            right=Side(style="thin",color="000000"),
            top=Side(style="thin",color="000000"),
            bottom=Side(style="thin",color="000000")
        )

        row=1

        cuentas=sorted(self.data.cuentas.keys())

        for cta in cuentas:
            movs=self.data.movimientos_por_cuenta(cta)
            if not movs:
                continue

            nombre=self.data.cuentas[cta].get("nombre","")

            # ENCABEZADO CUENTA
            cell=ws.cell(row=row,column=1,value=f"{cta} — {nombre}")
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=morado
            row+=2

            headers=["Fecha","Documento","Desglose","Debe","Haber","Saldo"]
            for c,h in enumerate(headers, start=1):
                cell=ws.cell(row=row,column=c,value=h)
                cell.font=Font(bold=True,color="FFFFFF")
                cell.fill=morado
                cell.border=borde
            row+=1

            saldo_acum=0
            total_debe=0
            total_haber=0

            for m in movs:
                concepto=m.get("concepto","").strip()
                if concepto:
                    des=concepto
                else:
                    des=m.get("nombre_cuenta","")

                debe=float(m.get("debe",0))
                haber=float(m.get("haber",0))

                saldo_acum+=(haber-debe)
                total_debe+=debe
                total_haber+=haber

                fila=[
                    m.get("fecha",""),
                    m.get("documento",""),
                    des,
                    debe,
                    haber,
                    saldo_acum
                ]

                for c,val in enumerate(fila,start=1):
                    cell=ws.cell(row=row,column=c,value=val)
                    cell.border=borde
                    if c>=4:
                        cell.alignment=Alignment(horizontal="right")
                row+=1

            # TOTAL
            for c in range(1,7):
                cell=ws.cell(row=row,column=c)
                cell.border=borde
                cell.fill=verde

            ws.cell(row=row,column=3,value="TOTAL").font=Font(bold=True)
            ws.cell(row=row,column=4,value=total_debe)
            ws.cell(row=row,column=5,value=total_haber)
            ws.cell(row=row,column=6,value=(total_haber-total_debe))

            row+=3

        wb.save(ruta)

    # ================================================================
    # EXPORTAR BALANCE SHILLONG (SUMAS & SALDOS)
    # ================================================================
    def _exportar_balance(self):

        ruta,_= QFileDialog.getSaveFileName(
            self,"Exportar Balance SHILLONG",
            "Balance_Sumas_Saldos.xlsx","Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb=openpyxl.Workbook()
        ws=wb.active

        morado=PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        verde=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        borde=Border(
            left=Side(style="thin",color="000000"),
            right=Side(style="thin",color="000000"),
            top=Side(style="thin",color="000000"),
            bottom=Side(style="thin",color="000000")
        )

        headers=["Cuenta","Nombre","Debe","Haber","Saldo"]

        row=1
        for c,h in enumerate(headers,start=1):
            cell=ws.cell(row=row,column=c,value=h)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=morado
            cell.border=borde
        row+=1

        resumen=defaultdict(lambda:{"nombre":"", "debe":0, "haber":0})

        for m in self.data.movimientos:
            cta=str(m.get("cuenta",""))
            resumen[cta]["nombre"]=m.get("nombre_cuenta","")
            resumen[cta]["debe"]+=float(m.get("debe",0))
            resumen[cta]["haber"]+=float(m.get("haber",0))

        total_debe=0
        total_haber=0

        for cta in sorted(resumen.keys()):
            d=resumen[cta]
            saldo=d["haber"]-d["debe"]

            fila=[cta, d["nombre"], d["debe"], d["haber"], saldo]

            for c,val in enumerate(fila,start=1):
                cell=ws.cell(row=row,column=c,value=val)
                cell.border=borde
                if c>=3:
                    cell.alignment=Alignment(horizontal="right")

            total_debe+=d["debe"]
            total_haber+=d["haber"]
            row+=1

        # TOTAL GENERAL
        for c in range(1,6):
            cell=ws.cell(row=row,column=c)
            cell.border=borde
            cell.fill=verde

        ws.cell(row=row,column=2,value="TOTAL GENERAL").font=Font(bold=True)
        ws.cell(row=row,column=3,value=total_debe)
        ws.cell(row=row,column=4,value=total_haber)
        ws.cell(row=row,column=5,value=(total_haber-total_debe))

        wb.save(ruta)
